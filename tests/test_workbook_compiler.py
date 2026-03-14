from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from Thesis_ML.config.schema_versions import (
    COMPILED_MANIFEST_SCHEMA_VERSION,
    WORKBOOK_SCHEMA_VERSION,
)
from Thesis_ML.orchestration.workbook_compiler import compile_workbook_file
from Thesis_ML.workbook.builder import build_workbook


def _make_workbook(path: Path) -> None:
    workbook = build_workbook()
    workbook.save(path)


def _set_executable_row(path: Path) -> None:
    workbook = load_workbook(path)
    ws = workbook["Experiment_Definitions"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col = {str(name): idx + 1 for idx, name in enumerate(headers)}

    ws.cell(2, col["experiment_id"], "E16")
    ws.cell(2, col["enabled"], "Yes")
    ws.cell(2, col["start_section"], "dataset_selection")
    ws.cell(2, col["end_section"], "evaluation")
    ws.cell(2, col["base_artifact_id"], "")
    ws.cell(2, col["target"], "coarse_affect")
    ws.cell(2, col["cv"], "within_subject_loso_session")
    ws.cell(2, col["model"], "ridge")
    ws.cell(2, col["subject"], "sub-001")
    ws.cell(2, col["train_subject"], "")
    ws.cell(2, col["test_subject"], "")
    ws.cell(2, col["filter_task"], "")
    ws.cell(2, col["filter_modality"], "")
    ws.cell(2, col["reuse_policy"], "auto")
    workbook.save(path)


def _header_map(ws, header_row: int) -> dict[str, int]:
    return {
        str(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(header_row, col).value is not None
    }


def _set_factorial_design(
    path: Path,
    *,
    add_constraint: bool = False,
    study_type: str = "full_factorial",
    intent: str = "exploratory",
) -> None:
    workbook = load_workbook(path)

    study_ws = workbook["Study_Design"]
    study_col = _header_map(study_ws, 2)
    study_ws.cell(3, study_col["study_id"], "S01")
    study_ws.cell(3, study_col["study_name"], "Factorial model/task study")
    study_ws.cell(3, study_col["enabled"], "Yes")
    study_ws.cell(3, study_col["study_type"], study_type)
    study_ws.cell(3, study_col["intent"], intent)
    study_ws.cell(3, study_col["start_section"], "dataset_selection")
    study_ws.cell(3, study_col["end_section"], "evaluation")
    study_ws.cell(3, study_col["primary_metric"], "balanced_accuracy")
    study_ws.cell(3, study_col["cv_scheme"], "within_subject_loso_session")
    study_ws.cell(3, study_col["replication_mode"], "fixed_repeats")
    study_ws.cell(3, study_col["num_repeats"], 1)
    study_ws.cell(3, study_col["random_seed_policy"], "fixed")

    factors_ws = workbook["Factors"]
    factors_col = _header_map(factors_ws, 2)
    factors_ws.cell(3, factors_col["study_id"], "S01")
    factors_ws.cell(3, factors_col["factor_name"], "model")
    factors_ws.cell(3, factors_col["parameter_path"], "model")
    factors_ws.cell(3, factors_col["factor_type"], "categorical")
    factors_ws.cell(3, factors_col["levels"], "ridge|logreg")

    factors_ws.cell(4, factors_col["study_id"], "S01")
    factors_ws.cell(4, factors_col["factor_name"], "filter_task")
    factors_ws.cell(4, factors_col["parameter_path"], "filter_task")
    factors_ws.cell(4, factors_col["factor_type"], "categorical")
    factors_ws.cell(4, factors_col["levels"], "passive|emo")

    fixed_ws = workbook["Fixed_Controls"]
    fixed_col = _header_map(fixed_ws, 2)
    fixed_ws.cell(3, fixed_col["study_id"], "S01")
    fixed_ws.cell(3, fixed_col["parameter_path"], "target")
    fixed_ws.cell(3, fixed_col["value"], "coarse_affect")
    fixed_ws.cell(4, fixed_col["study_id"], "S01")
    fixed_ws.cell(4, fixed_col["parameter_path"], "cv")
    fixed_ws.cell(4, fixed_col["value"], "within_subject_loso_session")
    fixed_ws.cell(5, fixed_col["study_id"], "S01")
    fixed_ws.cell(5, fixed_col["parameter_path"], "subject")
    fixed_ws.cell(5, fixed_col["value"], "sub-001")

    blocking_ws = workbook["Blocking_and_Replication"]
    blocking_col = _header_map(blocking_ws, 2)
    blocking_ws.cell(3, blocking_col["study_id"], "S01")
    blocking_ws.cell(3, blocking_col["block_type"], "none")
    blocking_ws.cell(3, blocking_col["repeat_id"], 1)

    if add_constraint:
        constraints_ws = workbook["Constraints"]
        constraints_col = _header_map(constraints_ws, 2)
        constraints_ws.cell(3, constraints_col["study_id"], "S01")
        constraints_ws.cell(3, constraints_col["if_factor"], "model")
        constraints_ws.cell(3, constraints_col["if_level"], "logreg")
        constraints_ws.cell(3, constraints_col["disallow_factor"], "filter_task")
        constraints_ws.cell(3, constraints_col["disallow_level"], "emo")
        constraints_ws.cell(3, constraints_col["reason"], "example exclusion")

    workbook.save(path)


def _set_rigor_metadata(
    path: Path,
    *,
    study_id: str = "S01",
    confirmatory_lock_applied: str = "No",
    primary_contrast: str = "ridge - logreg",
    multiplicity_handling: str = "none",
    interpretation_rules: str = "Descriptive, non-causal interpretation only.",
) -> None:
    workbook = load_workbook(path)

    checklist_ws = workbook["Study_Rigor_Checklist"]
    checklist_col = _header_map(checklist_ws, 2)
    checklist_ws.cell(3, checklist_col["study_id"], study_id)
    checklist_ws.cell(3, checklist_col["leakage_risk_reviewed"], "Yes")
    checklist_ws.cell(3, checklist_col["deployment_boundary_defined"], "Yes")
    checklist_ws.cell(3, checklist_col["unit_of_analysis_defined"], "Yes")
    checklist_ws.cell(3, checklist_col["data_hierarchy_defined"], "Yes")
    checklist_ws.cell(3, checklist_col["missing_data_plan"], "Median imputation with flag.")
    checklist_ws.cell(3, checklist_col["class_imbalance_plan"], "Class weighting enabled.")
    checklist_ws.cell(3, checklist_col["subgroup_plan"], "Report by task and modality.")
    checklist_ws.cell(3, checklist_col["reporting_checklist_completed"], "Yes")
    checklist_ws.cell(3, checklist_col["risk_of_bias_reviewed"], "Yes")
    checklist_ws.cell(3, checklist_col["confirmatory_lock_applied"], confirmatory_lock_applied)

    analysis_ws = workbook["Analysis_Plan"]
    analysis_col = _header_map(analysis_ws, 2)
    analysis_ws.cell(3, analysis_col["study_id"], study_id)
    analysis_ws.cell(3, analysis_col["primary_contrast"], primary_contrast)
    analysis_ws.cell(3, analysis_col["secondary_contrasts"], "task=emo - task=passive")
    analysis_ws.cell(3, analysis_col["aggregation_level"], "cell")
    analysis_ws.cell(3, analysis_col["uncertainty_method"], "bootstrap")
    analysis_ws.cell(3, analysis_col["multiplicity_handling"], multiplicity_handling)
    analysis_ws.cell(3, analysis_col["interaction_reporting_policy"], "descriptive_only")
    analysis_ws.cell(3, analysis_col["interpretation_rules"], interpretation_rules)

    workbook.save(path)


def test_compile_workbook_file_success(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_executable_row(workbook_path)

    manifest = compile_workbook_file(workbook_path)

    assert manifest.schema_version == WORKBOOK_SCHEMA_VERSION
    assert manifest.compiled_manifest_schema_version == COMPILED_MANIFEST_SCHEMA_VERSION
    assert len(manifest.experiments) == 1
    assert manifest.experiments[0].experiment_id == "E16"
    assert len(manifest.trial_specs) == 1
    trial = manifest.trial_specs[0]
    assert trial.start_section == "dataset_selection"
    assert trial.end_section == "evaluation"
    assert trial.params["target"] == "coarse_affect"
    assert trial.params["cv"] == "within_subject_loso_session"
    assert trial.params["model"] == "ridge"


def test_build_workbook_includes_factorial_design_sheets() -> None:
    workbook = build_workbook()
    for sheet in [
        "Study_Design",
        "Study_Rigor_Checklist",
        "Analysis_Plan",
        "Factors",
        "Fixed_Controls",
        "Constraints",
        "Blocking_and_Replication",
        "Generated_Design_Matrix",
        "Effect_Summaries",
    ]:
        assert sheet in workbook.sheetnames

    study_headers = _header_map(workbook["Study_Design"], 2)
    for column_name in [
        "generalization_claim",
        "nested_cv",
        "external_validation_planned",
        "blocking_strategy",
        "randomization_strategy",
        "replication_strategy",
        "stopping_rule",
    ]:
        assert column_name in study_headers


def test_compile_workbook_missing_required_columns_raises(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    ws = workbook["Experiment_Definitions"]
    ws["C1"] = "start_section_broken"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="missing required columns"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_invalid_section_name_raises(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_executable_row(workbook_path)
    workbook = load_workbook(workbook_path)
    ws = workbook["Experiment_Definitions"]
    ws["C2"] = "invalid_section_name"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Invalid start_section"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_invalid_base_artifact_usage_raises(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_executable_row(workbook_path)
    workbook = load_workbook(workbook_path)
    ws = workbook["Experiment_Definitions"]
    ws["E2"] = "artifact_example_123"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Invalid base artifact usage"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_with_search_space_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_executable_row(workbook_path)

    workbook = load_workbook(workbook_path)
    defs_ws = workbook["Experiment_Definitions"]
    search_ws = workbook["Search_Spaces"]
    headers = [defs_ws.cell(1, col).value for col in range(1, defs_ws.max_column + 1)]
    defs_col = {str(name): idx + 1 for idx, name in enumerate(headers)}
    defs_ws.cell(2, defs_col["search_space_id"], "SS01")

    search_headers = [search_ws.cell(2, col).value for col in range(1, search_ws.max_column + 1)]
    search_col = {str(name): idx + 1 for idx, name in enumerate(search_headers)}
    search_ws.cell(3, search_col["enabled"], "Yes")
    search_ws.cell(4, search_col["enabled"], "Yes")
    workbook.save(workbook_path)

    manifest = compile_workbook_file(workbook_path)
    assert len(manifest.search_spaces) == 1
    assert manifest.search_spaces[0].search_space_id == "SS01"
    assert manifest.trial_specs[0].search_space_id == "SS01"


def test_compile_workbook_full_factorial_expands_design_cells(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)

    manifest = compile_workbook_file(workbook_path)
    study_trials = [trial for trial in manifest.trial_specs if trial.study_id == "S01"]
    assert len(study_trials) == 4
    assert len(manifest.generated_design_matrix) == 4


def test_compile_workbook_constraints_remove_invalid_combinations(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=True)

    manifest = compile_workbook_file(workbook_path)
    study_trials = [trial for trial in manifest.trial_specs if trial.study_id == "S01"]
    assert len(study_trials) == 3


def test_compile_workbook_fixed_controls_propagate_to_generated_trials(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)

    manifest = compile_workbook_file(workbook_path)
    study_trials = [trial for trial in manifest.trial_specs if trial.study_id == "S01"]
    assert study_trials
    for trial in study_trials:
        assert trial.params["target"] == "coarse_affect"
        assert trial.params["cv"] == "within_subject_loso_session"
        assert trial.params["subject"] == "sub-001"


def test_compile_workbook_exploratory_with_rigor_metadata_passes(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)

    workbook = load_workbook(workbook_path)
    study_ws = workbook["Study_Design"]
    study_col = _header_map(study_ws, 2)
    study_ws.cell(3, study_col["generalization_claim"], "Within-subject session transfer only.")
    study_ws.cell(3, study_col["nested_cv"], "No")
    study_ws.cell(3, study_col["external_validation_planned"], "No")
    study_ws.cell(3, study_col["blocking_strategy"], "none")
    study_ws.cell(3, study_col["randomization_strategy"], "fixed_order")
    study_ws.cell(3, study_col["replication_strategy"], "fixed_repeats")
    study_ws.cell(3, study_col["stopping_rule"], "Run all planned cells.")
    workbook.save(workbook_path)
    _set_rigor_metadata(workbook_path)

    manifest = compile_workbook_file(workbook_path)

    assert [spec.study_id for spec in manifest.study_rigor_checklists] == ["S01"]
    assert [spec.study_id for spec in manifest.analysis_plans] == ["S01"]
    assert manifest.validation_warnings == []


def test_compile_workbook_malformed_rigor_checklist_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)
    _set_rigor_metadata(workbook_path)

    workbook = load_workbook(workbook_path)
    checklist_ws = workbook["Study_Rigor_Checklist"]
    checklist_col = _header_map(checklist_ws, 2)
    checklist_ws.cell(3, checklist_col["leakage_risk_reviewed"], "maybe")
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Study_Rigor_Checklist.*Use Yes/No"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_malformed_analysis_plan_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)
    _set_rigor_metadata(workbook_path)

    workbook = load_workbook(workbook_path)
    analysis_ws = workbook["Analysis_Plan"]
    analysis_col = _header_map(analysis_ws, 2)
    analysis_ws.cell(3, analysis_col["aggregation_level"], "invalid_level")
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Invalid analysis plan 'S01'"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_duplicate_rigor_rows_fail(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)
    _set_rigor_metadata(workbook_path)

    workbook = load_workbook(workbook_path)
    checklist_ws = workbook["Study_Rigor_Checklist"]
    checklist_col = _header_map(checklist_ws, 2)
    checklist_ws.cell(4, checklist_col["study_id"], "S01")
    checklist_ws.cell(4, checklist_col["leakage_risk_reviewed"], "Yes")
    checklist_ws.cell(4, checklist_col["deployment_boundary_defined"], "Yes")
    checklist_ws.cell(4, checklist_col["unit_of_analysis_defined"], "Yes")
    checklist_ws.cell(4, checklist_col["data_hierarchy_defined"], "Yes")
    checklist_ws.cell(4, checklist_col["missing_data_plan"], "same")
    checklist_ws.cell(4, checklist_col["class_imbalance_plan"], "same")
    checklist_ws.cell(4, checklist_col["subgroup_plan"], "same")
    checklist_ws.cell(4, checklist_col["reporting_checklist_completed"], "Yes")
    checklist_ws.cell(4, checklist_col["risk_of_bias_reviewed"], "Yes")
    checklist_ws.cell(4, checklist_col["confirmatory_lock_applied"], "No")
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="duplicate entries for study_id 'S01'"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_rigor_unknown_study_reference_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)

    workbook = load_workbook(workbook_path)
    checklist_ws = workbook["Study_Rigor_Checklist"]
    checklist_col = _header_map(checklist_ws, 2)
    checklist_ws.cell(3, checklist_col["study_id"], "UNKNOWN")
    checklist_ws.cell(3, checklist_col["leakage_risk_reviewed"], "Yes")
    checklist_ws.cell(3, checklist_col["deployment_boundary_defined"], "Yes")
    checklist_ws.cell(3, checklist_col["unit_of_analysis_defined"], "Yes")
    checklist_ws.cell(3, checklist_col["data_hierarchy_defined"], "Yes")
    checklist_ws.cell(3, checklist_col["missing_data_plan"], "plan")
    checklist_ws.cell(3, checklist_col["class_imbalance_plan"], "plan")
    checklist_ws.cell(3, checklist_col["subgroup_plan"], "plan")
    checklist_ws.cell(3, checklist_col["reporting_checklist_completed"], "Yes")
    checklist_ws.cell(3, checklist_col["risk_of_bias_reviewed"], "Yes")
    checklist_ws.cell(3, checklist_col["confirmatory_lock_applied"], "No")
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Study_Rigor_Checklist row .* unknown study_id"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_confirmatory_missing_key_rigor_fields_fails(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False, intent="confirmatory")
    workbook = load_workbook(workbook_path)
    study_ws = workbook["Study_Design"]
    study_col = _header_map(study_ws, 2)
    study_ws.cell(3, study_col["generalization_claim"], "Within-dataset confirmatory scope.")
    workbook.save(workbook_path)
    _set_rigor_metadata(
        workbook_path,
        confirmatory_lock_applied="No",
        primary_contrast="",
        multiplicity_handling="",
        interpretation_rules="",
    )

    with pytest.raises(ValueError, match="confirmatory_lock_applied=Yes"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_exploratory_without_rigor_sheets_still_compiles(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, add_constraint=False)
    workbook = load_workbook(workbook_path)
    study_ws = workbook["Study_Design"]
    study_col = _header_map(study_ws, 2)
    study_ws.cell(3, study_col["generalization_claim"], "")
    study_ws.cell(3, study_col["primary_metric"], "")
    study_ws.cell(3, study_col["cv_scheme"], "")
    workbook.save(workbook_path)

    manifest = compile_workbook_file(workbook_path)

    study_trials = [trial for trial in manifest.trial_specs if trial.study_id == "S01"]
    assert len(study_trials) == 4
    assert any("missing generalization_claim" in warning for warning in manifest.validation_warnings)
    assert any("no Study_Rigor_Checklist entry" in warning for warning in manifest.validation_warnings)
    assert any("no Analysis_Plan entry" in warning for warning in manifest.validation_warnings)


def test_compile_workbook_fractional_factorial_raises_clear_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path, study_type="fractional_factorial")

    with pytest.raises(ValueError, match="fractional_factorial"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_factor_referencing_unknown_study_raises(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    ws = workbook["Factors"]
    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    col = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}
    ws.cell(3, col["study_id"], "UNKNOWN_STUDY")
    ws.cell(3, col["factor_name"], "model")
    ws.cell(3, col["parameter_path"], "model")
    ws.cell(3, col["factor_type"], "categorical")
    ws.cell(3, col["levels"], "ridge|logreg")
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="unknown study_id"):
        compile_workbook_file(workbook_path)


def test_compile_workbook_unsupported_schema_version_raises(tmp_path: Path) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_executable_row(workbook_path)
    workbook = load_workbook(workbook_path)
    readme = workbook["README"]
    readme["A46"] = "workbook_schema_version"
    readme["B46"] = "workbook-v999"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Unsupported workbook schema version"):
        compile_workbook_file(workbook_path)
