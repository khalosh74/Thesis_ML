from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_MACHINE_STATUS_SHEET = "Machine_Status"
_TRIAL_RESULTS_SHEET = "Trial_Results"
_SUMMARY_OUTPUTS_SHEET = "Summary_Outputs"
_RUN_LOG_SHEET = "Run_Log"

_MACHINE_STATUS_COLUMNS = [
    "machine_id",
    "hostname",
    "environment_name",
    "python_version",
    "gpu",
    "status",
    "last_checked",
    "notes",
]
_TRIAL_RESULTS_COLUMNS = [
    "trial_id",
    "experiment_id",
    "run_id",
    "status",
    "primary_metric_name",
    "primary_metric_value",
    "report_path",
    "metrics_path",
    "artifact_bundle",
    "notes",
]
_SUMMARY_OUTPUTS_COLUMNS = [
    "summary_type",
    "summary_key",
    "primary_metric_name",
    "primary_metric_value",
    "run_id",
    "experiment_id",
    "start_section",
    "end_section",
    "model",
    "cv",
    "target",
    "xai_method",
    "report_path",
    "notes",
]
_RUN_LOG_COLUMNS = [
    "Run_ID",
    "Experiment_ID",
    "Run_Date",
    "Dataset_Name",
    "Data_Subset",
    "Data_Slice_ID",
    "Grouping_Strategy_ID",
    "Code_Commit_or_Version",
    "Config_File_or_Path",
    "Random_Seed",
    "Target",
    "Split_ID_or_Fold_Definition",
    "Train_Group_Rule",
    "Test_Group_Rule",
    "Transfer_Direction",
    "Session_Coverage",
    "Task_Coverage",
    "Modality_Coverage",
    "Model",
    "Feature_Set",
    "Run_Type",
    "Affects_Frozen_Pipeline",
    "Eligible_for_Method_Decision",
    "Sample_Count",
    "Class_Counts",
    "Imbalance_Status",
    "Leakage_Check_Status",
    "Primary_Metric_Value",
    "Secondary_Metric_1",
    "Secondary_Metric_2",
    "Robustness_Output_Summary",
    "Result_Summary",
    "Preliminary_Interpretation",
    "Reviewed",
    "Used_in_Thesis",
    "Artifact_Path",
    "Notes",
]


def _sanitize_version_tag(version_tag: str) -> str:
    text = str(version_tag).strip()
    if not text:
        raise ValueError("version_tag must be non-empty.")
    return re.sub(r"[^A-Za-z0-9._-]+", "_", text)


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = str(cell.value).strip() if cell.value is not None else ""
        if key:
            mapping[key] = idx
    return mapping


def _assert_required_columns(
    ws: Worksheet,
    *,
    header_row: int,
    required_columns: list[str],
) -> dict[str, int]:
    mapping = _header_map(ws, header_row)
    missing = [name for name in required_columns if name not in mapping]
    if missing:
        raise ValueError(
            f"Sheet '{ws.title}' missing required columns: {', '.join(missing)}"
        )
    return mapping


def _is_row_empty(ws: Worksheet, row_index: int, key_column_index: int) -> bool:
    value = ws.cell(row=row_index, column=key_column_index).value
    if value is None:
        return True
    return not str(value).strip()


def _find_next_row(
    ws: Worksheet,
    *,
    key_column_index: int,
    data_start_row: int,
) -> int:
    row_index = data_start_row
    while True:
        if _is_row_empty(ws, row_index, key_column_index):
            return row_index
        row_index += 1


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_column: int) -> None:
    for col in range(1, max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst._style = src._style  # type: ignore[attr-defined]


def _update_table_refs(ws: Worksheet, required_row: int) -> None:
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if required_row > max_row:
            table.ref = f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(required_row, max_col).coordinate}"


def _apply_hyperlink(cell, value: Any) -> None:
    text = str(value).strip()
    if not text:
        return
    path = Path(text)
    if path.exists():
        cell.hyperlink = str(path.resolve())
        cell.style = "Hyperlink"


def _append_row(
    ws: Worksheet,
    *,
    header_map: dict[str, int],
    row_payload: dict[str, Any],
    key_column: str,
    data_start_row: int,
    style_template_row: int,
    max_column: int,
    hyperlink_columns: set[str] | None = None,
) -> int:
    key_col_idx = header_map[key_column]
    row_index = _find_next_row(
        ws,
        key_column_index=key_col_idx,
        data_start_row=data_start_row,
    )
    if row_index != style_template_row:
        _copy_row_style(ws, style_template_row, row_index, max_column)
    for key, value in row_payload.items():
        col_idx = header_map.get(key)
        if col_idx is None:
            continue
        cell = ws.cell(row=row_index, column=col_idx)
        cell.value = value
        if hyperlink_columns and key in hyperlink_columns and value:
            _apply_hyperlink(cell, value)
    _update_table_refs(ws, row_index)
    return row_index


def _save_versioned_workbook(
    wb: Workbook,
    *,
    source_workbook_path: Path,
    version_tag: str,
    output_dir: Path | None,
) -> Path:
    safe_tag = _sanitize_version_tag(version_tag)
    resolved_output_dir = output_dir if output_dir is not None else source_workbook_path.parent
    resolved_output_dir.mkdir(parents=True, exist_ok=True)
    output_path = resolved_output_dir / f"{source_workbook_path.stem}__results_{safe_tag}{source_workbook_path.suffix}"
    if output_path.exists():
        raise ValueError(f"Refusing to overwrite existing workbook write-back target: {output_path}")
    wb.save(output_path)
    return output_path


def write_workbook_results(
    *,
    source_workbook_path: Path,
    version_tag: str,
    machine_status_rows: list[dict[str, Any]],
    trial_result_rows: list[dict[str, Any]],
    summary_output_rows: list[dict[str, Any]] | None = None,
    run_log_rows: list[dict[str, Any]] | None = None,
    append_run_log: bool = True,
    output_dir: Path | None = None,
) -> Path:
    workbook_path = Path(source_workbook_path)
    wb = load_workbook(workbook_path, data_only=False)

    if _MACHINE_STATUS_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook missing required write-back sheet: '{_MACHINE_STATUS_SHEET}'")
    if _TRIAL_RESULTS_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook missing required write-back sheet: '{_TRIAL_RESULTS_SHEET}'")
    if _SUMMARY_OUTPUTS_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook missing required write-back sheet: '{_SUMMARY_OUTPUTS_SHEET}'")

    machine_ws = wb[_MACHINE_STATUS_SHEET]
    machine_headers = _assert_required_columns(
        machine_ws,
        header_row=2,
        required_columns=_MACHINE_STATUS_COLUMNS,
    )
    for row in machine_status_rows:
        _append_row(
            machine_ws,
            header_map=machine_headers,
            row_payload=row,
            key_column="machine_id",
            data_start_row=3,
            style_template_row=3,
            max_column=len(_MACHINE_STATUS_COLUMNS),
        )

    trial_ws = wb[_TRIAL_RESULTS_SHEET]
    trial_headers = _assert_required_columns(
        trial_ws,
        header_row=2,
        required_columns=_TRIAL_RESULTS_COLUMNS,
    )
    for row in trial_result_rows:
        _append_row(
            trial_ws,
            header_map=trial_headers,
            row_payload=row,
            key_column="trial_id",
            data_start_row=3,
            style_template_row=3,
            max_column=len(_TRIAL_RESULTS_COLUMNS),
            hyperlink_columns={"report_path", "metrics_path", "artifact_bundle"},
        )

    summary_ws = wb[_SUMMARY_OUTPUTS_SHEET]
    summary_headers = _assert_required_columns(
        summary_ws,
        header_row=2,
        required_columns=_SUMMARY_OUTPUTS_COLUMNS,
    )
    for row in summary_output_rows or []:
        _append_row(
            summary_ws,
            header_map=summary_headers,
            row_payload=row,
            key_column="summary_key",
            data_start_row=3,
            style_template_row=3,
            max_column=len(_SUMMARY_OUTPUTS_COLUMNS),
            hyperlink_columns={"report_path"},
        )

    if append_run_log and run_log_rows:
        if _RUN_LOG_SHEET not in wb.sheetnames:
            raise ValueError(f"Workbook missing optional write-back sheet: '{_RUN_LOG_SHEET}'")
        run_log_ws = wb[_RUN_LOG_SHEET]
        run_log_headers = _assert_required_columns(
            run_log_ws,
            header_row=1,
            required_columns=_RUN_LOG_COLUMNS,
        )
        for row in run_log_rows:
            _append_row(
                run_log_ws,
                header_map=run_log_headers,
                row_payload=row,
                key_column="Run_ID",
                data_start_row=2,
                style_template_row=2,
                max_column=len(_RUN_LOG_COLUMNS),
                hyperlink_columns={"Artifact_Path", "Config_File_or_Path"},
            )

    return _save_versioned_workbook(
        wb,
        source_workbook_path=workbook_path,
        version_tag=version_tag,
        output_dir=output_dir,
    )
