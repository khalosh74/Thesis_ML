from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from Thesis_ML.workbook.template_primitives import (
    COL,
    add_table,
    set_widths,
    style_body,
    style_header,
)


def fill_simple_structured_sheet(
    ws: Worksheet,
    *,
    columns: list[str],
    table_name: str,
    title: str,
    width_map: dict[str, float],
    starter_rows: list[dict[str, str]] | None = None,
) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(size=12, bold=True)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws.cell(1, 1).alignment = Alignment(horizontal="left")
    for col_idx_value, header in enumerate(columns, start=1):
        ws.cell(2, col_idx_value, header)
    style_header(ws, 2, len(columns))

    if starter_rows:
        for row_idx, row in enumerate(starter_rows, start=3):
            for col_idx_value, name in enumerate(columns, start=1):
                ws.cell(row_idx, col_idx_value, row.get(name, ""))

    last = 61
    style_body(ws, 3, last, 1, len(columns))
    end_col = get_column_letter(len(columns))
    add_table(ws, table_name, f"A2:{end_col}{last}", style="TableStyleMedium6")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{end_col}{last}"
    set_widths(ws, width_map)
    return last


__all__ = ["fill_simple_structured_sheet"]
