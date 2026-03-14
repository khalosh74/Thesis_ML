from __future__ import annotations

from pathlib import Path
from typing import Any

import nibabel as nib
import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from Thesis_ML.config.schema_versions import (
    COMPILED_MANIFEST_SCHEMA_VERSION,
    WORKBOOK_SCHEMA_VERSION,
)
from Thesis_ML.data.index_dataset import build_dataset_index
from Thesis_ML.workbook.builder import build_workbook


def _write_nifti(path: Path, data: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    image = nib.Nifti1Image(data.astype(np.float32), affine=np.eye(4, dtype=np.float64))
    nib.save(image, str(path))


def _create_tiny_glm_session(glm_dir: Path, labels: list[str]) -> None:
    glm_dir.mkdir(parents=True, exist_ok=True)

    shape = (3, 3, 3)
    mask = np.zeros(shape, dtype=np.float32)
    mask[1:, 1:, 1:] = 1.0
    _write_nifti(glm_dir / "mask.nii", mask)
    pd.Series(labels).to_csv(glm_dir / "regressor_labels.csv", index=False, header=False)

    for idx, label in enumerate(labels, start=1):
        beta = np.full(shape, fill_value=float(idx), dtype=np.float32)
        if "_anger_" in label:
            beta[1:, 1:, 1:] += 2.0
        if "_happiness_" in label:
            beta[1:, 1:, 1:] -= 2.0
        _write_nifti(glm_dir / f"beta_{idx:04d}.nii", beta)


@pytest.fixture
def acceptance_sample_data(tmp_path: Path) -> dict[str, Path]:
    labels = [
        "run-1_passive_anger_audio",
        "run-1_passive_happiness_audio",
        "run-1_passive_anger_video",
        "run-1_passive_happiness_video",
    ]
    data_root = tmp_path / "Data"
    _create_tiny_glm_session(data_root / "sub-001" / "ses-01" / "BAS2", labels)
    _create_tiny_glm_session(data_root / "sub-001" / "ses-02" / "BAS2", labels)

    index_csv = tmp_path / "dataset_index.csv"
    build_dataset_index(data_root=data_root, out_csv=index_csv)

    return {
        "data_root": data_root,
        "index_csv": index_csv,
        "cache_dir": tmp_path / "cache",
        "output_root": tmp_path / "outputs" / "artifacts" / "decision_support",
        "workbook_output_dir": tmp_path / "outputs" / "workbooks",
    }


@pytest.fixture
def acceptance_sample_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    wb = build_workbook()
    wb.save(workbook_path)

    workbook = load_workbook(workbook_path)
    ws = workbook["Experiment_Definitions"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}

    ws.cell(2, col["experiment_id"], "E16")
    ws.cell(2, col["enabled"], "Yes")
    ws.cell(2, col["start_section"], "dataset_selection")
    ws.cell(2, col["end_section"], "evaluation")
    ws.cell(2, col["base_artifact_id"], "")
    ws.cell(2, col["target"], "coarse_affect")
    ws.cell(2, col["cv"], "within_subject_loso_session")
    ws.cell(2, col["model"], "ridge")
    ws.cell(2, col["subject"], "sub-001")
    ws.cell(2, col["train_subject"], "")
    ws.cell(2, col["test_subject"], "")
    ws.cell(2, col["filter_task"], "")
    ws.cell(2, col["filter_modality"], "")
    ws.cell(2, col["reuse_policy"], "auto")
    ws.cell(2, col["search_space_id"], "")
    workbook.save(workbook_path)
    return workbook_path


@pytest.fixture
def acceptance_factorial_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "thesis_experiment_program_factorial.xlsx"
    wb = build_workbook()
    wb.save(workbook_path)

    workbook = load_workbook(workbook_path)

    study_ws = workbook["Study_Design"]
    study_headers = [study_ws.cell(2, col).value for col in range(1, study_ws.max_column + 1)]
    study_col = {str(name): idx + 1 for idx, name in enumerate(study_headers) if name is not None}
    study_ws.cell(3, study_col["study_id"], "S01")
    study_ws.cell(3, study_col["study_name"], "Acceptance factorial study")
    study_ws.cell(3, study_col["enabled"], "Yes")
    study_ws.cell(3, study_col["study_type"], "full_factorial")
    study_ws.cell(3, study_col["intent"], "exploratory")
    study_ws.cell(3, study_col["start_section"], "dataset_selection")
    study_ws.cell(3, study_col["end_section"], "evaluation")
    study_ws.cell(3, study_col["primary_metric"], "balanced_accuracy")
    study_ws.cell(3, study_col["cv_scheme"], "within_subject_loso_session")
    study_ws.cell(3, study_col["num_repeats"], 1)
    study_ws.cell(3, study_col["replication_mode"], "fixed_repeats")
    study_ws.cell(3, study_col["random_seed_policy"], "fixed")

    factors_ws = workbook["Factors"]
    factor_headers = [factors_ws.cell(2, col).value for col in range(1, factors_ws.max_column + 1)]
    factor_col = {str(name): idx + 1 for idx, name in enumerate(factor_headers) if name is not None}
    factors_ws.cell(3, factor_col["study_id"], "S01")
    factors_ws.cell(3, factor_col["factor_name"], "model")
    factors_ws.cell(3, factor_col["parameter_path"], "model")
    factors_ws.cell(3, factor_col["factor_type"], "categorical")
    factors_ws.cell(3, factor_col["levels"], "ridge|logreg")

    fixed_ws = workbook["Fixed_Controls"]
    fixed_headers = [fixed_ws.cell(2, col).value for col in range(1, fixed_ws.max_column + 1)]
    fixed_col = {str(name): idx + 1 for idx, name in enumerate(fixed_headers) if name is not None}
    fixed_ws.cell(3, fixed_col["study_id"], "S01")
    fixed_ws.cell(3, fixed_col["parameter_path"], "target")
    fixed_ws.cell(3, fixed_col["value"], "coarse_affect")
    fixed_ws.cell(4, fixed_col["study_id"], "S01")
    fixed_ws.cell(4, fixed_col["parameter_path"], "subject")
    fixed_ws.cell(4, fixed_col["value"], "sub-001")

    block_ws = workbook["Blocking_and_Replication"]
    block_headers = [block_ws.cell(2, col).value for col in range(1, block_ws.max_column + 1)]
    block_col = {str(name): idx + 1 for idx, name in enumerate(block_headers) if name is not None}
    block_ws.cell(3, block_col["study_id"], "S01")
    block_ws.cell(3, block_col["block_type"], "none")
    block_ws.cell(3, block_col["repeat_id"], 1)

    workbook.save(workbook_path)
    return workbook_path


@pytest.fixture
def acceptance_expected_manifest_shape() -> dict[str, Any]:
    return {
        "schema_version": WORKBOOK_SCHEMA_VERSION,
        "compiled_manifest_schema_version": COMPILED_MANIFEST_SCHEMA_VERSION,
        "experiment_id": "E16",
        "trial_count": 1,
        "required_trial_params": {
            "target": "coarse_affect",
            "cv": "within_subject_loso_session",
            "model": "ridge",
        },
    }


@pytest.fixture
def acceptance_expected_output_shape() -> dict[str, str]:
    return {
        "trial_status": "completed",
        "summary_type": "best_full_pipeline",
        "run_log_result": "completed",
    }
