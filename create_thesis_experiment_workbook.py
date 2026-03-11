from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUT_XLSX = Path("thesis_experiment_program.xlsx")

SHEET_ORDER = [
    "README",
    "Master_Experiments",
    "Run_Log",
    "Decision_Log",
    "Confirmatory_Set",
    "Thesis_Map",
    "Dictionary_Validation",
    "Dashboard",
    "Claim_Ledger",
    "AI_Usage_Log",
    "Ethics_Governance_Notes",
]

STAGE_V2 = [
    "Stage 1 - Target lock",
    "Stage 2 - Split lock",
    "Stage 3 - Model lock",
    "Stage 4 - Feature/preprocessing lock",
    "Stage 5 - Confirmatory analysis",
    "Stage 6 - Robustness analysis",
    "Stage 7 - Exploratory extension",
]

MASTER_COLUMNS = [
    "Experiment_ID",
    "Short_Title",
    "Category",
    "Evidential_Role",
    "Stage",
    "Priority",
    "Decision_Supported",
    "Exact_Question",
    "Hypothesis_or_Expectation",
    "Manipulated_Factor",
    "Held_Constant_Controls",
    "Dataset_Scope",
    "Target_Definition",
    "Split_Logic",
    "Leakage_Risk",
    "Leakage_Control",
    "Models_Compared",
    "Primary_Metric",
    "Secondary_Metrics",
    "Robustness_Checks",
    "Decision_Criterion",
    "Success_Pattern",
    "Failure_or_Warning_Pattern",
    "Threats_to_Validity",
    "Reporting_Destination",
    "Interpretation_Boundary",
    "Status",
    "Owner",
    "Outcome_Summary",
    "Decision_Taken",
    "Notes",
    "Experiment_Ready",
]

RUN_LOG_COLUMNS = [
    "Run_ID",
    "Experiment_ID",
    "Run_Date",
    "Dataset_Name",
    "Data_Subset",
    "Code_Commit_or_Version",
    "Config_File_or_Path",
    "Random_Seed",
    "Target",
    "Split_ID_or_Fold_Definition",
    "Model",
    "Feature_Set",
    "Run_Type",
    "Affects_Frozen_Pipeline",
    "Eligible_for_Method_Decision",
    "Primary_Metric_Value",
    "Secondary_Metric_1",
    "Secondary_Metric_2",
    "Robustness_Output_Summary",
    "Result_Summary",
    "Preliminary_Interpretation",
    "Reviewed",
    "Used_in_Thesis",
    "Artifact_Path",
    "Notes",
]

DECISION_COLUMNS = [
    "Decision_ID",
    "Decision_Topic",
    "Stage",
    "Candidate_Options",
    "Evidence_Experiments",
    "Decision_Criteria",
    "Chosen_Option",
    "Why_Chosen",
    "Why_Not_Others",
    "Risks_or_Tradeoffs",
    "Grounding_Section_or_Source_Packet",
    "Date_Locked",
    "Freeze_Status",
    "Thesis_Use",
]

CONFIRMATORY_COLUMNS = [
    "Experiment_ID",
    "Short_Title",
    "Final_Pipeline_Version",
    "Claim_Supported",
    "Status",
    "Eligible_to_Run",
    "Completed",
    "Required_for_Main_Claim",
    "Ready_for_Chapter_4",
    "Confirmatory_Eligible",
    "Main_Result_Summary",
    "Robustness_Status",
]

THESIS_MAP_COLUMNS = [
    "Experiment_ID",
    "Chapter",
    "Section",
    "Purpose_in_Thesis",
    "Main_or_Supporting",
    "Method_Choice_or_Method_Application_or_Discussion",
    "What_Claim_or_Decision_It_Supports",
    "Notes_for_Writing",
]

CLAIM_LEDGER_COLUMNS = [
    "Claim_ID",
    "Claim_Text",
    "Claim_Type",
    "Supported_By_Experiments",
    "Evidence_Status",
    "Writing_Location",
    "Interpretation_Limit",
    "Notes",
]

AI_USAGE_COLUMNS = [
    "Entry_ID",
    "Date",
    "Task",
    "Tool",
    "Input_Summary",
    "Output_Summary",
    "What_Was_Adopted",
    "Human_Verification_Status",
    "Thesis_Use",
    "Notes",
]

ETHICS_COLUMNS = [
    "Note_ID",
    "Experiment_ID_or_Decision_ID",
    "Topic",
    "Risk_or_Concern",
    "Mitigation_or_Response",
    "Thesis_Section",
    "Status",
    "Notes",
]

VOCABS = {
    "Category": [
        "Target-definition",
        "Split-design",
        "Model-comparison",
        "Preprocessing-feature",
        "Robustness",
        "Confirmatory core",
        "External exploratory",
    ],
    "Evidential_Role": [
        "Primary confirmatory",
        "Primary-supporting robustness",
        "Secondary decision-support",
        "Secondary decision-support / robustness",
        "Exploratory extension",
    ],
    "Stage": STAGE_V2,
    "Priority": ["Critical", "High", "Medium", "Low"],
    "Status": ["Planned", "Not started", "Running", "Completed", "On hold", "Dropped"],
    "Reporting_Destination": [
        "Chapter 3 Method choice",
        "Chapter 4 Main results",
        "Chapter 4 Supporting robustness",
        "Chapter 5 Discussion",
        "Appendix",
    ],
    "Reviewed": ["No", "Yes"],
    "Used_in_Thesis": ["No", "Yes"],
    "Freeze_Status": ["Open", "Locked"],
    "Run_Type": ["Pilot", "Decision-support", "Confirmatory", "Robustness", "Exploratory"],
    "Affects_Frozen_Pipeline": ["No", "Yes"],
    "Eligible_for_Method_Decision": ["No", "Yes"],
    "Claim_Type": ["Method-choice", "Confirmatory finding", "Robustness support", "Exploratory observation"],
    "Evidence_Status": ["open", "partial", "supported"],
    "Writing_Location": ["Chapter 3", "Chapter 4", "Chapter 5", "Appendix"],
    "Human_Verification_Status": ["Not reviewed", "Partially verified", "Verified"],
    "Ethics_Status": ["Open", "Monitoring", "Closed"],
    "Ethics_Topic": ["Data governance", "Bias/fairness", "Interpretation limits", "AI/tool transparency", "Reproducibility", "Other"],
    "Required_for_Main_Claim": ["No", "Yes"],
    "Ready_for_Chapter_4": ["No", "Yes"],
    "YesNo": ["No", "Yes"],
}

DEFINITIONS = [
    ("confirmatory", "Pre-specified, locked-pipeline evidence analysis supporting main thesis claims."),
    ("decision-support", "Method-choice analysis performed before locking confirmatory settings."),
    ("exploratory", "Hypothesis-generating extension outside confirmatory core."),
    ("construct validity", "How well labels/features operationalize intended constructs."),
    ("internal validity", "Protection against confounds and leakage in design and execution."),
    ("statistical conclusion validity", "Reliability of metric-based inference."),
    ("external validity", "Extent findings transfer to other data contexts."),
    ("leakage", "Train-test information contamination that inflates apparent performance."),
    ("domain shift", "Distribution shift between train and test domains."),
    ("interpretability stability", "Consistency of model-behavior explanation signals across folds."),
]

COL = {
    "header_bg": "1F4E78",
    "header_fg": "FFFFFF",
    "title_bg": "D9E1F2",
    "zebra": "F8FAFC",
    "confirmatory": "E2F0D9",
    "exploratory": "FCE4D6",
    "critical": "FFF2CC",
    "dropped": "E7E6E6",
    "missing": "FBE5E7",
    "ok": "E2F0D9",
    "bad": "FBE5E7",
    "open": "FCE4D6",
}

THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=COL["header_bg"])
        cell.font = Font(color=COL["header_fg"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def style_body(ws, r1: int, r2: int, c1: int, c2: int, zebra: bool = True) -> None:
    fill = PatternFill("solid", fgColor=COL["zebra"])
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if zebra and r % 2 == 0:
                cell.fill = fill


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_list_validation(ws, formula: str, col: int, start: int, end: int, allow_blank: bool = True) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    ws.add_data_validation(dv)
    letter = get_column_letter(col)
    dv.add(f"{letter}{start}:{letter}{end}")


def add_named_list(wb: Workbook, list_name: str, sheet_name: str, col: int, start: int, end: int) -> None:
    letter = get_column_letter(col)
    wb.defined_names.add(DefinedName(name=list_name, attr_text=f"'{sheet_name}'!${letter}${start}:${letter}${end}"))


def build_experiments() -> list[dict[str, str]]:
    base = {
        "Priority": "Medium",
        "Decision_Supported": "Methodological decision support.",
        "Hypothesis_or_Expectation": "Evaluate under leakage-aware claim-matched split logic before interpretation.",
        "Held_Constant_Controls": "Dataset index version, train-only fitting policy, metric definitions, seed policy, artifact schema.",
        "Dataset_Scope": "Internal repeated-session BAS2 dataset (sub-001 and sub-002) unless explicitly external.",
        "Target_Definition": "coarse_affect unless this experiment manipulates target granularity.",
        "Split_Logic": "Claim-matched leakage-safe split strategy.",
        "Leakage_Risk": "Performance inflation if split logic or fit scope leaks train/test information.",
        "Leakage_Control": "Predefined split manifests, train-only fitting, frozen transfer where applicable, no test-driven tuning.",
        "Models_Compared": "ridge baseline unless model family is manipulated.",
        "Primary_Metric": "Balanced accuracy",
        "Secondary_Metrics": "Macro-F1; class-wise recall; accuracy",
        "Robustness_Checks": "Fold-level checks and stress controls where relevant.",
        "Decision_Criterion": "Pre-specified criterion recorded before interpretation.",
        "Success_Pattern": "Stable claim-consistent performance without leakage indicators.",
        "Failure_or_Warning_Pattern": "Weak-split inflation, instability, class collapse, or directionally inconsistent evidence.",
        "Threats_to_Validity": "Class imbalance, sample limits, domain shift, and model/split sensitivity.",
        "Reporting_Destination": "Chapter 3 Method choice",
        "Interpretation_Boundary": "Interpret only under stated split logic/evidence tier; no causal or localization claims.",
        "Status": "Planned",
        "Owner": "Khaled (thesis author)",
        "Outcome_Summary": "",
        "Decision_Taken": "",
        "Notes": "",
    }
    rows = [
        {"Experiment_ID": "E01", "Short_Title": "Target granularity experiment", "Category": "Target-definition", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 1 - Target lock", "Priority": "High", "Decision_Supported": "Final target lock", "Exact_Question": "Should the thesis use fine-grained emotion labels, three-class coarse affect, or binary valence-like labels?", "Hypothesis_or_Expectation": "coarse_affect likely balances construct validity, class balance, and learnability.", "Manipulated_Factor": "Target definition", "Target_Definition": "fine emotion vs coarse_affect vs binary valence-like", "Split_Logic": "within_subject_loso_session for primary comparison; frozen transfer as secondary check", "Decision_Criterion": "Prefer target balancing construct validity, class balance, sample sufficiency, and learnability under claim-matched evaluation."},
        {"Experiment_ID": "E02", "Short_Title": "Task pooling experiment", "Category": "Target-definition", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 1 - Target lock", "Decision_Supported": "Task pooling policy", "Exact_Question": "Should the main target be learned across all tasks together, or separately by task?", "Manipulated_Factor": "Pooling strategy by task"},
        {"Experiment_ID": "E03", "Short_Title": "Modality pooling experiment", "Category": "Target-definition", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 1 - Target lock", "Decision_Supported": "Modality pooling policy", "Exact_Question": "Should audio, video, and audiovisual conditions be pooled or modeled separately?", "Manipulated_Factor": "Pooling strategy by modality"},
        {"Experiment_ID": "E04", "Short_Title": "Split-strength stress test", "Category": "Split-design", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 2 - Split lock", "Priority": "High", "Decision_Supported": "Primary split policy", "Exact_Question": "How much do conclusions change under weaker split strategies compared with session-held-out evaluation?", "Manipulated_Factor": "Split logic", "Split_Logic": "record-wise random split vs weaker grouped split vs within-person leave-one-session-out", "Leakage_Risk": "High under weak splits", "Decision_Criterion": "Retain strictest split matching claim; weaker splits only as inflation/leakage demonstrations."},
        {"Experiment_ID": "E05", "Short_Title": "Cross-person transfer design", "Category": "Split-design", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 2 - Split lock", "Priority": "High", "Decision_Supported": "Cross-person transfer framing", "Exact_Question": "How should cross-person generalization be framed and evaluated?", "Manipulated_Factor": "Transfer protocol", "Split_Logic": "frozen train-on-A/test-on-B and reverse direction", "Leakage_Control": "Fit only on train subject; no test subject re-fit/re-tune", "Decision_Criterion": "Keep frozen directional transfer as clean cross-case test.", "Interpretation_Boundary": "Transfer evidence under domain shift, not population-level generalization."},
        {"Experiment_ID": "E06", "Short_Title": "Linear model family comparison", "Category": "Model-comparison", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 3 - Model lock", "Priority": "High", "Decision_Supported": "Model family lock", "Exact_Question": "Which linear model is the most appropriate default: ridge, logistic regression, or linear SVM?", "Manipulated_Factor": "Model family", "Models_Compared": "ridge vs logreg vs linearsvc", "Decision_Criterion": "Prefer simplest model with comparable predictive performance and better robustness/stability."},
        {"Experiment_ID": "E07", "Short_Title": "Class weighting experiment", "Category": "Model-comparison", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 3 - Model lock", "Decision_Supported": "Weighting policy", "Exact_Question": "Does class weighting improve fairness across classes without creating instability?", "Manipulated_Factor": "Weighting strategy"},
        {"Experiment_ID": "E08", "Short_Title": "Hyperparameter strategy experiment", "Category": "Model-comparison", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 3 - Model lock", "Decision_Supported": "Tuning policy", "Exact_Question": "Are fixed conservative hyperparameters sufficient, or does light nested tuning materially improve results?", "Manipulated_Factor": "Tuning strategy", "Decision_Criterion": "Keep fixed settings unless nested tuning yields a clear and stable gain."},
        {"Experiment_ID": "E09", "Short_Title": "Whole-brain versus ROI experiment", "Category": "Preprocessing-feature", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 4 - Feature/preprocessing lock", "Decision_Supported": "Feature representation lock", "Exact_Question": "Should the primary representation remain whole-brain masked voxels or use theory-justified ROIs?", "Manipulated_Factor": "Feature space"},
        {"Experiment_ID": "E10", "Short_Title": "Dimensionality reduction experiment", "Category": "Preprocessing-feature", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 4 - Feature/preprocessing lock", "Decision_Supported": "Feature reduction policy", "Exact_Question": "Does unsupervised dimensionality reduction improve performance or stability enough to justify added complexity?", "Manipulated_Factor": "Feature reduction strategy"},
        {"Experiment_ID": "E11", "Short_Title": "Scaling and normalization experiment", "Category": "Preprocessing-feature", "Evidential_Role": "Secondary decision-support", "Stage": "Stage 4 - Feature/preprocessing lock", "Decision_Supported": "Scaling policy lock", "Exact_Question": "How sensitive are the results to scaling strategy?", "Manipulated_Factor": "Scaling policy"},
        {"Experiment_ID": "E12", "Short_Title": "Permutation test experiment", "Category": "Robustness", "Evidential_Role": "Primary-supporting robustness", "Stage": "Stage 6 - Robustness analysis", "Priority": "High", "Decision_Supported": "Chance-level distinguishability support", "Exact_Question": "Are the main predictive results distinguishable from chance under the exact same split logic?", "Manipulated_Factor": "Label structure (true vs permuted)", "Split_Logic": "exact confirmatory split logic with train-only permutation", "Secondary_Metrics": "Empirical p-value; null distribution summary", "Reporting_Destination": "Chapter 4 Supporting robustness"},
        {"Experiment_ID": "E13", "Short_Title": "Trivial baseline experiment", "Category": "Robustness", "Evidential_Role": "Primary-supporting robustness", "Stage": "Stage 6 - Robustness analysis", "Priority": "High", "Decision_Supported": "Non-trivial predictive value support", "Exact_Question": "Does the final model clearly outperform trivial baselines?", "Manipulated_Factor": "Baseline type", "Models_Compared": "Final locked model vs trivial baselines", "Reporting_Destination": "Chapter 4 Supporting robustness"},
        {"Experiment_ID": "E14", "Short_Title": "Stability of explanation experiment", "Category": "Robustness", "Evidential_Role": "Primary-supporting robustness", "Stage": "Stage 6 - Robustness analysis", "Priority": "High", "Decision_Supported": "Interpretability stability support", "Exact_Question": "Are linear coefficients or importance patterns stable across held-out-session folds?", "Manipulated_Factor": "Fold identity", "Primary_Metric": "Mean pairwise coefficient correlation", "Secondary_Metrics": "Sign consistency; top-k overlap", "Reporting_Destination": "Chapter 4 Supporting robustness", "Interpretation_Boundary": "Model-behavior evidence only; not direct neural localization."},
        {"Experiment_ID": "E15", "Short_Title": "Sensitivity to subset exclusion", "Category": "Robustness", "Evidential_Role": "Secondary decision-support / robustness", "Stage": "Stage 6 - Robustness analysis", "Decision_Supported": "Sensitivity framing", "Exact_Question": "Do the main conclusions depend excessively on one subset, such as one task or modality?", "Manipulated_Factor": "Subset inclusion", "Reporting_Destination": "Chapter 5 Discussion"},
        {"Experiment_ID": "E16", "Short_Title": "Final within-person confirmatory analysis", "Category": "Confirmatory core", "Evidential_Role": "Primary confirmatory", "Stage": "Stage 5 - Confirmatory analysis", "Priority": "Critical", "Decision_Supported": "Primary confirmatory claim", "Exact_Question": "Does the final locked pipeline show meaningful within-person held-out-session generalization?", "Manipulated_Factor": "None; final frozen pipeline", "Target_Definition": "Locked target after Stage 1 target lock", "Split_Logic": "within_subject_loso_session", "Models_Compared": "Final locked model", "Decision_Criterion": "Primary evidential core of thesis.", "Reporting_Destination": "Chapter 4 Main results"},
        {"Experiment_ID": "E17", "Short_Title": "Final cross-person transfer analysis", "Category": "Confirmatory core", "Evidential_Role": "Primary confirmatory", "Stage": "Stage 5 - Confirmatory analysis", "Priority": "Critical", "Decision_Supported": "Secondary confirmatory transfer claim", "Exact_Question": "Does the final locked pipeline transfer meaningfully across individuals under frozen directional transfer?", "Manipulated_Factor": "None; final frozen pipeline", "Target_Definition": "Locked target after Stage 1 target lock", "Split_Logic": "frozen_cross_person_transfer both directions", "Models_Compared": "Final locked model", "Decision_Criterion": "Interpret as cross-case transfer evidence under domain shift.", "Reporting_Destination": "Chapter 4 Main results", "Interpretation_Boundary": "Directional transfer evidence only; not population-level generalization."},
        {"Experiment_ID": "E18", "Short_Title": "External open-source split replication", "Category": "External exploratory", "Evidential_Role": "Exploratory extension", "Stage": "Stage 7 - Exploratory extension", "Priority": "Low", "Decision_Supported": "External split-consistency exploration", "Exact_Question": "Does the same leakage-aware split logic materially affect conclusions on a compatible external open-source dataset?", "Manipulated_Factor": "Dataset source", "Dataset_Scope": "External open-source dataset if available and approved", "Reporting_Destination": "Appendix"},
        {"Experiment_ID": "E19", "Short_Title": "External open-source model portability", "Category": "External exploratory", "Evidential_Role": "Exploratory extension", "Stage": "Stage 7 - Exploratory extension", "Priority": "Low", "Decision_Supported": "External model portability exploration", "Exact_Question": "Does the relative ranking of reasonable model choices remain similar on external data?", "Manipulated_Factor": "Dataset source with same small model set", "Dataset_Scope": "External open-source dataset if available and approved", "Models_Compared": "ridge vs logreg vs linearsvc", "Reporting_Destination": "Appendix"},
    ]
    out: list[dict[str, str]] = []
    for row in rows:
        item = dict(base)
        item.update(row)
        out.append(item)
    return out


def fill_readme_sheet(ws) -> None:
    ws.merge_cells("A1:I1")
    ws["A1"] = "Thesis Experiment Program Workbook (v2)"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws["A1"].alignment = Alignment(horizontal="left")

    blocks = [
        ("Purpose", "Scientific control system for thesis experiment governance: pre-interpretation design, lock tracking, leakage-aware execution traceability, and chapter-ready evidence mapping."),
        ("Evidence tiers", "Primary confirmatory; Primary-supporting robustness; Secondary decision-support; Exploratory extension. Each tier has explicit interpretation boundaries."),
        ("Stage system", "Stage 1 Target lock -> Stage 2 Split lock -> Stage 3 Model lock -> Stage 4 Feature/preprocessing lock -> Stage 5 Confirmatory analysis -> Stage 6 Robustness analysis -> Stage 7 Exploratory extension."),
        ("Freeze policy", "Confirmatory eligibility requires D01-D07 locked in Decision_Log. Chapter 4 readiness additionally requires required confirmatory/supporting items completed and marked Ready_for_Chapter_4=YES."),
        ("Governance additions", "Claim_Ledger enforces claim discipline. AI_Usage_Log supports tool transparency and human verification traceability. Ethics_Governance_Notes supports risk/mitigation accountability."),
        ("Workflow", "Master_Experiments -> Run_Log -> Decision_Log -> Confirmatory_Set -> Claim_Ledger/AI_Usage_Log/Ethics_Governance_Notes -> Thesis_Map -> Dashboard."),
    ]
    row = 3
    for title, text in blocks:
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="EEF3FB")
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=9)
        ws.cell(row=row, column=2, value=text).alignment = Alignment(wrap_text=True, vertical="top")
        for rr in (row, row + 1):
            for cc in range(1, 10):
                ws.cell(rr, cc).border = THIN
        row += 3
    ws["A22"] = "This workbook is aligned to thesis method-choice/method-application workflow and reporting governance requirements."
    ws.merge_cells("A22:I23")
    ws["A22"].fill = PatternFill("solid", fgColor="FFF8E1")
    ws["A22"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A22"].border = THIN
    set_widths(ws, {"A": 22, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24, "G": 24, "H": 24, "I": 24})
    ws.freeze_panes = "A2"


def fill_master_sheet(ws) -> int:
    for i, h in enumerate(MASTER_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(MASTER_COLUMNS))

    rows = build_experiments()
    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(MASTER_COLUMNS[:-1], start=1):
            ws.cell(r, c, row.get(name, ""))
        ws.cell(
            r,
            32,
            f'=IF($AA{r}="Dropped","N/A",IF(AND($H{r}<>"",$J{r}<>"",$K{r}<>"",$M{r}<>"",$N{r}<>"",$P{r}<>"",$R{r}<>"",$U{r}<>"",$V{r}<>"",$W{r}<>"",$Z{r}<>"",$Y{r}<>""),"READY","INCOMPLETE"))',
        )

    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(MASTER_COLUMNS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}{last}"

    set_widths(
        ws,
        {
            "A": 13, "B": 30, "C": 20, "D": 32, "E": 28, "F": 10, "G": 24, "H": 46, "I": 34, "J": 24,
            "K": 42, "L": 30, "M": 30, "N": 34, "O": 24, "P": 36, "Q": 24, "R": 18, "S": 26, "T": 36,
            "U": 36, "V": 34, "W": 36, "X": 34, "Y": 30, "Z": 36, "AA": 14, "AB": 18, "AC": 32,
            "AD": 28, "AE": 32, "AF": 16,
        },
    )

    add_list_validation(ws, "=List_Category", 3, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Evidential_Role", 4, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Stage", 5, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Priority", 6, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Reporting_Destination", 25, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Status", 27, 2, 500, allow_blank=False)

    rng = f"A2:AF{max(last, 200)}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=['$D2="Primary confirmatory"'], fill=PatternFill("solid", fgColor=COL["confirmatory"])))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['ISNUMBER(SEARCH("Exploratory",$D2))'], fill=PatternFill("solid", fgColor=COL["exploratory"])))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['$F2="Critical"'], fill=PatternFill("solid", fgColor=COL["critical"])))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['$AA2="Dropped"'], fill=PatternFill("solid", fgColor=COL["dropped"]), font=Font(color="6E6E6E", italic=True)))
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=['OR($H2="",$J2="",$K2="",$M2="",$N2="",$P2="",$R2="",$U2="",$V2="",$W2="",$Z2="",$Y2="")'],
            fill=PatternFill("solid", fgColor=COL["missing"]),
        ),
    )
    return last


def fill_run_log_sheet(ws) -> int:
    for i, h in enumerate(RUN_LOG_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(RUN_LOG_COLUMNS))
    last = 41
    style_body(ws, 2, last, 1, len(RUN_LOG_COLUMNS))
    add_table(ws, "RunLogTable", f"A1:Y{last}", style="TableStyleMedium9")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Y{last}"

    add_list_validation(ws, "=List_Run_Type", 13, 2, 1000)
    add_list_validation(ws, "=List_Affects_Frozen_Pipeline", 14, 2, 1000)
    add_list_validation(ws, "=List_Eligible_for_Method_Decision", 15, 2, 1000)
    add_list_validation(ws, "=List_Reviewed", 22, 2, 1000)
    add_list_validation(ws, "=List_Used_in_Thesis", 23, 2, 1000)

    set_widths(
        ws,
        {
            "A": 18, "B": 13, "C": 12, "D": 20, "E": 24, "F": 22, "G": 30, "H": 12, "I": 16, "J": 24,
            "K": 16, "L": 22, "M": 18, "N": 18, "O": 22, "P": 18, "Q": 18, "R": 18, "S": 30, "T": 28,
            "U": 34, "V": 12, "W": 14, "X": 36, "Y": 28,
        }
    )
    return last


def fill_decision_log_sheet(ws) -> int:
    for i, h in enumerate(DECISION_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(DECISION_COLUMNS))

    rows = [
        ("D01", "Final target choice", "Stage 1 - Target lock", "fine emotion; coarse_affect; binary valence-like", "E01,E02,E03", "Construct validity + class balance + learnability under strict split", "To be locked after method-choice evidence review", "", "", "Tradeoff between construct validity and learnability", "To be linked after background/method revision", "", "Open", "Chapter 3 method lock feeding Chapter 4 confirmatory analyses"),
        ("D02", "Final split design", "Stage 2 - Split lock", "within_subject_loso_session; weaker alternatives for stress testing", "E04,E05", "Prefer strictest split matching claim while documenting weak-split inflation", "To be locked after split-strength evidence review", "", "", "Weak split inflation risk", "To be linked after method section split rationale update", "", "Open", "Defines primary inference logic and leakage controls"),
        ("D03", "Final model family", "Stage 3 - Model lock", "ridge; logreg; linearsvc", "E06", "Simplicity + robustness + comparable performance", "To be locked after model comparison", "", "", "Complexity vs stability tradeoff", "To be linked after model-choice subsection revision", "", "Open", "Sets final confirmatory model family"),
        ("D04", "Final weighting policy", "Stage 3 - Model lock", "No class weighting; balanced weighting", "E07", "Minority-class fairness gain without instability", "To be locked after weighting sensitivity", "", "", "Fairness-stability tension", "To be linked after class-imbalance discussion draft", "", "Open", "Controls fairness assumptions in final pipeline"),
        ("D05", "Final tuning policy", "Stage 3 - Model lock", "Fixed explicit settings; light nested tuning", "E08", "Keep fixed unless clear stable gain from tuning", "To be locked after tuning strategy review", "", "", "Overfitting and complexity risk with tuning", "To be linked after reproducibility subsection update", "", "Open", "Determines model configuration policy"),
        ("D06", "Final feature representation", "Stage 4 - Feature/preprocessing lock", "Whole-brain masked voxels; theory-justified ROI", "E09,E10", "Validity + stability + complexity tradeoff", "To be locked after feature-space evidence review", "", "", "Representation choice impacts validity and interpretability", "To be linked after feature-method section update", "", "Open", "Locks final feature representation"),
        ("D07", "Final scaling policy", "Stage 4 - Feature/preprocessing lock", "No scaling; standard/robust scaling (train-only)", "E11", "Train-only scaling with best stability and simplicity", "To be locked after scaling sensitivity", "", "", "Scaling can alter class boundaries", "To be linked after preprocessing subsection revision", "", "Open", "Locks preprocessing policy"),
        ("D08", "Use of external data", "Stage 7 - Exploratory extension", "No external data; external replication/portability", "E18,E19", "External work remains exploratory and non-core", "To be decided after confirmatory core completion", "", "", "Comparability and external validity limits", "To be linked after discussion/limitations revision", "", "Open", "Controls exploratory external scope and appendix use"),
    ]

    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)

    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(DECISION_COLUMNS))
    add_table(ws, "DecisionLogTable", f"A1:N{last}", style="TableStyleMedium4")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{last}"
    add_list_validation(ws, "=List_Freeze_Status", 13, 2, 500, allow_blank=False)
    ws.conditional_formatting.add(f"M2:M{max(last,200)}", FormulaRule(formula=['$M2="Locked"'], fill=PatternFill("solid", fgColor=COL["ok"])))
    ws.conditional_formatting.add(f"M2:M{max(last,200)}", FormulaRule(formula=['$M2="Open"'], fill=PatternFill("solid", fgColor=COL["open"])))
    set_widths(
        ws,
        {"A": 12, "B": 24, "C": 30, "D": 36, "E": 18, "F": 34, "G": 34, "H": 22, "I": 22, "J": 28, "K": 34, "L": 12, "M": 14, "N": 30},
    )
    return last


def fill_confirmatory_sheet(ws) -> int:
    for i, h in enumerate(CONFIRMATORY_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(CONFIRMATORY_COLUMNS))

    rows = [
        ("E12", "Permutation test experiment", "To be fixed after freeze lock", "Chance-distinguishability support", "YES"),
        ("E13", "Trivial baseline experiment", "To be fixed after freeze lock", "Non-trivial predictive value support", "YES"),
        ("E14", "Stability of explanation experiment", "To be fixed after freeze lock", "Model-behavior robustness support", "YES"),
        ("E16", "Final within-person confirmatory analysis", "To be fixed after freeze lock", "Primary within-person claim", "YES"),
        ("E17", "Final cross-person transfer analysis", "To be fixed after freeze lock", "Secondary transfer claim", "YES"),
    ]
    for r, row in enumerate(rows, start=2):
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, row[1])
        ws.cell(r, 3, row[2])
        ws.cell(r, 4, row[3])
        ws.cell(r, 5, f'=IFERROR(INDEX(Master_Experiments!$AA:$AA,MATCH($A{r},Master_Experiments!$A:$A,0)),"Planned")')
        ws.cell(
            r,
            10,
            '=IF(AND(COUNTIFS(Decision_Log!$A:$A,"D01",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D02",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D03",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D04",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D05",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D06",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D07",Decision_Log!$M:$M,"Locked")>0),"YES","NO")',
        )
        ws.cell(r, 6, f'=IF(AND($J{r}="YES",IFERROR(INDEX(Master_Experiments!$AF:$AF,MATCH($A{r},Master_Experiments!$A:$A,0)),"INCOMPLETE")="READY"),"YES","NO")')
        ws.cell(r, 7, f'=IF(IFERROR(INDEX(Master_Experiments!$AA:$AA,MATCH($A{r},Master_Experiments!$A:$A,0)),"Planned")="Completed","YES","NO")')
        ws.cell(r, 8, row[4])
        ws.cell(r, 9, "NO")
        ws.cell(r, 11, f'=IFERROR(INDEX(Master_Experiments!$AC:$AC,MATCH($A{r},Master_Experiments!$A:$A,0)),"")')
        ws.cell(r, 12, "Planned")

    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(CONFIRMATORY_COLUMNS))
    add_table(ws, "ConfirmatoryTable", f"A1:L{last}", style="TableStyleMedium10")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{last}"
    add_list_validation(ws, "=List_Status", 5, 2, 500)
    add_list_validation(ws, "=List_Required_for_Main_Claim", 8, 2, 500)
    add_list_validation(ws, "=List_Ready_for_Chapter_4", 9, 2, 500)
    add_list_validation(ws, "=List_Status", 12, 2, 500)
    set_widths(
        ws,
        {"A": 13, "B": 34, "C": 28, "D": 34, "E": 14, "F": 18, "G": 12, "H": 20, "I": 18, "J": 22, "K": 32, "L": 20},
    )
    ws.conditional_formatting.add(f"J2:J{max(last,200)}", FormulaRule(formula=['$J2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])))
    ws.conditional_formatting.add(f"F2:F{max(last,200)}", FormulaRule(formula=['$F2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])))
    ws.conditional_formatting.add(f"G2:G{max(last,200)}", FormulaRule(formula=['$G2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])))
    ws.conditional_formatting.add(f"I2:I{max(last,200)}", FormulaRule(formula=['$I2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])))
    return last


def fill_thesis_map_sheet(ws) -> int:
    for i, h in enumerate(THESIS_MAP_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(THESIS_MAP_COLUMNS))
    rows = [
        ("E01", "Chapter 3", "3.3 Target definition", "Target granularity method choice", "Supporting", "Method choice", "Locks target definition", "Reference Decision D01"),
        ("E02", "Chapter 3", "3.3 Task strategy", "Task pooling method choice", "Supporting", "Method choice", "Task pooling policy", ""),
        ("E03", "Chapter 3", "3.3 Modality strategy", "Modality pooling method choice", "Supporting", "Method choice", "Modality pooling policy", ""),
        ("E04", "Chapter 3", "3.4 Split design", "Split-strength stress evidence", "Supporting", "Method choice", "Justifies strict split", ""),
        ("E05", "Chapter 3", "3.4 Transfer protocol", "Cross-person framing", "Supporting", "Method choice", "Directional transfer design", ""),
        ("E06", "Chapter 3", "3.5 Model family", "Model family selection", "Supporting", "Method choice", "Final model family lock", ""),
        ("E07", "Chapter 3", "3.5 Weighting policy", "Class weighting decision", "Supporting", "Method choice", "Weighting policy lock", ""),
        ("E08", "Chapter 3", "3.5 Tuning policy", "Hyperparameter policy", "Supporting", "Method choice", "Fixed vs tuning decision", ""),
        ("E09", "Chapter 3", "3.2 Feature space", "Whole-brain vs ROI decision", "Supporting", "Method choice", "Feature representation lock", ""),
        ("E10", "Chapter 3", "3.2 Dimensionality", "Reduction strategy decision", "Supporting", "Method choice", "Complexity/stability tradeoff", ""),
        ("E11", "Chapter 3", "3.2 Scaling", "Scaling sensitivity decision", "Supporting", "Method choice", "Scaling policy lock", ""),
        ("E12", "Chapter 4", "4.3 Robustness", "Permutation support", "Supporting", "Method application", "Beyond-chance support", ""),
        ("E13", "Chapter 4", "4.3 Robustness", "Trivial baseline support", "Supporting", "Method application", "Non-trivial value support", ""),
        ("E14", "Chapter 4", "4.3 Robustness", "Interpretability stability support", "Supporting", "Method application", "Model-behavior robustness", "No localization claim"),
        ("E15", "Chapter 5", "5.2 Limitations", "Subset sensitivity discussion", "Supporting", "Discussion", "Sensitivity/limitations support", ""),
        ("E16", "Chapter 4", "4.1 Main results", "Within-person confirmatory evidence", "Main", "Method application", "Primary thesis claim", ""),
        ("E17", "Chapter 4", "4.2 Transfer results", "Cross-person transfer evidence", "Main", "Method application", "Secondary transfer claim", ""),
        ("E18", "Appendix", "A. External exploratory", "External split replication", "Supporting", "Discussion", "Exploratory external consistency", ""),
        ("E19", "Appendix", "A. External exploratory", "External model portability", "Supporting", "Discussion", "Exploratory portability", ""),
    ]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(THESIS_MAP_COLUMNS))
    add_table(ws, "ThesisMapTable", f"A1:H{last}", style="TableStyleMedium6")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last}"
    set_widths(ws, {"A": 13, "B": 12, "C": 24, "D": 30, "E": 14, "F": 44, "G": 34, "H": 30})
    return last


def fill_claim_ledger_sheet(ws) -> int:
    for i, h in enumerate(CLAIM_LEDGER_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(CLAIM_LEDGER_COLUMNS))
    starter = [
        ("C01", "Primary within-person held-out-session decoding demonstrates meaningful generalization under locked pipeline.", "Confirmatory finding", "E16,E12,E13,E14", "open", "Chapter 4", "Interpret only within claim-matched split and current dataset scope.", ""),
        ("C02", "Frozen directional cross-person transfer indicates cross-case portability under domain shift.", "Confirmatory finding", "E17,E12,E13", "open", "Chapter 4", "Not a population-level generalization claim.", ""),
        ("C03", "Method lock decisions reduce leakage risk and interpretation drift.", "Method-choice", "E01,E04,E05,E06,E09,E11", "partial", "Chapter 3", "Governance claim; not a performance claim.", ""),
    ]
    for r, row in enumerate(starter, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    last = 31
    style_body(ws, 2, last, 1, len(CLAIM_LEDGER_COLUMNS))
    add_table(ws, "ClaimLedgerTable", f"A1:H{last}", style="TableStyleMedium8")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last}"
    add_list_validation(ws, "=List_Claim_Type", 3, 2, 1000)
    add_list_validation(ws, "=List_Evidence_Status", 5, 2, 1000)
    add_list_validation(ws, "=List_Writing_Location", 6, 2, 1000)
    set_widths(ws, {"A": 12, "B": 48, "C": 20, "D": 24, "E": 14, "F": 14, "G": 38, "H": 28})
    return last


def fill_ai_usage_sheet(ws) -> int:
    for i, h in enumerate(AI_USAGE_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(AI_USAGE_COLUMNS))
    ws.cell(2, 1, "AI001")
    ws.cell(2, 4, "GPT-5.3-Codex")
    ws.cell(2, 5, "Workbook package revision and governance alignment")
    ws.cell(2, 8, "Not reviewed")
    ws.cell(2, 9, "No")
    last = 41
    style_body(ws, 2, last, 1, len(AI_USAGE_COLUMNS))
    add_table(ws, "AIUsageTable", f"A1:J{last}", style="TableStyleMedium5")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{last}"
    add_list_validation(ws, "=List_Human_Verification_Status", 8, 2, 1000)
    add_list_validation(ws, "=List_Used_in_Thesis", 9, 2, 1000)
    set_widths(ws, {"A": 12, "B": 12, "C": 24, "D": 18, "E": 32, "F": 32, "G": 30, "H": 24, "I": 14, "J": 28})
    return last


def fill_ethics_sheet(ws) -> int:
    for i, h in enumerate(ETHICS_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(ETHICS_COLUMNS))
    starter = [
        ("EN01", "E16", "Interpretation limits", "Risk of over-claiming confirmatory evidence beyond split/domain scope", "Enforce interpretation boundaries in Claim_Ledger and Chapter 5 limitations", "Chapter 5", "Open", ""),
        ("EN02", "D08", "AI/tool transparency", "Insufficient traceability of AI-assisted drafting decisions", "Maintain AI_Usage_Log with human verification status before thesis inclusion", "Chapter 2/Appendix", "Open", ""),
    ]
    for r, row in enumerate(starter, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    last = 31
    style_body(ws, 2, last, 1, len(ETHICS_COLUMNS))
    add_table(ws, "EthicsTable", f"A1:H{last}", style="TableStyleMedium7")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last}"
    add_list_validation(ws, "=List_Ethics_Topic", 3, 2, 1000)
    add_list_validation(ws, "=List_Ethics_Status", 7, 2, 1000)
    set_widths(ws, {"A": 12, "B": 22, "C": 22, "D": 34, "E": 34, "F": 20, "G": 12, "H": 26})
    return last


def fill_dictionary_sheet(ws, wb: Workbook) -> None:
    ws["A1"] = "Controlled vocabularies (drives dropdown validation)"
    ws["A1"].font = Font(size=12, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws.merge_cells("A1:U1")

    names = [
        "Category",
        "Evidential_Role",
        "Stage",
        "Priority",
        "Status",
        "Reporting_Destination",
        "Reviewed",
        "Used_in_Thesis",
        "Freeze_Status",
        "Run_Type",
        "Affects_Frozen_Pipeline",
        "Eligible_for_Method_Decision",
        "Claim_Type",
        "Evidence_Status",
        "Writing_Location",
        "Human_Verification_Status",
        "Ethics_Status",
        "Ethics_Topic",
        "Required_for_Main_Claim",
        "Ready_for_Chapter_4",
        "YesNo",
    ]
    for c, name in enumerate(names, start=1):
        ws.cell(2, c, name)
        ws.cell(2, c).font = Font(color=COL["header_fg"], bold=True)
        ws.cell(2, c).fill = PatternFill("solid", fgColor=COL["header_bg"])
        items = VOCABS[name]
        for r, item in enumerate(items, start=3):
            ws.cell(r, c, item)
            ws.cell(r, c).border = THIN
            if r % 2 == 0:
                ws.cell(r, c).fill = PatternFill("solid", fgColor=COL["zebra"])
        add_named_list(wb, f"List_{name}", ws.title, c, 3, 2 + len(items))

    ws["V1"] = "Term definitions"
    ws["V1"].font = Font(size=12, bold=True)
    ws["V1"].fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws.merge_cells("V1:W1")
    ws["V2"] = "Term"
    ws["W2"] = "Definition"
    for cell in ("V2", "W2"):
        ws[cell].font = Font(color=COL["header_fg"], bold=True)
        ws[cell].fill = PatternFill("solid", fgColor=COL["header_bg"])
    for r, (term, definition) in enumerate(DEFINITIONS, start=3):
        ws.cell(r, 22, term)
        ws.cell(r, 23, definition)
        ws.cell(r, 22).border = THIN
        ws.cell(r, 23).border = THIN
        ws.cell(r, 23).alignment = Alignment(wrap_text=True, vertical="top")
        if r % 2 == 0:
            ws.cell(r, 22).fill = PatternFill("solid", fgColor=COL["zebra"])
            ws.cell(r, 23).fill = PatternFill("solid", fgColor=COL["zebra"])

    for c in range(1, 22):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.column_dimensions["V"].width = 24
    ws.column_dimensions["W"].width = 84
    ws.freeze_panes = "A3"


def fill_dashboard_sheet(ws) -> None:
    ws.merge_cells("A1:I1")
    ws["A1"] = "Thesis Experiment Program Dashboard (v2)"
    ws["A1"].font = Font(size=15, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COL["title_bg"])

    ws["A3"] = "Program state summary"
    ws["A3"].font = Font(size=12, bold=True)
    ws["A3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("A3:B3")

    labels = [
        "Total number of experiments",
        "Completed experiments",
        "Critical experiments completed",
        "Open decisions",
        "Locked decisions",
        "Target locked?",
        "Split locked?",
        "Model locked?",
        "Feature/preprocessing locked?",
        "Confirmatory eligibility achieved?",
        "Confirmatory ready for Chapter 4?",
        "Experiments mapped to Chapter 4 main results",
        "Experiments mapped only to Appendix",
        "Claims with Evidence_Status = supported",
        "Claims with Evidence_Status = partial",
        "Claims with Evidence_Status = open",
        "AI log entries",
        "Open ethics/governance notes",
    ]
    formulas = [
        '=COUNTA(Master_Experiments!$A$2:$A$500)',
        '=COUNTIF(Master_Experiments!$AA$2:$AA$500,"Completed")',
        '=COUNTIFS(Master_Experiments!$F$2:$F$500,"Critical",Master_Experiments!$AA$2:$AA$500,"Completed")',
        '=COUNTIF(Decision_Log!$M$2:$M$500,"Open")',
        '=COUNTIF(Decision_Log!$M$2:$M$500,"Locked")',
        '=IF(COUNTIFS(Decision_Log!$A:$A,"D01",Decision_Log!$M:$M,"Locked")>0,"YES","NO")',
        '=IF(COUNTIFS(Decision_Log!$A:$A,"D02",Decision_Log!$M:$M,"Locked")>0,"YES","NO")',
        '=IF(AND(COUNTIFS(Decision_Log!$A:$A,"D03",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D04",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D05",Decision_Log!$M:$M,"Locked")>0),"YES","NO")',
        '=IF(AND(COUNTIFS(Decision_Log!$A:$A,"D06",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D07",Decision_Log!$M:$M,"Locked")>0),"YES","NO")',
        '=IF(COUNTIFS(Confirmatory_Set!$J:$J,"YES")=COUNTIFS(Confirmatory_Set!$A:$A,"<>"),"YES","NO")',
        '=IF(AND(B13="YES",COUNTIFS(Confirmatory_Set!$H:$H,"YES",Confirmatory_Set!$G:$G,"YES",Confirmatory_Set!$I:$I,"YES")=COUNTIFS(Confirmatory_Set!$H:$H,"YES")),"YES","NO")',
        '=COUNTIFS(Thesis_Map!$B$2:$B$500,"Chapter 4",Thesis_Map!$E$2:$E$500,"Main")',
        '=COUNTIF(Thesis_Map!$B$2:$B$500,"Appendix")',
        '=COUNTIF(Claim_Ledger!$E$2:$E$500,"supported")',
        '=COUNTIF(Claim_Ledger!$E$2:$E$500,"partial")',
        '=COUNTIF(Claim_Ledger!$E$2:$E$500,"open")',
        '=COUNTA(AI_Usage_Log!$A$2:$A$500)',
        '=COUNTIF(Ethics_Governance_Notes!$G$2:$G$500,"Open")',
    ]

    for i, (label, formula) in enumerate(zip(labels, formulas, strict=True), start=4):
        ws.cell(i, 1, label)
        ws.cell(i, 2, formula)
        ws.cell(i, 1).border = THIN
        ws.cell(i, 2).border = THIN
        if i % 2 == 0:
            ws.cell(i, 1).fill = PatternFill("solid", fgColor=COL["zebra"])
            ws.cell(i, 2).fill = PatternFill("solid", fgColor=COL["zebra"])

    ws["D3"] = "Experiments by category"
    ws["D3"].font = Font(size=12, bold=True)
    ws["D3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("D3:E3")
    for i, _ in enumerate(VOCABS["Category"], start=4):
        ws.cell(i, 4, f"=Dictionary_Validation!A{i-1}")
        ws.cell(i, 5, f"=COUNTIF(Master_Experiments!$C:$C,D{i})")
        ws.cell(i, 4).border = THIN
        ws.cell(i, 5).border = THIN

    ws["G3"] = "Experiments by evidential role"
    ws["G3"].font = Font(size=12, bold=True)
    ws["G3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("G3:H3")
    for i, _ in enumerate(VOCABS["Evidential_Role"], start=4):
        ws.cell(i, 7, f"=Dictionary_Validation!B{i-1}")
        ws.cell(i, 8, f"=COUNTIF(Master_Experiments!$D:$D,G{i})")
        ws.cell(i, 7).border = THIN
        ws.cell(i, 8).border = THIN
        ws.cell(i, 7).alignment = Alignment(wrap_text=True)

    set_widths(ws, {"A": 46, "B": 16, "C": 3, "D": 26, "E": 10, "F": 3, "G": 36, "H": 10, "I": 3})
    ws.freeze_panes = "A4"
    ws.conditional_formatting.add("B9:B14", FormulaRule(formula=['B9="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])))
    ws.conditional_formatting.add("B9:B14", FormulaRule(formula=['B9="NO"'], fill=PatternFill("solid", fgColor=COL["bad"])))


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEET_ORDER:
        wb.create_sheet(name)

    fill_readme_sheet(wb["README"])
    fill_master_sheet(wb["Master_Experiments"])
    fill_run_log_sheet(wb["Run_Log"])
    fill_decision_log_sheet(wb["Decision_Log"])
    fill_confirmatory_sheet(wb["Confirmatory_Set"])
    fill_thesis_map_sheet(wb["Thesis_Map"])
    fill_dictionary_sheet(wb["Dictionary_Validation"], wb)
    fill_dashboard_sheet(wb["Dashboard"])
    fill_claim_ledger_sheet(wb["Claim_Ledger"])
    fill_ai_usage_sheet(wb["AI_Usage_Log"])
    fill_ethics_sheet(wb["Ethics_Governance_Notes"])

    wb.calculation.fullCalcOnLoad = True
    wb["README"]["A25"] = "Generated by create_thesis_experiment_workbook.py (v2)"
    wb["README"]["A25"].font = Font(italic=True, color="4B5563")
    return wb


def validate(path: Path) -> dict[str, str]:
    wb = load_workbook(path)
    sheets = [ws.title for ws in wb.worksheets]
    sheet_ok = sheets == SHEET_ORDER
    missing_sheets = [s for s in SHEET_ORDER if s not in sheets]

    master = wb["Master_Experiments"]
    confirm = wb["Confirmatory_Set"]
    dash = wb["Dashboard"]
    dictionary = wb["Dictionary_Validation"]

    dv_count = sum(
        len(wb[name].data_validations.dataValidation)
        for name in ["Master_Experiments", "Run_Log", "Decision_Log", "Confirmatory_Set", "Claim_Ledger", "AI_Usage_Log", "Ethics_Governance_Notes"]
    )
    stage_values = {master[f"E{r}"].value for r in range(2, master.max_row + 1)}
    stage_vocab = set(STAGE_V2)
    stage_consistent = stage_values.issubset(stage_vocab)

    return {
        "sheet_order_ok": str(sheet_ok),
        "missing_sheets": ", ".join(missing_sheets) if missing_sheets else "None",
        "sheet_count": str(len(sheets)),
        "data_validations_found": str(dv_count),
        "experiment_ready_formula_present": str(bool(master["AF2"].value)),
        "confirmatory_formula_present": str(bool(confirm["F2"].value and confirm["G2"].value and confirm["J2"].value)),
        "dashboard_formula_present": str(bool(dash["B13"].value and dash["B14"].value)),
        "stage_vocab_consistent": str(stage_consistent),
        "stage_vocab_rows": str(len([v for v in stage_values if v is not None])),
        "dictionary_stage_head": str(dictionary["C3"].value),
    }


def main() -> None:
    wb = build_workbook()
    wb.save(OUT_XLSX)
    summary = validate(OUT_XLSX)
    print("Created workbook:", OUT_XLSX.resolve())
    print("Sheet order valid:", summary["sheet_order_ok"])
    print("Missing required sheets:", summary["missing_sheets"])
    print("Sheet count:", summary["sheet_count"])
    print("Data validations found:", summary["data_validations_found"])
    print("Experiment_Ready formula present:", summary["experiment_ready_formula_present"])
    print("Confirmatory formulas present:", summary["confirmatory_formula_present"])
    print("Dashboard formulas present:", summary["dashboard_formula_present"])
    print("Stage vocabulary consistent:", summary["stage_vocab_consistent"])
    print("Stage rows detected:", summary["stage_vocab_rows"])


if __name__ == "__main__":
    main()
