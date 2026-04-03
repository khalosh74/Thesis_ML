from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

from Thesis_ML.workbook.template_constants import (
    AI_USAGE_COLUMNS,
    CLAIM_LEDGER_COLUMNS,
    CONFIRMATORY_COLUMNS,
    DATA_SELECTION_COLUMNS,
    DECISION_COLUMNS,
    DEFINITIONS,
    ETHICS_COLUMNS,
    GROUPING_STRATEGY_COLUMNS,
    RUN_LOG_COLUMNS,
    THESIS_MAP_COLUMNS,
    VOCABS,
)
from Thesis_ML.workbook.template_primitives import (
    COL,
    THIN,
    add_dynamic_named_list,
    add_list_validation,
    add_named_list,
    add_table,
    col_idx,
    set_widths,
    style_body,
    style_header,
)


def fill_data_selection_design_sheet(ws, wb: Workbook) -> int:
    for i, h in enumerate(DATA_SELECTION_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(DATA_SELECTION_COLUMNS))

    rows = [
        {
            "Data_Slice_ID": "DS01",
            "Slice_Name": "All subjects, all sessions, pooled task, pooled modality, coarse_affect",
            "Purpose": "Global baseline for target/split/model method-choice comparisons.",
            "Subject_Scope": "all_subjects",
            "Session_Scope": "all_sessions",
            "Time_Window": "full_study_window",
            "Task_Scope": "pooled_all_tasks",
            "Modality_Scope": "pooled_all_modalities",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect",
            "Inclusion_Rule": "All QC-passed repeated-session beta maps with valid labels.",
            "Exclusion_Rule": "Drop failed-preprocessing rows and missing targets.",
            "Class_Balance_Policy": "as_is",
            "Minimum_Samples_Rule": "Prefer >=20 per class before claim-level interpretation.",
            "Leakage_Risk": "medium",
            "Valid_Use_Case": "method_choice",
            "Thesis_Use": "method_choice",
            "Notes": "",
        },
        {
            "Data_Slice_ID": "DS02",
            "Slice_Name": "Subject-specific pooled-task pooled-modality",
            "Purpose": "Primary within-person held-out-session claim support.",
            "Subject_Scope": "single_subject",
            "Session_Scope": "held_out_session",
            "Time_Window": "full_study_window",
            "Task_Scope": "pooled_all_tasks",
            "Modality_Scope": "pooled_all_modalities",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect",
            "Inclusion_Rule": "One subject at a time with all eligible sessions.",
            "Exclusion_Rule": "No cross-subject pooling inside fold-level fit.",
            "Class_Balance_Policy": "weighted_only",
            "Minimum_Samples_Rule": "Per-class floor checked per subject.",
            "Leakage_Risk": "low",
            "Valid_Use_Case": "confirmatory_within_person",
            "Thesis_Use": "main_confirmatory",
            "Notes": "Matches held-out-session inference boundary.",
        },
        {
            "Data_Slice_ID": "DS03",
            "Slice_Name": "Task-specific pooled-modality",
            "Purpose": "Task pooling vs task-specific method-choice and diagnostics.",
            "Subject_Scope": "all_subjects",
            "Session_Scope": "all_sessions",
            "Time_Window": "full_study_window",
            "Task_Scope": "task_specific_other",
            "Modality_Scope": "pooled_all_modalities",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect by task",
            "Inclusion_Rule": "Restrict each run to one task family.",
            "Exclusion_Rule": "Exclude off-task records for that run.",
            "Class_Balance_Policy": "as_is",
            "Minimum_Samples_Rule": "Minimum counts checked per task-specific slice.",
            "Leakage_Risk": "medium",
            "Valid_Use_Case": "method_choice",
            "Thesis_Use": "method_choice",
            "Notes": "",
        },
        {
            "Data_Slice_ID": "DS04",
            "Slice_Name": "Modality-specific pooled-task",
            "Purpose": "Modality pooling vs modality-specific method-choice checks.",
            "Subject_Scope": "all_subjects",
            "Session_Scope": "all_sessions",
            "Time_Window": "full_study_window",
            "Task_Scope": "pooled_all_tasks",
            "Modality_Scope": "audio_only",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect",
            "Inclusion_Rule": "One modality family per run (audio/video/audiovisual variants).",
            "Exclusion_Rule": "Exclude other modalities in each modality-specific run.",
            "Class_Balance_Policy": "as_is",
            "Minimum_Samples_Rule": "Check sample sufficiency per modality variant.",
            "Leakage_Risk": "medium",
            "Valid_Use_Case": "method_choice",
            "Thesis_Use": "method_choice",
            "Notes": "Seed modality is audio_only; clone row for other modalities.",
        },
        {
            "Data_Slice_ID": "DS05",
            "Slice_Name": "Subject-specific task x modality subset",
            "Purpose": "Fine-grained diagnostic sensitivity slices.",
            "Subject_Scope": "single_subject",
            "Session_Scope": "all_sessions",
            "Time_Window": "full_study_window",
            "Task_Scope": "task_specific_other",
            "Modality_Scope": "audiovisual_only",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect",
            "Inclusion_Rule": "Single subject + selected task + selected modality combination.",
            "Exclusion_Rule": "Exclude all records outside the chosen intersection.",
            "Class_Balance_Policy": "min_count_threshold",
            "Minimum_Samples_Rule": "Hard floor required before reporting.",
            "Leakage_Risk": "high",
            "Valid_Use_Case": "diagnostic_only",
            "Thesis_Use": "discussion_limitations",
            "Notes": "",
        },
        {
            "Data_Slice_ID": "DS06",
            "Slice_Name": "Early sessions only",
            "Purpose": "Temporal drift checks and early-to-late transfer design.",
            "Subject_Scope": "all_subjects",
            "Session_Scope": "early_only",
            "Time_Window": "early_window",
            "Task_Scope": "pooled_all_tasks",
            "Modality_Scope": "pooled_all_modalities",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect",
            "Inclusion_Rule": "Use sessions pre-registered as early window.",
            "Exclusion_Rule": "Exclude late-window sessions.",
            "Class_Balance_Policy": "as_is",
            "Minimum_Samples_Rule": "Check class floor by temporal window.",
            "Leakage_Risk": "medium",
            "Valid_Use_Case": "robustness_support",
            "Thesis_Use": "supporting_robustness",
            "Notes": "",
        },
        {
            "Data_Slice_ID": "DS07",
            "Slice_Name": "Late sessions only",
            "Purpose": "Temporal drift checks and early-to-late transfer design.",
            "Subject_Scope": "all_subjects",
            "Session_Scope": "late_only",
            "Time_Window": "late_window",
            "Task_Scope": "pooled_all_tasks",
            "Modality_Scope": "pooled_all_modalities",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect",
            "Inclusion_Rule": "Use sessions pre-registered as late window.",
            "Exclusion_Rule": "Exclude early-window sessions.",
            "Class_Balance_Policy": "as_is",
            "Minimum_Samples_Rule": "Check class floor by temporal window.",
            "Leakage_Risk": "medium",
            "Valid_Use_Case": "robustness_support",
            "Thesis_Use": "supporting_robustness",
            "Notes": "",
        },
        {
            "Data_Slice_ID": "DS08",
            "Slice_Name": "Balanced subset only",
            "Purpose": "Imbalance-aware sensitivity diagnostics.",
            "Subject_Scope": "all_subjects",
            "Session_Scope": "all_sessions",
            "Time_Window": "full_study_window",
            "Task_Scope": "pooled_all_tasks",
            "Modality_Scope": "pooled_all_modalities",
            "Target_Type": "coarse_affect",
            "Label_Set": "coarse_affect",
            "Inclusion_Rule": "Only include rows passing balancing policy.",
            "Exclusion_Rule": "Exclude classes below predefined minimum count floor.",
            "Class_Balance_Policy": "balanced_subset",
            "Minimum_Samples_Rule": "Set explicit class floor in analysis config.",
            "Leakage_Risk": "low",
            "Valid_Use_Case": "robustness_support",
            "Thesis_Use": "supporting_robustness",
            "Notes": "Do not replace confirmatory as-is evidence with this slice.",
        },
    ]

    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(DATA_SELECTION_COLUMNS, start=1):
            ws.cell(r, c, row.get(name, ""))

    last = 41
    style_body(ws, 2, last, 1, len(DATA_SELECTION_COLUMNS))
    add_table(ws, "DataSelectionTable", f"A1:R{last}", style="TableStyleMedium3")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{last}"

    add_list_validation(
        ws, "=List_Subject_Scope", col_idx(DATA_SELECTION_COLUMNS, "Subject_Scope"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Session_Scope", col_idx(DATA_SELECTION_COLUMNS, "Session_Scope"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Task_Scope", col_idx(DATA_SELECTION_COLUMNS, "Task_Scope"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Modality_Scope", col_idx(DATA_SELECTION_COLUMNS, "Modality_Scope"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Target_Type", col_idx(DATA_SELECTION_COLUMNS, "Target_Type"), 2, 1000
    )
    add_list_validation(
        ws,
        "=List_Class_Balance_Policy",
        col_idx(DATA_SELECTION_COLUMNS, "Class_Balance_Policy"),
        2,
        1000,
    )
    add_list_validation(
        ws, "=List_Leakage_Risk_Level", col_idx(DATA_SELECTION_COLUMNS, "Leakage_Risk"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Use_Case", col_idx(DATA_SELECTION_COLUMNS, "Valid_Use_Case"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Thesis_Use_Tag", col_idx(DATA_SELECTION_COLUMNS, "Thesis_Use"), 2, 1000
    )

    set_widths(
        ws,
        {
            "A": 13,
            "B": 42,
            "C": 34,
            "D": 18,
            "E": 16,
            "F": 18,
            "G": 20,
            "H": 22,
            "I": 16,
            "J": 22,
            "K": 34,
            "L": 32,
            "M": 20,
            "N": 24,
            "O": 14,
            "P": 22,
            "Q": 22,
            "R": 30,
        },
    )

    add_dynamic_named_list(
        wb,
        "List_Data_Slice_ID",
        ws.title,
        col_idx(DATA_SELECTION_COLUMNS, "Data_Slice_ID"),
        2,
    )
    return last


def fill_grouping_strategy_map_sheet(ws, wb: Workbook) -> int:
    for i, h in enumerate(GROUPING_STRATEGY_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(GROUPING_STRATEGY_COLUMNS))

    rows = [
        {
            "Grouping_Strategy_ID": "GS01",
            "Strategy_Name": "within-subject LOSO-session",
            "Split_Family": "within_subject",
            "Train_Group_Unit": "subject_session",
            "Test_Group_Unit": "session",
            "Grouping_Level": "within_subject",
            "Train_Rule": "Per subject: train on all but one session.",
            "Test_Rule": "Per subject: test on held-out session.",
            "Leakage_Safeguard": "No held-out session rows in any train-side fit/tune stage.",
            "Suitable_For": "confirmatory_within_person",
            "Not_Suitable_For": "confirmatory_cross_person_transfer",
            "Interpretation_Boundary": "Supports within-person held-out-session claim only.",
            "Typical_Experiments": "E04,E16",
            "Thesis_Section": "Chapter 4 Main results",
            "Notes": "",
        },
        {
            "Grouping_Strategy_ID": "GS02",
            "Strategy_Name": "pooled-task within-subject LOSO-session",
            "Split_Family": "within_subject",
            "Train_Group_Unit": "subject_session",
            "Test_Group_Unit": "session",
            "Grouping_Level": "within_subject",
            "Train_Rule": "Within each subject, pool tasks and hold out one session.",
            "Test_Rule": "Evaluate on held-out session with pooled task scope.",
            "Leakage_Safeguard": "No test-session information in pooled train transforms.",
            "Suitable_For": "confirmatory_within_person",
            "Not_Suitable_For": "confirmatory_cross_person_transfer",
            "Interpretation_Boundary": "Within-person conclusion under pooled task policy.",
            "Typical_Experiments": "E02,E16",
            "Thesis_Section": "Chapter 3 Method choice",
            "Notes": "",
        },
        {
            "Grouping_Strategy_ID": "GS03",
            "Strategy_Name": "task-specific within-subject LOSO-session",
            "Split_Family": "within_subject",
            "Train_Group_Unit": "subject_session",
            "Test_Group_Unit": "session",
            "Grouping_Level": "task_level",
            "Train_Rule": "Within each subject/task, hold out one session.",
            "Test_Rule": "Test on same-task held-out session.",
            "Leakage_Safeguard": "Task-specific split manifests and train-only preprocessing.",
            "Suitable_For": "method_choice",
            "Not_Suitable_For": "confirmatory_cross_person_transfer",
            "Interpretation_Boundary": "Task-conditioned within-person evidence only.",
            "Typical_Experiments": "E02,E15",
            "Thesis_Section": "Chapter 3 Method choice",
            "Notes": "",
        },
        {
            "Grouping_Strategy_ID": "GS04",
            "Strategy_Name": "modality-specific within-subject LOSO-session",
            "Split_Family": "within_subject",
            "Train_Group_Unit": "subject_session",
            "Test_Group_Unit": "session",
            "Grouping_Level": "modality_level",
            "Train_Rule": "Within each subject/modality, hold out one session.",
            "Test_Rule": "Test on same-modality held-out session.",
            "Leakage_Safeguard": "Modality-specific fold manifests with strict train-only fit.",
            "Suitable_For": "method_choice",
            "Not_Suitable_For": "confirmatory_cross_person_transfer",
            "Interpretation_Boundary": "Modality-conditioned within-person evidence only.",
            "Typical_Experiments": "E03,E15",
            "Thesis_Section": "Chapter 3 Method choice",
            "Notes": "",
        },
        {
            "Grouping_Strategy_ID": "GS05",
            "Strategy_Name": "early-train late-test",
            "Split_Family": "temporal",
            "Train_Group_Unit": "session",
            "Test_Group_Unit": "session",
            "Grouping_Level": "session_level",
            "Train_Rule": "Train on pre-defined early sessions.",
            "Test_Rule": "Test on disjoint late sessions.",
            "Leakage_Safeguard": "Fixed temporal boundary and no late-session tuning.",
            "Suitable_For": "robustness_support",
            "Not_Suitable_For": "confirmatory_within_person",
            "Interpretation_Boundary": "Temporal transfer sensitivity, not the primary confirmatory claim.",
            "Typical_Experiments": "E04,E15",
            "Thesis_Section": "Chapter 4 Supporting robustness",
            "Notes": "",
        },
        {
            "Grouping_Strategy_ID": "GS06",
            "Strategy_Name": "frozen cross-person transfer A->B",
            "Split_Family": "cross_person",
            "Train_Group_Unit": "subject",
            "Test_Group_Unit": "subject",
            "Grouping_Level": "across_subject",
            "Train_Rule": "Fit on subject A only (frozen pipeline).",
            "Test_Rule": "Evaluate on subject B without refit/re-tune.",
            "Leakage_Safeguard": "No test-subject fitting; strict direction lock.",
            "Suitable_For": "confirmatory_cross_person_transfer",
            "Not_Suitable_For": "confirmatory_within_person",
            "Interpretation_Boundary": "Directional transfer under domain shift only.",
            "Typical_Experiments": "E05,E17",
            "Thesis_Section": "Chapter 4 Main results",
            "Notes": "",
        },
        {
            "Grouping_Strategy_ID": "GS07",
            "Strategy_Name": "frozen cross-person transfer B->A",
            "Split_Family": "cross_person",
            "Train_Group_Unit": "subject",
            "Test_Group_Unit": "subject",
            "Grouping_Level": "across_subject",
            "Train_Rule": "Fit on subject B only (frozen pipeline).",
            "Test_Rule": "Evaluate on subject A without refit/re-tune.",
            "Leakage_Safeguard": "No test-subject fitting; strict direction lock.",
            "Suitable_For": "confirmatory_cross_person_transfer",
            "Not_Suitable_For": "confirmatory_within_person",
            "Interpretation_Boundary": "Directional transfer under domain shift only.",
            "Typical_Experiments": "E05,E17",
            "Thesis_Section": "Chapter 4 Main results",
            "Notes": "",
        },
        {
            "Grouping_Strategy_ID": "GS08",
            "Strategy_Name": "weak split inflation demo",
            "Split_Family": "weak_split_demo",
            "Train_Group_Unit": "custom_group",
            "Test_Group_Unit": "custom_group",
            "Grouping_Level": "mixed_level",
            "Train_Rule": "Use intentionally weaker split variant for contrast.",
            "Test_Rule": "Evaluate inflation magnitude vs strict split.",
            "Leakage_Safeguard": "Tag as weak split and exclude from confirmatory claims.",
            "Suitable_For": "weak_split_demo_only",
            "Not_Suitable_For": "confirmatory_within_person",
            "Interpretation_Boundary": "Inflation demonstration only; non-confirmatory.",
            "Typical_Experiments": "E04",
            "Thesis_Section": "Chapter 3 Method choice",
            "Notes": "Permitted only as cautionary demonstration.",
        },
        {
            "Grouping_Strategy_ID": "GS09",
            "Strategy_Name": "task x modality diagnostic split",
            "Split_Family": "diagnostic",
            "Train_Group_Unit": "task_modality",
            "Test_Group_Unit": "task_modality",
            "Grouping_Level": "task_modality_level",
            "Train_Rule": "Train within selected task/modality intersections.",
            "Test_Rule": "Test on matched or held-out intersections per protocol.",
            "Leakage_Safeguard": "Pre-registered intersection mapping and no adaptive relabeling.",
            "Suitable_For": "diagnostic_only",
            "Not_Suitable_For": "confirmatory_within_person",
            "Interpretation_Boundary": "Diagnostic sensitivity only; not core evidence.",
            "Typical_Experiments": "E15",
            "Thesis_Section": "Chapter 5 Discussion",
            "Notes": "",
        },
    ]

    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(GROUPING_STRATEGY_COLUMNS, start=1):
            ws.cell(r, c, row.get(name, ""))

    last = 41
    style_body(ws, 2, last, 1, len(GROUPING_STRATEGY_COLUMNS))
    add_table(ws, "GroupingStrategyTable", f"A1:O{last}", style="TableStyleMedium4")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{last}"

    add_list_validation(
        ws, "=List_Split_Family", col_idx(GROUPING_STRATEGY_COLUMNS, "Split_Family"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Group_Unit", col_idx(GROUPING_STRATEGY_COLUMNS, "Train_Group_Unit"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Group_Unit", col_idx(GROUPING_STRATEGY_COLUMNS, "Test_Group_Unit"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Grouping_Level", col_idx(GROUPING_STRATEGY_COLUMNS, "Grouping_Level"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Use_Case", col_idx(GROUPING_STRATEGY_COLUMNS, "Suitable_For"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Use_Case", col_idx(GROUPING_STRATEGY_COLUMNS, "Not_Suitable_For"), 2, 1000
    )
    add_list_validation(
        ws,
        "=List_Reporting_Destination",
        col_idx(GROUPING_STRATEGY_COLUMNS, "Thesis_Section"),
        2,
        1000,
    )

    set_widths(
        ws,
        {
            "A": 16,
            "B": 36,
            "C": 16,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 30,
            "H": 30,
            "I": 32,
            "J": 20,
            "K": 22,
            "L": 36,
            "M": 20,
            "N": 24,
            "O": 26,
        },
    )

    add_dynamic_named_list(
        wb,
        "List_Grouping_Strategy_ID",
        ws.title,
        col_idx(GROUPING_STRATEGY_COLUMNS, "Grouping_Strategy_ID"),
        2,
    )
    return last


def fill_data_profile_sheet(ws) -> int:
    ws.merge_cells("A1:J1")
    ws["A1"] = "Data Profile and Imbalance Diagnostics"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws["A1"].alignment = Alignment(horizontal="left")

    ws["A3"] = "How to use"
    ws["A3"].font = Font(bold=True)
    ws["A3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("B3:J4")
    ws["B3"] = (
        "Paste exported counts into Count columns (or maintain manually). "
        "Formulas compute shares, imbalance flags, and sparsity flags."
    )
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    for rr in range(3, 5):
        for cc in range(1, 11):
            ws.cell(rr, cc).border = THIN

    ws["A6"] = "Threshold controls"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws["A7"] = "Sparse count threshold"
    ws["A8"] = "Mild imbalance ratio (max/min)"
    ws["A9"] = "Severe imbalance ratio (max/min)"
    ws["B7"] = 20
    ws["B8"] = 2
    ws["B9"] = 3
    for rr in range(7, 10):
        ws.cell(rr, 1).border = THIN
        ws.cell(rr, 2).border = THIN
        if rr % 2 == 0:
            ws.cell(rr, 1).fill = PatternFill("solid", fgColor=COL["zebra"])
            ws.cell(rr, 2).fill = PatternFill("solid", fgColor=COL["zebra"])

    ws["D6"] = "Profile summary"
    ws["D6"].font = Font(bold=True)
    ws["D6"].fill = PatternFill("solid", fgColor="EEF3FB")
    summary_labels = [
        ("Total_subject_samples", "=SUM(B13:B20)"),
        ("Total_session_samples", "=SUM(B25:B34)"),
        ("Total_task_samples", "=SUM(B39:B45)"),
        ("Total_modality_samples", "=SUM(B50:B55)"),
        ("Total_class_samples", "=SUM(B60:B65)"),
        ("Class_imbalance_ratio_max_min", '=IFERROR(MAX(B60:B65)/MIN(B60:B65),"")'),
        (
            "Class_imbalance_flag",
            '=IF(B60="","",IF(OR(MIN(B60:B65)=0,G12>$B$9),"SEVERE_IMBALANCE",IF(G12>$B$8,"MILD_IMBALANCE","BALANCED")))',
        ),
    ]
    for i, (label, formula) in enumerate(summary_labels, start=7):
        ws.cell(i, 4, label)
        ws.cell(i, 7, formula)
        ws.cell(i, 4).border = THIN
        ws.cell(i, 7).border = THIN
        if i % 2 == 0:
            ws.cell(i, 4).fill = PatternFill("solid", fgColor=COL["zebra"])
            ws.cell(i, 7).fill = PatternFill("solid", fgColor=COL["zebra"])

    def write_count_block(
        start_row: int,
        title: str,
        key_header: str,
        keys: list[str],
        table_name: str,
    ) -> int:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        ws.cell(start_row, 1, title)
        ws.cell(start_row, 1).font = Font(size=11, bold=True)
        ws.cell(start_row, 1).fill = PatternFill("solid", fgColor="EEF3FB")
        ws.cell(start_row, 1).border = THIN

        header_row = start_row + 1
        headers = [key_header, "Count", "Share_of_Block", "Flag", "Notes"]
        for c, h in enumerate(headers, start=1):
            ws.cell(header_row, c, h)
        style_header(ws, header_row, len(headers))

        data_start = header_row + 1
        data_end = data_start + len(keys) - 1
        for r, key in enumerate(keys, start=data_start):
            ws.cell(r, 1, key)
            ws.cell(r, 3, f"=IFERROR(B{r}/SUM($B${data_start}:$B${data_end}),0)")
            ws.cell(r, 4, f'=IF(B{r}="","",IF(B{r}<$B$7,"SPARSE","OK"))')
        style_body(ws, data_start, data_end, 1, 5)
        add_table(ws, table_name, f"A{header_row}:E{data_end}", style="TableStyleMedium6")
        return data_end

    write_count_block(
        11,
        "Counts by subject",
        "Subject_ID",
        ["sub-001", "sub-002", "sub-003", "sub-004", "sub-005", "sub-006", "sub-007", "sub-008"],
        "ProfileSubjectTable",
    )
    write_count_block(
        23,
        "Counts by session",
        "Session_ID",
        [
            "ses-01",
            "ses-02",
            "ses-03",
            "ses-04",
            "ses-05",
            "ses-06",
            "ses-07",
            "ses-08",
            "ses-09",
            "ses-10",
        ],
        "ProfileSessionTable",
    )
    write_count_block(
        37,
        "Counts by task",
        "Task_ID",
        [
            "pooled_all_tasks",
            "passive_only",
            "emo_only",
            "rating_only",
            "task_specific_other",
            "custom_task_1",
            "custom_task_2",
        ],
        "ProfileTaskTable",
    )
    write_count_block(
        48,
        "Counts by modality",
        "Modality_ID",
        [
            "pooled_all_modalities",
            "audio_only",
            "video_only",
            "audiovisual_only",
            "custom_modality_1",
            "custom_modality_2",
        ],
        "ProfileModalityTable",
    )
    write_count_block(
        58,
        "Counts by target/class",
        "Target_or_Class_ID",
        [
            "coarse_affect_neg",
            "coarse_affect_neu",
            "coarse_affect_pos",
            "binary_neg",
            "binary_pos",
            "custom_class_1",
        ],
        "ProfileTargetClassTable",
    )

    slice_start = 69
    ws.merge_cells(start_row=slice_start, start_column=1, end_row=slice_start, end_column=5)
    ws.cell(slice_start, 1, "Counts by Data_Slice_ID")
    ws.cell(slice_start, 1).font = Font(size=11, bold=True)
    ws.cell(slice_start, 1).fill = PatternFill("solid", fgColor="EEF3FB")
    ws.cell(slice_start, 1).border = THIN
    for c, h in enumerate(["Data_Slice_ID", "Count", "Share_of_Block", "Flag", "Notes"], start=1):
        ws.cell(slice_start + 1, c, h)
    style_header(ws, slice_start + 1, 5)

    data_start = slice_start + 2
    data_end = data_start + 9
    for r in range(data_start, data_end + 1):
        source_row = r - data_start + 2
        ws.cell(r, 1, f'=IFERROR(Data_Selection_Design!$A{source_row},"")')
        ws.cell(r, 3, f"=IFERROR(B{r}/SUM($B${data_start}:$B${data_end}),0)")
        ws.cell(r, 4, f'=IF(B{r}="","",IF(B{r}<$B$7,"SPARSE","OK"))')
    style_body(ws, data_start, data_end, 1, 5)
    add_table(
        ws, "ProfileDataSliceTable", f"A{slice_start + 1}:E{data_end}", style="TableStyleMedium6"
    )

    class_flag_range = "$B$60:$B$65"
    for r in range(60, 66):
        ws.cell(
            r,
            4,
            f'=IF(B{r}="","",IF(OR(MIN({class_flag_range})=0,MAX({class_flag_range})/MIN({class_flag_range})>$B$9),'
            f'"SEVERE_IMBALANCE",IF(MAX({class_flag_range})/MIN({class_flag_range})>$B$8,"MILD_IMBALANCE","BALANCED")))',
        )

    set_widths(
        ws,
        {"A": 26, "B": 14, "C": 16, "D": 20, "E": 38, "F": 4, "G": 24, "H": 14, "I": 12, "J": 12},
    )
    ws.freeze_panes = "A12"
    return data_end


def fill_run_log_sheet(ws) -> int:
    for i, h in enumerate(RUN_LOG_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(RUN_LOG_COLUMNS))
    last = 41
    style_body(ws, 2, last, 1, len(RUN_LOG_COLUMNS))
    end_col = get_column_letter(len(RUN_LOG_COLUMNS))
    add_table(ws, "RunLogTable", f"A1:{end_col}{last}", style="TableStyleMedium9")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{end_col}{last}"

    add_list_validation(
        ws, "=List_Data_Slice_ID", col_idx(RUN_LOG_COLUMNS, "Data_Slice_ID"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Grouping_Strategy_ID", col_idx(RUN_LOG_COLUMNS, "Grouping_Strategy_ID"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Transfer_Direction", col_idx(RUN_LOG_COLUMNS, "Transfer_Direction"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Session_Scope", col_idx(RUN_LOG_COLUMNS, "Session_Coverage"), 2, 1000
    )
    add_list_validation(ws, "=List_Task_Scope", col_idx(RUN_LOG_COLUMNS, "Task_Coverage"), 2, 1000)
    add_list_validation(
        ws, "=List_Modality_Scope", col_idx(RUN_LOG_COLUMNS, "Modality_Coverage"), 2, 1000
    )
    add_list_validation(ws, "=List_Run_Type", col_idx(RUN_LOG_COLUMNS, "Run_Type"), 2, 1000)
    add_list_validation(
        ws,
        "=List_Affects_Frozen_Pipeline",
        col_idx(RUN_LOG_COLUMNS, "Affects_Frozen_Pipeline"),
        2,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Eligible_for_Method_Decision",
        col_idx(RUN_LOG_COLUMNS, "Eligible_for_Method_Decision"),
        2,
        1000,
    )
    add_list_validation(
        ws, "=List_Imbalance_Status", col_idx(RUN_LOG_COLUMNS, "Imbalance_Status"), 2, 1000
    )
    add_list_validation(
        ws, "=List_Leakage_Check_Status", col_idx(RUN_LOG_COLUMNS, "Leakage_Check_Status"), 2, 1000
    )
    add_list_validation(ws, "=List_Reviewed", col_idx(RUN_LOG_COLUMNS, "Reviewed"), 2, 1000)
    add_list_validation(
        ws, "=List_Used_in_Thesis", col_idx(RUN_LOG_COLUMNS, "Used_in_Thesis"), 2, 1000
    )

    set_widths(
        ws,
        {
            "A": 18,
            "B": 13,
            "C": 12,
            "D": 20,
            "E": 22,
            "F": 12,
            "G": 16,
            "H": 22,
            "I": 30,
            "J": 12,
            "K": 16,
            "L": 24,
            "M": 22,
            "N": 22,
            "O": 18,
            "P": 18,
            "Q": 18,
            "R": 20,
            "S": 16,
            "T": 22,
            "U": 18,
            "V": 18,
            "W": 22,
            "X": 12,
            "Y": 20,
            "Z": 18,
            "AA": 20,
            "AB": 18,
            "AC": 18,
            "AD": 18,
            "AE": 30,
            "AF": 28,
            "AG": 34,
            "AH": 12,
            "AI": 14,
            "AJ": 36,
            "AK": 28,
        },
    )
    return last


def fill_decision_log_sheet(ws) -> int:
    for i, h in enumerate(DECISION_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(DECISION_COLUMNS))

    rows = [
        (
            "D01",
            "Final target choice",
            "Stage 1 - Target lock",
            "fine emotion; coarse_affect; binary valence-like",
            "E01,E02,E03",
            "Construct validity + class balance + learnability under strict split",
            "To be locked after method-choice evidence review",
            "",
            "",
            "Tradeoff between construct validity and learnability",
            "To be linked after background/method revision",
            "",
            "Open",
            "Chapter 3 method lock feeding Chapter 4 confirmatory analyses",
        ),
        (
            "D02",
            "Final split design",
            "Stage 2 - Split lock",
            "within_subject_loso_session; weaker alternatives for stress testing",
            "E04,E05",
            "Prefer strictest split matching claim while documenting weak-split inflation",
            "To be locked after split-strength evidence review",
            "",
            "",
            "Weak split inflation risk",
            "To be linked after method section split rationale update",
            "",
            "Open",
            "Defines primary inference logic and leakage controls",
        ),
        (
            "D03",
            "Final model family",
            "Stage 3 - Model lock",
            "ridge; logreg; linearsvc",
            "E06",
            "Simplicity + robustness + comparable performance",
            "To be locked after model comparison",
            "",
            "",
            "Complexity vs stability tradeoff",
            "To be linked after model-choice subsection revision",
            "",
            "Open",
            "Sets final confirmatory model family",
        ),
        (
            "D04",
            "Final weighting policy",
            "Stage 3 - Model lock",
            "No class weighting; balanced weighting",
            "E07",
            "Minority-class fairness gain without instability",
            "To be locked after weighting sensitivity",
            "",
            "",
            "Fairness-stability tension",
            "To be linked after class-imbalance discussion draft",
            "",
            "Open",
            "Controls fairness assumptions in final pipeline",
        ),
        (
            "D05",
            "Final tuning policy",
            "Stage 3 - Model lock",
            "Fixed explicit settings; light nested tuning",
            "E08",
            "Keep fixed unless clear stable gain from tuning",
            "To be locked after tuning strategy review",
            "",
            "",
            "Overfitting and complexity risk with tuning",
            "To be linked after reproducibility subsection update",
            "",
            "Open",
            "Determines model configuration policy",
        ),
        (
            "D06",
            "Final feature representation",
            "Stage 4 - Feature/preprocessing lock",
            "Whole-brain masked voxels; theory-justified ROI",
            "E10",
            "Lock on non-ROI representation evidence; use E09 as advisory/supporting ROI sensitivity context only.",
            "To be locked after feature-space evidence review",
            "",
            "",
            "Representation choice impacts validity and interpretability",
            "To be linked after feature-method section update",
            "",
            "Open",
            "Locks final feature representation",
        ),
        (
            "D07",
            "Final scaling policy",
            "Stage 4 - Feature/preprocessing lock",
            "No scaling; standard/robust scaling (train-only)",
            "E11",
            "Train-only scaling with best stability and simplicity",
            "To be locked after scaling sensitivity",
            "",
            "",
            "Scaling can alter class boundaries",
            "To be linked after preprocessing subsection revision",
            "",
            "Open",
            "Locks preprocessing policy",
        ),
        (
            "D08",
            "Use of external data",
            "Stage 7 - Exploratory extension",
            "No external data; external replication/portability",
            "E18,E19",
            "External work remains exploratory and non-core",
            "To be decided after confirmatory core completion",
            "",
            "",
            "Comparability and external validity limits",
            "To be linked after discussion/limitations revision",
            "",
            "Open",
            "Controls exploratory external scope and appendix use",
        ),
    ]

    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)

    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(DECISION_COLUMNS))
    add_table(ws, "DecisionLogTable", f"A1:N{last}", style="TableStyleMedium4")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{last}"
    add_list_validation(ws, "=List_Freeze_Status", 13, 2, 500, allow_blank=False)
    ws.conditional_formatting.add(
        f"M2:M{max(last, 200)}",
        FormulaRule(formula=['$M2="Locked"'], fill=PatternFill("solid", fgColor=COL["ok"])),
    )
    ws.conditional_formatting.add(
        f"M2:M{max(last, 200)}",
        FormulaRule(formula=['$M2="Open"'], fill=PatternFill("solid", fgColor=COL["open"])),
    )
    set_widths(
        ws,
        {
            "A": 12,
            "B": 24,
            "C": 30,
            "D": 36,
            "E": 18,
            "F": 34,
            "G": 34,
            "H": 22,
            "I": 22,
            "J": 28,
            "K": 34,
            "L": 12,
            "M": 14,
            "N": 30,
        },
    )
    return last


def fill_confirmatory_sheet(ws) -> int:
    for i, h in enumerate(CONFIRMATORY_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(CONFIRMATORY_COLUMNS))

    rows = [
        (
            "E12",
            "Permutation test experiment",
            "To be fixed after freeze lock",
            "Chance-distinguishability support",
            "YES",
        ),
        (
            "E13",
            "Trivial baseline experiment",
            "To be fixed after freeze lock",
            "Non-trivial predictive value support",
            "YES",
        ),
        (
            "E14",
            "Stability of explanation experiment",
            "To be fixed after freeze lock",
            "Model-behavior robustness support",
            "YES",
        ),
        (
            "E16",
            "Final within-person confirmatory analysis",
            "To be fixed after freeze lock",
            "Primary within-person claim",
            "YES",
        ),
        (
            "E17",
            "Final cross-person transfer analysis",
            "To be fixed after freeze lock",
            "Secondary transfer claim",
            "YES",
        ),
    ]
    for r, row in enumerate(rows, start=2):
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, row[1])
        ws.cell(r, 3, row[2])
        ws.cell(r, 4, row[3])
        ws.cell(
            r,
            5,
            f'=IFERROR(INDEX(Master_Experiments!$AA:$AA,MATCH($A{r},Master_Experiments!$A:$A,0)),"Planned")',
        )
        ws.cell(
            r,
            10,
            '=IF(AND(COUNTIFS(Decision_Log!$A:$A,"D01",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D02",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D03",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D04",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D05",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D06",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D07",Decision_Log!$M:$M,"Locked")>0),"YES","NO")',
        )
        ws.cell(
            r,
            6,
            f'=IF(AND($J{r}="YES",IFERROR(INDEX(Master_Experiments!$AF:$AF,MATCH($A{r},Master_Experiments!$A:$A,0)),"INCOMPLETE")="READY"),"YES","NO")',
        )
        ws.cell(
            r,
            7,
            f'=IF(IFERROR(INDEX(Master_Experiments!$AA:$AA,MATCH($A{r},Master_Experiments!$A:$A,0)),"Planned")="Completed","YES","NO")',
        )
        ws.cell(r, 8, row[4])
        ws.cell(r, 9, "NO")
        ws.cell(
            r,
            11,
            f'=IFERROR(INDEX(Master_Experiments!$AC:$AC,MATCH($A{r},Master_Experiments!$A:$A,0)),"")',
        )
        ws.cell(r, 12, "Planned")

    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(CONFIRMATORY_COLUMNS))
    add_table(ws, "ConfirmatoryTable", f"A1:L{last}", style="TableStyleMedium10")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{last}"
    add_list_validation(ws, "=List_Status", 5, 2, 500)
    add_list_validation(ws, "=List_Required_for_Main_Claim", 8, 2, 500)
    add_list_validation(ws, "=List_Ready_for_Chapter_4", 9, 2, 500)
    add_list_validation(ws, "=List_Status", 12, 2, 500)
    set_widths(
        ws,
        {
            "A": 13,
            "B": 34,
            "C": 28,
            "D": 34,
            "E": 14,
            "F": 18,
            "G": 12,
            "H": 20,
            "I": 18,
            "J": 22,
            "K": 32,
            "L": 20,
        },
    )
    ws.conditional_formatting.add(
        f"J2:J{max(last, 200)}",
        FormulaRule(formula=['$J2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])),
    )
    ws.conditional_formatting.add(
        f"F2:F{max(last, 200)}",
        FormulaRule(formula=['$F2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])),
    )
    ws.conditional_formatting.add(
        f"G2:G{max(last, 200)}",
        FormulaRule(formula=['$G2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])),
    )
    ws.conditional_formatting.add(
        f"I2:I{max(last, 200)}",
        FormulaRule(formula=['$I2="YES"'], fill=PatternFill("solid", fgColor=COL["ok"])),
    )
    return last


def fill_thesis_map_sheet(ws) -> int:
    for i, h in enumerate(THESIS_MAP_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(THESIS_MAP_COLUMNS))
    rows = [
        (
            "E01",
            "Chapter 3",
            "3.3 Target definition",
            "Target granularity method choice",
            "Supporting",
            "Method choice",
            "Locks target definition",
            "Reference Decision D01",
        ),
        (
            "E02",
            "Chapter 3",
            "3.3 Task strategy",
            "Task pooling method choice",
            "Supporting",
            "Method choice",
            "Task pooling policy",
            "",
        ),
        (
            "E03",
            "Chapter 3",
            "3.3 Modality strategy",
            "Modality pooling method choice",
            "Supporting",
            "Method choice",
            "Modality pooling policy",
            "",
        ),
        (
            "E04",
            "Chapter 3",
            "3.4 Split design",
            "Split-strength stress evidence",
            "Supporting",
            "Method choice",
            "Justifies strict split",
            "",
        ),
        (
            "E05",
            "Chapter 3",
            "3.4 Transfer protocol",
            "Cross-person framing",
            "Supporting",
            "Method choice",
            "Directional transfer design",
            "",
        ),
        (
            "E06",
            "Chapter 3",
            "3.5 Model family",
            "Model family selection",
            "Supporting",
            "Method choice",
            "Final model family lock",
            "",
        ),
        (
            "E07",
            "Chapter 3",
            "3.5 Weighting policy",
            "Class weighting decision",
            "Supporting",
            "Method choice",
            "Weighting policy lock",
            "",
        ),
        (
            "E08",
            "Chapter 3",
            "3.5 Tuning policy",
            "Hyperparameter policy",
            "Supporting",
            "Method choice",
            "Fixed vs tuning decision",
            "",
        ),
        (
            "E09",
            "Chapter 3",
            "3.2 Feature space",
            "Whole-brain vs simpler atlas-based ROI comparison",
            "Supporting",
            "Method choice",
            "Advisory ROI sensitivity check (not a lock dependency)",
            "",
        ),
        (
            "E10",
            "Chapter 3",
            "3.2 Dimensionality",
            "Reduction strategy decision",
            "Supporting",
            "Method choice",
            "Complexity/stability tradeoff",
            "",
        ),
        (
            "E11",
            "Chapter 3",
            "3.2 Scaling",
            "Scaling sensitivity decision",
            "Supporting",
            "Method choice",
            "Scaling policy lock",
            "",
        ),
        (
            "E12",
            "Chapter 4",
            "4.3 Robustness",
            "Permutation support",
            "Supporting",
            "Method application",
            "Beyond-chance support",
            "",
        ),
        (
            "E13",
            "Chapter 4",
            "4.3 Robustness",
            "Trivial baseline support",
            "Supporting",
            "Method application",
            "Non-trivial value support",
            "",
        ),
        (
            "E14",
            "Chapter 4",
            "4.3 Robustness",
            "Interpretability stability support",
            "Supporting",
            "Method application",
            "Model-behavior robustness",
            "No localization claim",
        ),
        (
            "E15",
            "Chapter 5",
            "5.2 Limitations",
            "Subset sensitivity discussion",
            "Supporting",
            "Discussion",
            "Sensitivity/limitations support",
            "",
        ),
        (
            "E16",
            "Chapter 4",
            "4.1 Main results",
            "Within-person confirmatory evidence",
            "Main",
            "Method application",
            "Primary thesis claim",
            "",
        ),
        (
            "E17",
            "Chapter 4",
            "4.2 Transfer results",
            "Cross-person transfer evidence",
            "Main",
            "Method application",
            "Secondary transfer claim",
            "",
        ),
        (
            "E18",
            "Appendix",
            "A. External exploratory",
            "External split replication",
            "Supporting",
            "Discussion",
            "Exploratory external consistency",
            "",
        ),
        (
            "E19",
            "Appendix",
            "A. External exploratory",
            "External model portability",
            "Supporting",
            "Discussion",
            "Exploratory portability",
            "",
        ),
    ]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(THESIS_MAP_COLUMNS))
    add_table(ws, "ThesisMapTable", f"A1:H{last}", style="TableStyleMedium6")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last}"
    set_widths(ws, {"A": 13, "B": 12, "C": 24, "D": 30, "E": 14, "F": 44, "G": 34, "H": 30})
    return last


def fill_claim_ledger_sheet(ws) -> int:
    for i, h in enumerate(CLAIM_LEDGER_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(CLAIM_LEDGER_COLUMNS))
    starter = [
        (
            "C01",
            "Primary within-person held-out-session decoding demonstrates meaningful generalization under locked pipeline.",
            "Confirmatory finding",
            "E16,E12,E13,E14",
            "open",
            "Chapter 4",
            "Interpret only within claim-matched split and current dataset scope.",
            "",
        ),
        (
            "C02",
            "Frozen directional cross-person transfer indicates cross-case portability under domain shift.",
            "Confirmatory finding",
            "E17,E12,E13",
            "open",
            "Chapter 4",
            "Not a population-level generalization claim.",
            "",
        ),
        (
            "C03",
            "Method lock decisions reduce leakage risk and interpretation drift.",
            "Method-choice",
            "E01,E04,E05,E06,E11",
            "partial",
            "Chapter 3",
            "Governance claim; not a performance claim. E09 is supporting/advisory sensitivity context only.",
            "",
        ),
    ]
    for r, row in enumerate(starter, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    last = 31
    style_body(ws, 2, last, 1, len(CLAIM_LEDGER_COLUMNS))
    add_table(ws, "ClaimLedgerTable", f"A1:H{last}", style="TableStyleMedium8")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last}"
    add_list_validation(ws, "=List_Claim_Type", 3, 2, 1000)
    add_list_validation(ws, "=List_Evidence_Status", 5, 2, 1000)
    add_list_validation(ws, "=List_Writing_Location", 6, 2, 1000)
    set_widths(ws, {"A": 12, "B": 48, "C": 20, "D": 24, "E": 14, "F": 14, "G": 38, "H": 28})
    return last


def fill_ai_usage_sheet(ws) -> int:
    for i, h in enumerate(AI_USAGE_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(AI_USAGE_COLUMNS))
    ws.cell(2, 1, "AI001")
    ws.cell(2, 4, "GPT-5.3-Codex")
    ws.cell(2, 5, "Workbook package revision and governance alignment")
    ws.cell(2, 8, "Not reviewed")
    ws.cell(2, 9, "No")
    last = 41
    style_body(ws, 2, last, 1, len(AI_USAGE_COLUMNS))
    add_table(ws, "AIUsageTable", f"A1:J{last}", style="TableStyleMedium5")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{last}"
    add_list_validation(ws, "=List_Human_Verification_Status", 8, 2, 1000)
    add_list_validation(ws, "=List_Used_in_Thesis", 9, 2, 1000)
    set_widths(
        ws,
        {"A": 12, "B": 12, "C": 24, "D": 18, "E": 32, "F": 32, "G": 30, "H": 24, "I": 14, "J": 28},
    )
    return last


def fill_ethics_sheet(ws) -> int:
    for i, h in enumerate(ETHICS_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(ETHICS_COLUMNS))
    starter = [
        (
            "EN01",
            "E16",
            "Interpretation limits",
            "Risk of over-claiming confirmatory evidence beyond split/domain scope",
            "Enforce interpretation boundaries in Claim_Ledger and Chapter 5 limitations",
            "Chapter 5",
            "Open",
            "",
        ),
        (
            "EN02",
            "D08",
            "AI/tool transparency",
            "Insufficient traceability of AI-assisted drafting decisions",
            "Maintain AI_Usage_Log with human verification status before thesis inclusion",
            "Chapter 2/Appendix",
            "Open",
            "",
        ),
    ]
    for r, row in enumerate(starter, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    last = 31
    style_body(ws, 2, last, 1, len(ETHICS_COLUMNS))
    add_table(ws, "EthicsTable", f"A1:H{last}", style="TableStyleMedium7")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last}"
    add_list_validation(ws, "=List_Ethics_Topic", 3, 2, 1000)
    add_list_validation(ws, "=List_Ethics_Status", 7, 2, 1000)
    set_widths(ws, {"A": 12, "B": 22, "C": 22, "D": 34, "E": 34, "F": 20, "G": 12, "H": 26})
    return last


def fill_dictionary_sheet(ws, wb: Workbook) -> None:
    ws["A1"] = "Controlled vocabularies (drives dropdown validation)"
    ws["A1"].font = Font(size=12, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COL["title_bg"])

    names = [
        "Category",
        "Evidential_Role",
        "Stage",
        "Priority",
        "Status",
        "Reporting_Destination",
        "Reviewed",
        "Used_in_Thesis",
        "Freeze_Status",
        "Run_Type",
        "Affects_Frozen_Pipeline",
        "Eligible_for_Method_Decision",
        "Claim_Type",
        "Evidence_Status",
        "Writing_Location",
        "Human_Verification_Status",
        "Ethics_Status",
        "Ethics_Topic",
        "Required_for_Main_Claim",
        "Ready_for_Chapter_4",
        "YesNo",
        "Subject_Scope",
        "Session_Scope",
        "Task_Scope",
        "Modality_Scope",
        "Class_Balance_Policy",
        "Split_Family",
        "Imbalance_Status",
        "Leakage_Check_Status",
        "Thesis_Use_Tag",
        "Use_Case",
        "Group_Unit",
        "Grouping_Level",
        "Transfer_Direction",
        "Leakage_Risk_Level",
        "Target_Type",
        "Execution_Section",
        "Reuse_Policy",
        "Search_Optimization_Mode",
        "Search_Parameter_Scope",
        "Study_Type",
        "Study_Intent",
        "Factor_Type",
        "Replication_Mode",
        "Replication_Strategy",
        "Blocking_Strategy",
        "Randomization_Strategy",
        "Random_Seed_Policy",
        "Block_Type",
        "Design_Cell_Status",
        "Aggregation_Level",
        "Uncertainty_Method",
        "Multiplicity_Handling",
        "Interaction_Reporting_Policy",
        "Study_Review_Disposition",
        "Study_Eligibility_Status",
    ]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(names))
    for c, name in enumerate(names, start=1):
        ws.cell(2, c, name)
        ws.cell(2, c).font = Font(color=COL["header_fg"], bold=True)
        ws.cell(2, c).fill = PatternFill("solid", fgColor=COL["header_bg"])
        items = VOCABS[name]
        for r, item in enumerate(items, start=3):
            ws.cell(r, c, item)
            ws.cell(r, c).border = THIN
            if r % 2 == 0:
                ws.cell(r, c).fill = PatternFill("solid", fgColor=COL["zebra"])
        add_named_list(wb, f"List_{name}", ws.title, c, 3, 2 + len(items))

    def_col = len(names) + 1
    term_col = def_col
    definition_col = def_col + 1
    ws.cell(1, term_col, "Term definitions")
    ws.cell(1, term_col).font = Font(size=12, bold=True)
    ws.cell(1, term_col).fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws.merge_cells(start_row=1, start_column=term_col, end_row=1, end_column=definition_col)
    ws.cell(2, term_col, "Term")
    ws.cell(2, definition_col, "Definition")
    for c in (term_col, definition_col):
        ws.cell(2, c).font = Font(color=COL["header_fg"], bold=True)
        ws.cell(2, c).fill = PatternFill("solid", fgColor=COL["header_bg"])
    for r, (term, definition) in enumerate(DEFINITIONS, start=3):
        ws.cell(r, term_col, term)
        ws.cell(r, definition_col, definition)
        ws.cell(r, term_col).border = THIN
        ws.cell(r, definition_col).border = THIN
        ws.cell(r, definition_col).alignment = Alignment(wrap_text=True, vertical="top")
        if r % 2 == 0:
            ws.cell(r, term_col).fill = PatternFill("solid", fgColor=COL["zebra"])
            ws.cell(r, definition_col).fill = PatternFill("solid", fgColor=COL["zebra"])

    for c in range(1, len(names) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.column_dimensions[get_column_letter(term_col)].width = 24
    ws.column_dimensions[get_column_letter(definition_col)].width = 84
    ws.freeze_panes = "A3"


def fill_dashboard_sheet(ws) -> None:
    ws.merge_cells("A1:N1")
    ws["A1"] = "Thesis Experiment Program Dashboard (v2)"
    ws["A1"].font = Font(size=15, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COL["title_bg"])

    ws["A3"] = "Program state summary"
    ws["A3"].font = Font(size=12, bold=True)
    ws["A3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("A3:B3")

    labels = [
        "Total number of experiments",
        "Completed experiments",
        "Critical experiments completed",
        "Open decisions",
        "Locked decisions",
        "Target locked?",
        "Split locked?",
        "Model locked?",
        "Feature/preprocessing locked?",
        "Confirmatory eligibility achieved?",
        "Confirmatory ready for Chapter 4?",
        "Experiments mapped to Chapter 4 main results",
        "Experiments mapped only to Appendix",
        "Claims with Evidence_Status = supported",
        "Claims with Evidence_Status = partial",
        "Claims with Evidence_Status = open",
        "AI log entries",
        "Open ethics/governance notes",
        "Runs flagged with imbalance warning",
        "Runs flagged with leakage warning/fail",
        "Distinct data slices used in Run_Log",
        "Distinct grouping strategies used in Run_Log",
        "Runs with Data_Slice_ID recorded",
        "Runs with Grouping_Strategy_ID recorded",
    ]
    formulas = [
        "=COUNTA(Master_Experiments!$A$2:$A$500)",
        '=COUNTIF(Master_Experiments!$AA$2:$AA$500,"Completed")',
        '=COUNTIFS(Master_Experiments!$F$2:$F$500,"Critical",Master_Experiments!$AA$2:$AA$500,"Completed")',
        '=COUNTIF(Decision_Log!$M$2:$M$500,"Open")',
        '=COUNTIF(Decision_Log!$M$2:$M$500,"Locked")',
        '=IF(COUNTIFS(Decision_Log!$A:$A,"D01",Decision_Log!$M:$M,"Locked")>0,"YES","NO")',
        '=IF(COUNTIFS(Decision_Log!$A:$A,"D02",Decision_Log!$M:$M,"Locked")>0,"YES","NO")',
        '=IF(AND(COUNTIFS(Decision_Log!$A:$A,"D03",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D04",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D05",Decision_Log!$M:$M,"Locked")>0),"YES","NO")',
        '=IF(AND(COUNTIFS(Decision_Log!$A:$A,"D06",Decision_Log!$M:$M,"Locked")>0,COUNTIFS(Decision_Log!$A:$A,"D07",Decision_Log!$M:$M,"Locked")>0),"YES","NO")',
        '=IF(COUNTIFS(Confirmatory_Set!$J:$J,"YES")=COUNTIFS(Confirmatory_Set!$A:$A,"<>"),"YES","NO")',
        '=IF(AND(B13="YES",COUNTIFS(Confirmatory_Set!$H:$H,"YES",Confirmatory_Set!$G:$G,"YES",Confirmatory_Set!$I:$I,"YES")=COUNTIFS(Confirmatory_Set!$H:$H,"YES")),"YES","NO")',
        '=COUNTIFS(Thesis_Map!$B$2:$B$500,"Chapter 4",Thesis_Map!$E$2:$E$500,"Main")',
        '=COUNTIF(Thesis_Map!$B$2:$B$500,"Appendix")',
        '=COUNTIF(Claim_Ledger!$E$2:$E$500,"supported")',
        '=COUNTIF(Claim_Ledger!$E$2:$E$500,"partial")',
        '=COUNTIF(Claim_Ledger!$E$2:$E$500,"open")',
        "=COUNTA(AI_Usage_Log!$A$2:$A$500)",
        '=COUNTIF(Ethics_Governance_Notes!$G$2:$G$500,"Open")',
        '=COUNTIFS(Run_Log!$Z$2:$Z$2000,"mild_imbalance")+COUNTIFS(Run_Log!$Z$2:$Z$2000,"severe_imbalance")',
        '=COUNTIFS(Run_Log!$AA$2:$AA$2000,"warning")+COUNTIFS(Run_Log!$AA$2:$AA$2000,"failed")',
        '=SUMPRODUCT((Run_Log!$F$2:$F$2000<>"")/COUNTIF(Run_Log!$F$2:$F$2000,Run_Log!$F$2:$F$2000&""))',
        '=SUMPRODUCT((Run_Log!$G$2:$G$2000<>"")/COUNTIF(Run_Log!$G$2:$G$2000,Run_Log!$G$2:$G$2000&""))',
        '=COUNTIF(Run_Log!$F$2:$F$2000,"<>")',
        '=COUNTIF(Run_Log!$G$2:$G$2000,"<>")',
    ]

    for i, (label, formula) in enumerate(zip(labels, formulas, strict=True), start=4):
        ws.cell(i, 1, label)
        ws.cell(i, 2, formula)
        ws.cell(i, 1).border = THIN
        ws.cell(i, 2).border = THIN
        if i % 2 == 0:
            ws.cell(i, 1).fill = PatternFill("solid", fgColor=COL["zebra"])
            ws.cell(i, 2).fill = PatternFill("solid", fgColor=COL["zebra"])

    ws["D3"] = "Experiments by category"
    ws["D3"].font = Font(size=12, bold=True)
    ws["D3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("D3:E3")
    for i, _ in enumerate(VOCABS["Category"], start=4):
        ws.cell(i, 4, f"=Dictionary_Validation!A{i - 1}")
        ws.cell(i, 5, f"=COUNTIF(Master_Experiments!$C:$C,D{i})")
        ws.cell(i, 4).border = THIN
        ws.cell(i, 5).border = THIN

    ws["G3"] = "Experiments by evidential role"
    ws["G3"].font = Font(size=12, bold=True)
    ws["G3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("G3:H3")
    for i, _ in enumerate(VOCABS["Evidential_Role"], start=4):
        ws.cell(i, 7, f"=Dictionary_Validation!B{i - 1}")
        ws.cell(i, 8, f"=COUNTIF(Master_Experiments!$D:$D,G{i})")
        ws.cell(i, 7).border = THIN
        ws.cell(i, 8).border = THIN
        ws.cell(i, 7).alignment = Alignment(wrap_text=True)

    ws["J3"] = "Experiments by data slice"
    ws["J3"].font = Font(size=12, bold=True)
    ws["J3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("J3:K3")
    for i in range(4, 16):
        src_row = i - 2
        ws.cell(i, 10, f'=IFERROR(Data_Selection_Design!$A{src_row},"")')
        ws.cell(
            i,
            11,
            f'=IF(J{i}="","",SUMPRODUCT((Run_Log!$F$2:$F$2000=J{i})*(Run_Log!$B$2:$B$2000<>"")/'
            f"IFERROR(COUNTIFS(Run_Log!$F$2:$F$2000,J{i},Run_Log!$B$2:$B$2000,Run_Log!$B$2:$B$2000),1)))",
        )
        ws.cell(i, 10).border = THIN
        ws.cell(i, 11).border = THIN

    ws["M3"] = "Experiments by grouping strategy"
    ws["M3"].font = Font(size=12, bold=True)
    ws["M3"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("M3:N3")
    for i in range(4, 16):
        src_row = i - 2
        ws.cell(i, 13, f'=IFERROR(Grouping_Strategy_Map!$A{src_row},"")')
        ws.cell(
            i,
            14,
            f'=IF(M{i}="","",SUMPRODUCT((Run_Log!$G$2:$G$2000=M{i})*(Run_Log!$B$2:$B$2000<>"")/'
            f"IFERROR(COUNTIFS(Run_Log!$G$2:$G$2000,M{i},Run_Log!$B$2:$B$2000,Run_Log!$B$2:$B$2000),1)))",
        )
        ws.cell(i, 13).border = THIN
        ws.cell(i, 14).border = THIN

    ws["D17"] = "Quick comparison counts (run-level)"
    ws["D17"].font = Font(size=12, bold=True)
    ws["D17"].fill = PatternFill("solid", fgColor="EEF3FB")
    ws.merge_cells("D17:E17")
    comp_labels = [
        "Pooled task runs",
        "Task-specific runs",
        "Pooled modality runs",
        "Modality-specific runs",
        "Within-person runs",
        "Cross-person runs",
        "Temporal split runs",
    ]
    comp_formulas = [
        '=COUNTIF(Run_Log!$Q$2:$Q$2000,"pooled_all_tasks")',
        '=COUNTIFS(Run_Log!$Q$2:$Q$2000,"<>",Run_Log!$Q$2:$Q$2000,"<>pooled_all_tasks")',
        '=COUNTIF(Run_Log!$R$2:$R$2000,"pooled_all_modalities")',
        '=COUNTIFS(Run_Log!$R$2:$R$2000,"<>",Run_Log!$R$2:$R$2000,"<>pooled_all_modalities")',
        '=SUMPRODUCT((Grouping_Strategy_Map!$C$2:$C$200="within_subject")*COUNTIF(Run_Log!$G$2:$G$2000,Grouping_Strategy_Map!$A$2:$A$200))',
        '=SUMPRODUCT((Grouping_Strategy_Map!$C$2:$C$200="cross_person")*COUNTIF(Run_Log!$G$2:$G$2000,Grouping_Strategy_Map!$A$2:$A$200))',
        '=SUMPRODUCT((Grouping_Strategy_Map!$C$2:$C$200="temporal")*COUNTIF(Run_Log!$G$2:$G$2000,Grouping_Strategy_Map!$A$2:$A$200))',
    ]
    for i, (label, formula) in enumerate(zip(comp_labels, comp_formulas, strict=True), start=18):
        ws.cell(i, 4, label)
        ws.cell(i, 5, formula)
        ws.cell(i, 4).border = THIN
        ws.cell(i, 5).border = THIN
        if i % 2 == 0:
            ws.cell(i, 4).fill = PatternFill("solid", fgColor=COL["zebra"])
            ws.cell(i, 5).fill = PatternFill("solid", fgColor=COL["zebra"])

    set_widths(
        ws,
        {
            "A": 46,
            "B": 16,
            "C": 3,
            "D": 30,
            "E": 12,
            "F": 3,
            "G": 36,
            "H": 12,
            "I": 3,
            "J": 24,
            "K": 12,
            "L": 3,
            "M": 28,
            "N": 12,
        },
    )
    ws.freeze_panes = "A4"
    ws.conditional_formatting.add(
        "B9:B14", FormulaRule(formula=['B9="YES"'], fill=PatternFill("solid", fgColor=COL["ok"]))
    )
    ws.conditional_formatting.add(
        "B9:B14", FormulaRule(formula=['B9="NO"'], fill=PatternFill("solid", fgColor=COL["bad"]))
    )
