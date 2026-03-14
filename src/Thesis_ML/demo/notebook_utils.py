from __future__ import annotations

import csv
import json
import os
import shutil
import sqlite3
import subprocess
import sys
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any


def _looks_like_repo_root(path: Path) -> bool:
    return (path / "pyproject.toml").exists() and (path / "src" / "Thesis_ML").exists()


def discover_project_root(
    *,
    start: Path | None = None,
    env_var: str = "THESIS_ML_PROJECT_ROOT",
) -> Path:
    """Resolve project root using optional env override and parent traversal."""
    override = os.getenv(env_var, "").strip()
    if override:
        candidate = Path(override).expanduser().resolve()
        if _looks_like_repo_root(candidate):
            return candidate
        raise FileNotFoundError(
            f"{env_var} is set but does not look like the repository root: {candidate}"
        )

    base = (start or Path.cwd()).resolve()
    for candidate in [base, *base.parents]:
        if _looks_like_repo_root(candidate):
            return candidate

    raise FileNotFoundError(
        f"Could not auto-detect repository root from '{base}'. Set {env_var} to the repo path."
    )


def resolve_user_path(
    value: str | Path | None,
    *,
    base: Path,
    default: str | Path,
) -> Path:
    """Resolve optional user override path against a base directory."""
    raw = str(value).strip() if value is not None else ""
    source = Path(raw) if raw else Path(default)
    path = source.expanduser()
    if not path.is_absolute():
        path = (base / path).resolve()
    return path


@dataclass(frozen=True)
class NotebookWorkspace:
    project_root: Path
    notebook_root: Path
    cache_dir: Path
    output_root: Path
    registry_dir: Path
    workbook_dir: Path
    workbook_output_dir: Path


def setup_notebook_workspace(
    notebook_slug: str,
    *,
    project_root: Path | None = None,
    create: bool = True,
) -> NotebookWorkspace:
    """Create and return notebook-local workspace directories."""
    root = project_root or discover_project_root()
    notebook_root = root / "outputs" / "notebook_demo" / notebook_slug
    workspace = NotebookWorkspace(
        project_root=root,
        notebook_root=notebook_root,
        cache_dir=notebook_root / "cache",
        output_root=notebook_root / "campaign_outputs",
        registry_dir=notebook_root / "registry",
        workbook_dir=notebook_root / "workbook",
        workbook_output_dir=notebook_root / "workbook_results",
    )
    if create:
        for directory in (
            workspace.notebook_root,
            workspace.cache_dir,
            workspace.output_root,
            workspace.registry_dir,
            workspace.workbook_dir,
            workspace.workbook_output_dir,
        ):
            directory.mkdir(parents=True, exist_ok=True)
    return workspace


def resolve_cli_command(*, preferred_executable: str, fallback_module: str) -> list[str]:
    """Use installed CLI executable when available; otherwise fallback to python -m module."""
    executable = shutil.which(preferred_executable)
    if executable:
        return [executable]
    return [sys.executable, "-m", fallback_module]


def run_command(
    cmd: Sequence[str],
    *,
    cwd: Path | None = None,
    check: bool = False,
) -> subprocess.CompletedProcess[str]:
    """Run a command, print command/stdout/stderr, and optionally raise on failure."""
    resolved = [str(part) for part in cmd]
    printable = " ".join(subprocess.list2cmdline([part]) for part in resolved)
    print(f"$ {printable}")
    result = subprocess.run(
        resolved,
        cwd=str(cwd) if cwd else None,
        capture_output=True,
        text=True,
        check=False,
    )
    print(f"[returncode] {result.returncode}")
    if result.stdout.strip():
        print("\n[stdout]\n" + result.stdout)
    if result.stderr.strip():
        print("\n[stderr]\n" + result.stderr)
    if check and result.returncode != 0:
        raise RuntimeError(
            f"Command failed with return code {result.returncode}: {' '.join(resolved)}"
        )
    return result


def parse_cli_key_value_lines(stdout_text: str) -> dict[str, str]:
    """Parse CLI lines formatted as '- key: value'."""
    result: dict[str, str] = {}
    for raw_line in stdout_text.splitlines():
        line = raw_line.strip()
        if not line.startswith("-") or ":" not in line:
            continue
        key, value = line[1:].split(":", 1)
        result[key.strip()] = value.strip()
    return result


def read_json_safe(path: Path | None) -> Any | None:
    """Read JSON from path; return None and print reason on missing/invalid file."""
    if path is None:
        return None
    target = Path(path)
    if not target.exists():
        print(f"[missing] {target}")
        return None
    try:
        return json.loads(target.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"[invalid json] {target}: {exc}")
        return None


def preview_csv_rows(path: Path, *, n: int = 10) -> list[list[str]] | None:
    """Return first n+1 rows (header + n rows) from CSV path."""
    target = Path(path)
    if not target.exists():
        print(f"[missing] {target}")
        return None
    rows: list[list[str]] = []
    with target.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        for idx, row in enumerate(reader):
            rows.append(row)
            if idx >= n:
                break
    return rows


def print_tree(root: Path, *, max_depth: int = 5, max_entries: int = 400) -> None:
    """Print a readable directory tree preview."""
    base = Path(root)
    print(f"Output tree under: {base}")
    if not base.exists():
        print("(missing)")
        return

    printed = 0
    total_seen = 0
    for item in sorted(base.rglob("*")):
        rel = item.relative_to(base)
        depth = len(rel.parts)
        if depth > max_depth:
            continue
        total_seen += 1
        if printed >= max_entries:
            continue
        indent = "  " * (depth - 1)
        suffix = "/" if item.is_dir() else ""
        print(f"{indent}- {rel.name}{suffix}")
        printed += 1
    if total_seen > printed:
        print(f"... truncated ({total_seen - printed} more entries at shown depth)")


def list_campaign_manifests(output_root: Path) -> list[Path]:
    """List campaign manifest files under campaign output root."""
    campaigns_dir = Path(output_root) / "campaigns"
    if not campaigns_dir.exists():
        return []
    return sorted(campaigns_dir.glob("*/campaign_manifest.json"), key=lambda p: p.stat().st_mtime)


def dedupe_paths(paths: Iterable[Path]) -> list[Path]:
    """Deduplicate paths by resolved string while preserving sorted order."""
    seen: set[str] = set()
    result: list[Path] = []
    for path in sorted(paths, key=lambda p: str(p)):
        key = str(path.resolve())
        if key in seen:
            continue
        seen.add(key)
        result.append(path)
    return result


def find_first_under(root: Path, filename: str) -> Path | None:
    """Return first matching filename under root, or None."""
    matches = sorted(Path(root).rglob(filename))
    return matches[0] if matches else None


def path_status(paths: Mapping[str, Path]) -> dict[str, dict[str, str | bool]]:
    """Return existence/type summary for notebook prerequisite display."""
    summary: dict[str, dict[str, str | bool]] = {}
    for key, path in paths.items():
        target = Path(path)
        summary[key] = {
            "path": str(target),
            "exists": target.exists(),
            "is_dir": target.is_dir(),
            "is_file": target.is_file(),
        }
    return summary


def execution_inputs_ready(*, index_csv: Path, data_root: Path) -> tuple[bool, list[str]]:
    """Check minimal data inputs needed for dry-run/full execution paths."""
    issues: list[str] = []
    if not Path(index_csv).exists():
        issues.append(f"Missing dataset index CSV: {index_csv}")
    if not Path(data_root).exists():
        issues.append(f"Missing data root: {data_root}")
    return (len(issues) == 0, issues)


def select_registry_experiment(
    registry_payload: Mapping[str, Any], experiment_id: str
) -> dict[str, Any]:
    """Select one experiment by ID from registry payload."""
    experiments = list(registry_payload.get("experiments", []))
    selected = next(
        (exp for exp in experiments if str(exp.get("experiment_id")) == str(experiment_id)),
        None,
    )
    if selected is None:
        available = ", ".join(sorted(str(exp.get("experiment_id")) for exp in experiments))
        raise ValueError(f"Experiment '{experiment_id}' not found. Available IDs: {available}")
    return dict(selected)


def create_single_experiment_registry(
    *,
    source_registry_path: Path,
    destination_registry_path: Path,
    experiment_id: str,
) -> dict[str, Any]:
    """Write notebook-local registry payload containing exactly one experiment."""
    payload = read_json_safe(source_registry_path)
    if not isinstance(payload, dict):
        raise ValueError(f"Invalid registry payload: {source_registry_path}")

    selected = select_registry_experiment(payload, experiment_id)
    output_payload = dict(payload)
    output_payload["experiments"] = [selected]
    destination_registry_path.parent.mkdir(parents=True, exist_ok=True)
    destination_registry_path.write_text(
        json.dumps(output_payload, indent=2) + "\n",
        encoding="utf-8",
    )
    return output_payload


def summarize_registry_experiment(experiment: Mapping[str, Any]) -> dict[str, Any]:
    """Build concise summary for notebook display before execution."""
    templates = list(experiment.get("variant_templates", []))
    supported_templates = [item for item in templates if bool(item.get("supported", True))]

    def collect_param_values(param_name: str) -> list[str]:
        values = set()
        for template in supported_templates:
            params = template.get("params", {})
            if not isinstance(params, dict):
                continue
            value = params.get(param_name)
            if value in (None, ""):
                continue
            values.add(str(value))
        return sorted(values)

    def as_text(values: list[str]) -> str:
        return ", ".join(values) if values else "<not explicitly fixed in params>"

    return {
        "experiment_id": experiment.get("experiment_id"),
        "title": experiment.get("title"),
        "stage": experiment.get("stage"),
        "manipulated_factor": experiment.get("manipulated_factor"),
        "primary_metric": experiment.get("primary_metric"),
        "secondary_metrics": experiment.get("secondary_metrics"),
        "target": as_text(collect_param_values("target")),
        "model": as_text(collect_param_values("model")),
        "cv": as_text(collect_param_values("cv")),
        "filter_task": as_text(collect_param_values("filter_task")),
        "filter_modality": as_text(collect_param_values("filter_modality")),
        "dataset_scope": experiment.get("dataset_scope"),
        "fixed_controls": experiment.get("fixed_controls"),
        "notes": experiment.get("notes"),
    }


def copy_file_to_workspace(
    *,
    source_path: Path,
    destination_path: Path,
    overwrite: bool = True,
) -> Path:
    """Copy source file to destination path in notebook workspace."""
    source = Path(source_path)
    destination = Path(destination_path)
    if not source.exists():
        raise FileNotFoundError(f"Source file not found: {source}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists() and not overwrite:
        return destination
    shutil.copy2(source, destination)
    return destination


def enable_single_workbook_experiment(
    *,
    workbook_path: Path,
    experiment_id: str | None = None,
    default_values: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Enable exactly one Experiment_Definitions row and disable all other experiment rows."""
    from openpyxl import load_workbook

    defaults = dict(default_values or {})
    defaults.setdefault("target", "coarse_affect")
    defaults.setdefault("cv", "within_subject_loso_session")
    defaults.setdefault("model", "ridge")

    workbook = load_workbook(workbook_path, data_only=False)
    ws = workbook["Experiment_Definitions"]
    header_map = {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value is not None and str(cell.value).strip()
    }
    required = ["experiment_id", "enabled", "target", "cv", "model"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError(f"Experiment_Definitions missing required columns: {missing}")

    rows: list[tuple[int, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=header_map["experiment_id"]).value
        exp_id = str(raw_id).strip() if raw_id is not None else ""
        if exp_id:
            rows.append((row_idx, exp_id))
    if not rows:
        raise ValueError("No experiment rows found in Experiment_Definitions.")

    selected_row: int | None = None
    selected_experiment_id: str | None = None
    if experiment_id:
        for row_idx, exp_id in rows:
            if exp_id == experiment_id:
                selected_row = row_idx
                selected_experiment_id = exp_id
                break
    if selected_row is None:
        selected_row, selected_experiment_id = rows[0]

    for row_idx, _ in rows:
        ws.cell(row=row_idx, column=header_map["enabled"]).value = (
            "Yes" if row_idx == selected_row else "No"
        )

    for field_name, default in defaults.items():
        col_idx = header_map.get(field_name)
        if col_idx is None:
            continue
        value = ws.cell(row=selected_row, column=col_idx).value
        text = str(value).strip() if value is not None else ""
        if not text:
            ws.cell(row=selected_row, column=col_idx).value = default

    workbook.save(workbook_path)
    payload = {
        key: ws.cell(row=selected_row, column=col_idx).value for key, col_idx in header_map.items()
    }
    enabled_count = sum(
        1
        for row_idx, _ in rows
        if str(ws.cell(row=row_idx, column=header_map["enabled"]).value).strip().lower()
        in {"yes", "y", "true", "1"}
    )
    return {
        "selected_experiment_id": selected_experiment_id,
        "selected_row": selected_row,
        "enabled_rows": enabled_count,
        "row_payload": payload,
    }


def summarize_workbook_sheet(
    workbook_path: Path,
    *,
    sheet_name: str,
    header_row: int,
    data_start_row: int | None = None,
    max_rows: int = 5,
) -> dict[str, Any]:
    """Return compact sheet summary with headers and first rows."""
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=False)
    ws = workbook[sheet_name]
    headers = [
        str(cell.value).strip()
        for cell in ws[header_row]
        if cell.value is not None and str(cell.value).strip()
    ]
    start_row = data_start_row if data_start_row is not None else header_row + 1
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + max_rows - 1, values_only=True):
        rows.append(list(row))
    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "headers": headers,
        "rows": rows,
    }


def preview_artifact_registry(db_path: Path, *, limit: int = 10) -> dict[str, Any]:
    """Return table list and recent artifact rows from SQLite artifact registry."""
    target = Path(db_path)
    if not target.exists():
        return {"exists": False, "path": str(target), "tables": [], "artifacts": []}

    with sqlite3.connect(target) as connection:
        tables = [
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
            )
        ]
        artifacts: list[tuple[Any, ...]] = []
        if "artifacts" in tables:
            artifacts = connection.execute(
                (
                    "SELECT artifact_id, artifact_type, run_id, status, path "
                    "FROM artifacts ORDER BY rowid DESC LIMIT ?"
                ),
                (int(limit),),
            ).fetchall()
    return {
        "exists": True,
        "path": str(target),
        "tables": tables,
        "artifacts": artifacts,
    }
