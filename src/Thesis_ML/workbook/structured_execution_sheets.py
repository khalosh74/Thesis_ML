from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from Thesis_ML.workbook.template_primitives import (
    COL,
    add_dynamic_named_list,
    add_list_validation,
    add_table,
    col_idx,
    set_widths,
    style_body,
    style_header,
)


def fill_simple_structured_sheet(
    ws: Worksheet,
    *,
    columns: list[str],
    table_name: str,
    title: str,
    width_map: dict[str, float],
    starter_rows: list[dict[str, str]] | None = None,
) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(size=12, bold=True)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws.cell(1, 1).alignment = Alignment(horizontal="left")
    for col_idx_value, header in enumerate(columns, start=1):
        ws.cell(2, col_idx_value, header)
    style_header(ws, 2, len(columns))

    if starter_rows:
        for row_idx, row in enumerate(starter_rows, start=3):
            for col_idx_value, name in enumerate(columns, start=1):
                ws.cell(row_idx, col_idx_value, row.get(name, ""))

    last = 61
    style_body(ws, 3, last, 1, len(columns))
    end_col = get_column_letter(len(columns))
    add_table(ws, table_name, f"A2:{end_col}{last}", style="TableStyleMedium6")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{end_col}{last}"
    set_widths(ws, width_map)
    return last


def fill_artifact_registry(ws: Worksheet, *, artifact_registry_columns: list[str]) -> int:
    starter = [
        {
            "artifact_id": "",
            "artifact_type": "",
            "run_id": "",
            "status": "",
            "created_at": "",
            "path": "",
            "upstream_artifact_ids": "",
            "config_hash": "",
            "code_ref": "",
            "notes": "Optional workbook mirror of machine registry for audit snapshots.",
        }
    ]
    return fill_simple_structured_sheet(
        ws=ws,
        columns=artifact_registry_columns,
        table_name="WorkbookArtifactRegistryTable",
        title="Workbook Artifact Registry Snapshot",
        width_map={
            "A": 32,
            "B": 20,
            "C": 24,
            "D": 14,
            "E": 24,
            "F": 40,
            "G": 32,
            "H": 24,
            "I": 20,
            "J": 34,
        },
        starter_rows=starter,
    )


def fill_fixed_configs(ws: Worksheet, *, fixed_configs_columns: list[str]) -> int:
    starter = [
        {
            "config_key": "default_target",
            "config_value": "coarse_affect",
            "scope": "global",
            "locked": "No",
            "owner": "Khaled",
            "last_updated": "",
            "notes": "Machine-readable defaults for execution templates.",
        },
        {
            "config_key": "default_model",
            "config_value": "ridge",
            "scope": "global",
            "locked": "No",
            "owner": "Khaled",
            "last_updated": "",
            "notes": "",
        },
    ]
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=fixed_configs_columns,
        table_name="FixedConfigsTable",
        title="Fixed Configurations and Locks",
        width_map={
            "A": 28,
            "B": 28,
            "C": 14,
            "D": 10,
            "E": 16,
            "F": 16,
            "G": 36,
        },
        starter_rows=starter,
    )
    add_list_validation(ws, "=List_YesNo", col_idx(fixed_configs_columns, "locked"), 3, 1000)
    return last


def fill_objectives(ws: Worksheet, *, objectives_columns: list[str]) -> int:
    starter = [
        {
            "objective_id": "OBJ01",
            "objective_text": "Lock target definition under leakage-aware evaluation.",
            "stage": "Stage 1 - Target lock",
            "linked_experiment_id": "E01",
            "primary_metric": "balanced_accuracy",
            "success_criterion": "Decision D01 locked with pre-registered rationale.",
            "status": "Planned",
            "notes": "",
        }
    ]
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=objectives_columns,
        table_name="ObjectivesTable",
        title="Program Objectives",
        width_map={
            "A": 12,
            "B": 44,
            "C": 28,
            "D": 18,
            "E": 20,
            "F": 34,
            "G": 14,
            "H": 30,
        },
        starter_rows=starter,
    )
    add_list_validation(ws, "=List_Stage", col_idx(objectives_columns, "stage"), 3, 1000)
    add_list_validation(
        ws,
        "=List_Experiment_ID",
        col_idx(objectives_columns, "linked_experiment_id"),
        3,
        1000,
    )
    add_list_validation(ws, "=List_Status", col_idx(objectives_columns, "status"), 3, 1000)
    return last


def fill_machine_status(ws: Worksheet, *, machine_status_columns: list[str]) -> int:
    starter = [
        {
            "machine_id": "M01",
            "hostname": "",
            "environment_name": "thesis-dev",
            "python_version": "",
            "gpu": "",
            "status": "Monitoring",
            "last_checked": "",
            "notes": "Track execution environment snapshots used for thesis runs.",
        }
    ]
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=machine_status_columns,
        table_name="MachineStatusTable",
        title="Machine Status and Environment",
        width_map={
            "A": 12,
            "B": 20,
            "C": 22,
            "D": 18,
            "E": 20,
            "F": 14,
            "G": 16,
            "H": 34,
        },
        starter_rows=starter,
    )
    add_list_validation(
        ws,
        "=List_Ethics_Status",
        col_idx(machine_status_columns, "status"),
        3,
        1000,
    )
    return last


def fill_study_design(
    ws: Worksheet,
    wb,
    *,
    study_design_columns: list[str],
) -> int:
    starter = [
        {
            "study_id": "S01",
            "study_name": "Example full factorial study",
            "enabled": "No",
            "study_type": "full_factorial",
            "intent": "exploratory",
            "question": "How do model and task filter settings change balanced accuracy?",
            "generalization_claim": "Within-dataset transfer across held-out sessions only.",
            "start_section": "dataset_selection",
            "end_section": "evaluation",
            "base_artifact_id": "",
            "primary_metric": "balanced_accuracy",
            "secondary_metrics": "macro_f1|accuracy",
            "cv_scheme": "within_subject_loso_session",
            "nested_cv": "No",
            "external_validation_planned": "No",
            "blocking_strategy": "none",
            "randomization_strategy": "fixed_order",
            "replication_mode": "fixed_repeats",
            "replication_strategy": "fixed_repeats",
            "num_repeats": "1",
            "random_seed_policy": "fixed",
            "stopping_rule": "Complete all planned design cells.",
            "notes": "",
        }
    ]
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=study_design_columns,
        table_name="StudyDesignTable",
        title="Study Design (Factorial and Custom Matrices)",
        width_map={
            "A": 12,
            "B": 30,
            "C": 10,
            "D": 20,
            "E": 14,
            "F": 36,
            "G": 34,
            "H": 18,
            "I": 16,
            "J": 24,
            "K": 18,
            "L": 20,
            "M": 24,
            "N": 10,
            "O": 14,
            "P": 18,
            "Q": 18,
            "R": 18,
            "S": 18,
            "T": 14,
            "U": 18,
            "V": 24,
            "W": 34,
        },
        starter_rows=starter,
    )
    add_list_validation(ws, "=List_YesNo", col_idx(study_design_columns, "enabled"), 3, 1000)
    add_list_validation(
        ws,
        "=List_Study_Type",
        col_idx(study_design_columns, "study_type"),
        3,
        1000,
    )
    add_list_validation(ws, "=List_Study_Intent", col_idx(study_design_columns, "intent"), 3, 1000)
    add_list_validation(
        ws,
        "=List_Execution_Section",
        col_idx(study_design_columns, "start_section"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Execution_Section",
        col_idx(study_design_columns, "end_section"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Replication_Mode",
        col_idx(study_design_columns, "replication_mode"),
        3,
        1000,
    )
    add_list_validation(ws, "=List_YesNo", col_idx(study_design_columns, "nested_cv"), 3, 1000)
    add_list_validation(
        ws,
        "=List_YesNo",
        col_idx(study_design_columns, "external_validation_planned"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Blocking_Strategy",
        col_idx(study_design_columns, "blocking_strategy"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Randomization_Strategy",
        col_idx(study_design_columns, "randomization_strategy"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Replication_Strategy",
        col_idx(study_design_columns, "replication_strategy"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Random_Seed_Policy",
        col_idx(study_design_columns, "random_seed_policy"),
        3,
        1000,
    )
    add_dynamic_named_list(
        wb,
        "List_Study_ID",
        ws.title,
        col_idx(study_design_columns, "study_id"),
        3,
    )
    return last


def fill_factors(ws: Worksheet, *, factors_columns: list[str]) -> int:
    starter = [
        {
            "study_id": "S01",
            "factor_name": "model",
            "section_name": "",
            "parameter_path": "model",
            "factor_type": "categorical",
            "levels": "ridge|logreg",
        },
        {
            "study_id": "S01",
            "factor_name": "filter_task",
            "section_name": "",
            "parameter_path": "filter_task",
            "factor_type": "categorical",
            "levels": "passive|emo",
        },
    ]
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=factors_columns,
        table_name="FactorsTable",
        title="Factor Definitions",
        width_map={
            "A": 12,
            "B": 20,
            "C": 18,
            "D": 26,
            "E": 14,
            "F": 34,
        },
        starter_rows=starter,
    )
    add_list_validation(ws, "=List_Study_ID", col_idx(factors_columns, "study_id"), 3, 1000)
    add_list_validation(ws, "=List_Factor_Type", col_idx(factors_columns, "factor_type"), 3, 1000)
    add_list_validation(
        ws,
        "=List_Execution_Section",
        col_idx(factors_columns, "section_name"),
        3,
        1000,
    )
    return last


def fill_study_rigor_checklist(
    ws: Worksheet,
    *,
    study_rigor_checklist_columns: list[str],
) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=study_rigor_checklist_columns,
        table_name="StudyRigorChecklistTable",
        title="Study Rigor Checklist",
        width_map={
            "A": 12,
            "B": 16,
            "C": 22,
            "D": 22,
            "E": 22,
            "F": 28,
            "G": 28,
            "H": 24,
            "I": 30,
            "J": 22,
            "K": 20,
            "L": 22,
            "M": 34,
        },
        starter_rows=None,
    )
    add_list_validation(
        ws,
        "=List_Study_ID",
        col_idx(study_rigor_checklist_columns, "study_id"),
        3,
        1000,
    )
    for column_name in (
        "leakage_risk_reviewed",
        "deployment_boundary_defined",
        "unit_of_analysis_defined",
        "data_hierarchy_defined",
        "reporting_checklist_completed",
        "risk_of_bias_reviewed",
        "confirmatory_lock_applied",
    ):
        add_list_validation(
            ws,
            "=List_YesNo",
            col_idx(study_rigor_checklist_columns, column_name),
            3,
            1000,
        )
    return last


def fill_analysis_plan(ws: Worksheet, *, analysis_plan_columns: list[str]) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=analysis_plan_columns,
        table_name="AnalysisPlanTable",
        title="Analysis Plan",
        width_map={
            "A": 12,
            "B": 28,
            "C": 28,
            "D": 18,
            "E": 18,
            "F": 24,
            "G": 24,
            "H": 32,
            "I": 34,
        },
        starter_rows=None,
    )
    add_list_validation(ws, "=List_Study_ID", col_idx(analysis_plan_columns, "study_id"), 3, 1000)
    add_list_validation(
        ws,
        "=List_Aggregation_Level",
        col_idx(analysis_plan_columns, "aggregation_level"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Uncertainty_Method",
        col_idx(analysis_plan_columns, "uncertainty_method"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Multiplicity_Handling",
        col_idx(analysis_plan_columns, "multiplicity_handling"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Interaction_Reporting_Policy",
        col_idx(analysis_plan_columns, "interaction_reporting_policy"),
        3,
        1000,
    )
    return last


def fill_study_review(ws: Worksheet, *, study_review_columns: list[str]) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=study_review_columns,
        table_name="StudyReviewTable",
        title="Study Review (Machine-managed pre-execution guardrails)",
        width_map={
            "A": 12,
            "B": 24,
            "C": 14,
            "D": 14,
            "E": 24,
            "F": 12,
            "G": 10,
            "H": 26,
            "I": 26,
            "J": 26,
            "K": 24,
            "L": 28,
            "M": 18,
            "N": 16,
            "O": 34,
            "P": 34,
            "Q": 30,
            "R": 12,
            "S": 12,
            "T": 12,
            "U": 20,
            "V": 24,
            "W": 20,
            "X": 10,
            "Y": 16,
            "Z": 18,
            "AA": 18,
            "AB": 18,
            "AC": 16,
            "AD": 12,
            "AE": 18,
            "AF": 20,
            "AG": 20,
            "AH": 12,
            "AI": 12,
            "AJ": 12,
            "AK": 12,
            "AL": 36,
        },
        starter_rows=None,
    )
    add_list_validation(ws, "=List_Study_ID", col_idx(study_review_columns, "study_id"), 3, 1000)
    add_list_validation(
        ws,
        "=List_Study_Review_Disposition",
        col_idx(study_review_columns, "execution_disposition"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Study_Eligibility_Status",
        col_idx(study_review_columns, "execution_eligibility_status"),
        3,
        1000,
    )
    return last


def fill_fixed_controls(ws: Worksheet, *, fixed_controls_columns: list[str]) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=fixed_controls_columns,
        table_name="DesignFixedControlsTable",
        title="Fixed Controls",
        width_map={
            "A": 12,
            "B": 28,
            "C": 28,
        },
        starter_rows=None,
    )
    add_list_validation(
        ws,
        "=List_Study_ID",
        col_idx(fixed_controls_columns, "study_id"),
        3,
        1000,
    )
    return last


def fill_constraints(ws: Worksheet, *, constraints_columns: list[str]) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=constraints_columns,
        table_name="DesignConstraintsTable",
        title="Invalid Combination Constraints",
        width_map={
            "A": 12,
            "B": 20,
            "C": 18,
            "D": 22,
            "E": 18,
            "F": 36,
        },
        starter_rows=None,
    )
    add_list_validation(ws, "=List_Study_ID", col_idx(constraints_columns, "study_id"), 3, 1000)
    return last


def fill_blocking_and_replication(
    ws: Worksheet,
    *,
    blocking_and_replication_columns: list[str],
) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=blocking_and_replication_columns,
        table_name="BlockingReplicationTable",
        title="Blocking and Replication",
        width_map={
            "A": 12,
            "B": 14,
            "C": 20,
            "D": 10,
            "E": 12,
        },
        starter_rows=None,
    )
    add_list_validation(
        ws,
        "=List_Block_Type",
        col_idx(blocking_and_replication_columns, "block_type"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Study_ID",
        col_idx(blocking_and_replication_columns, "study_id"),
        3,
        1000,
    )
    return last


def fill_generated_design_matrix(
    ws: Worksheet,
    *,
    generated_design_matrix_columns: list[str],
) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=generated_design_matrix_columns,
        table_name="GeneratedDesignMatrixTable",
        title="Generated Design Matrix (Machine-managed)",
        width_map={
            "A": 12,
            "B": 22,
            "C": 22,
            "D": 34,
            "E": 18,
            "F": 16,
            "G": 24,
            "H": 36,
            "I": 14,
        },
        starter_rows=None,
    )
    add_list_validation(
        ws,
        "=List_Study_ID",
        col_idx(generated_design_matrix_columns, "study_id"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Execution_Section",
        col_idx(generated_design_matrix_columns, "start_section"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Execution_Section",
        col_idx(generated_design_matrix_columns, "end_section"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Design_Cell_Status",
        col_idx(generated_design_matrix_columns, "status"),
        3,
        1000,
    )
    return last


def fill_effect_summaries(ws: Worksheet, *, effect_summaries_columns: list[str]) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=effect_summaries_columns,
        table_name="EffectSummariesTable",
        title="Effect Summaries (Descriptive)",
        width_map={
            "A": 12,
            "B": 22,
            "C": 22,
            "D": 18,
            "E": 24,
            "F": 20,
            "G": 20,
            "H": 20,
            "I": 18,
            "J": 34,
        },
        starter_rows=None,
    )
    add_list_validation(ws, "=List_Study_ID", col_idx(effect_summaries_columns, "study_id"), 3, 1000)
    return last


def fill_trial_results(ws: Worksheet, *, trial_results_columns: list[str]) -> int:
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=trial_results_columns,
        table_name="TrialResultsTable",
        title="Trial Results (Machine-managed)",
        width_map={
            "A": 22,
            "B": 14,
            "C": 24,
            "D": 14,
            "E": 22,
            "F": 18,
            "G": 34,
            "H": 34,
            "I": 28,
            "J": 30,
            "K": 12,
            "L": 22,
            "M": 34,
            "N": 34,
        },
        starter_rows=None,
    )
    add_list_validation(
        ws, "=List_Experiment_ID", col_idx(trial_results_columns, "experiment_id"), 3, 1000
    )
    add_list_validation(ws, "=List_Status", col_idx(trial_results_columns, "status"), 3, 1000)
    return last


def fill_summary_outputs(ws: Worksheet, *, summary_outputs_columns: list[str]) -> int:
    return fill_simple_structured_sheet(
        ws=ws,
        columns=summary_outputs_columns,
        table_name="SummaryOutputsTable",
        title="Machine Summary Outputs (Best Runs and Patterns)",
        width_map={
            "A": 18,
            "B": 22,
            "C": 20,
            "D": 18,
            "E": 26,
            "F": 14,
            "G": 18,
            "H": 18,
            "I": 14,
            "J": 24,
            "K": 18,
            "L": 20,
            "M": 34,
            "N": 36,
        },
        starter_rows=None,
    )


def fill_experiment_definitions(ws: Worksheet, *, experiment_definitions_columns: list[str]) -> int:
    for idx, header in enumerate(experiment_definitions_columns, start=1):
        ws.cell(1, idx, header)
    style_header(ws, 1, len(experiment_definitions_columns))

    seed_rows = [
        {
            "experiment_id": "E16",
            "enabled": "No",
            "start_section": "dataset_selection",
            "end_section": "evaluation",
            "base_artifact_id": "",
            "target": "coarse_affect",
            "cv": "within_subject_loso_session",
            "model": "ridge",
            "subject": "sub-001",
            "train_subject": "",
            "test_subject": "",
            "filter_task": "",
            "filter_modality": "",
            "reuse_policy": "auto",
            "search_space_id": "",
        },
        {
            "experiment_id": "E17",
            "enabled": "No",
            "start_section": "dataset_selection",
            "end_section": "evaluation",
            "base_artifact_id": "",
            "target": "coarse_affect",
            "cv": "frozen_cross_person_transfer",
            "model": "ridge",
            "subject": "",
            "train_subject": "sub-001",
            "test_subject": "sub-002",
            "filter_task": "",
            "filter_modality": "",
            "reuse_policy": "auto",
            "search_space_id": "",
        },
    ]
    for row_idx, row in enumerate(seed_rows, start=2):
        for col_idx_value, name in enumerate(experiment_definitions_columns, start=1):
            ws.cell(row_idx, col_idx_value, row.get(name, ""))

    last = 81
    style_body(ws, 2, last, 1, len(experiment_definitions_columns))
    add_table(ws, "ExperimentDefinitionsTable", f"A1:O{last}", style="TableStyleMedium2")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{last}"

    add_list_validation(
        ws,
        "=List_Experiment_ID",
        col_idx(experiment_definitions_columns, "experiment_id"),
        2,
        1000,
    )
    add_list_validation(
        ws,
        "=List_YesNo",
        col_idx(experiment_definitions_columns, "enabled"),
        2,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Execution_Section",
        col_idx(experiment_definitions_columns, "start_section"),
        2,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Execution_Section",
        col_idx(experiment_definitions_columns, "end_section"),
        2,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Reuse_Policy",
        col_idx(experiment_definitions_columns, "reuse_policy"),
        2,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Search_Space_ID",
        col_idx(experiment_definitions_columns, "search_space_id"),
        2,
        1000,
    )

    set_widths(
        ws,
        {
            "A": 14,
            "B": 10,
            "C": 20,
            "D": 20,
            "E": 28,
            "F": 18,
            "G": 28,
            "H": 14,
            "I": 12,
            "J": 14,
            "K": 14,
            "L": 16,
            "M": 18,
            "N": 18,
            "O": 16,
        },
    )
    return last


def fill_search_spaces(
    ws: Worksheet,
    wb,
    *,
    search_spaces_columns: list[str],
) -> int:
    seed_rows = [
        {
            "search_space_id": "SS01",
            "enabled": "No",
            "optimization_mode": "deterministic_grid",
            "parameter_name": "model",
            "parameter_values": "ridge|logreg|linearsvc",
            "parameter_scope": "parameter",
            "objective_metric": "balanced_accuracy",
            "max_trials": "",
            "notes": "Example model-family search.",
        },
        {
            "search_space_id": "SS01",
            "enabled": "No",
            "optimization_mode": "deterministic_grid",
            "parameter_name": "start_section",
            "parameter_values": "dataset_selection|feature_matrix_load",
            "parameter_scope": "segment",
            "objective_metric": "balanced_accuracy",
            "max_trials": "",
            "notes": "Example section-start search.",
        },
    ]
    last = fill_simple_structured_sheet(
        ws=ws,
        columns=search_spaces_columns,
        table_name="SearchSpacesTable",
        title="Search Space Definitions",
        width_map={
            "A": 16,
            "B": 10,
            "C": 20,
            "D": 24,
            "E": 34,
            "F": 14,
            "G": 20,
            "H": 12,
            "I": 32,
        },
        starter_rows=seed_rows,
    )
    add_list_validation(ws, "=List_YesNo", col_idx(search_spaces_columns, "enabled"), 3, 1000)
    add_list_validation(
        ws,
        "=List_Search_Optimization_Mode",
        col_idx(search_spaces_columns, "optimization_mode"),
        3,
        1000,
    )
    add_list_validation(
        ws,
        "=List_Search_Parameter_Scope",
        col_idx(search_spaces_columns, "parameter_scope"),
        3,
        1000,
    )
    add_dynamic_named_list(
        wb,
        "List_Search_Space_ID",
        ws.title,
        col_idx(search_spaces_columns, "search_space_id"),
        3,
    )
    return last


__all__ = [
    "fill_analysis_plan",
    "fill_artifact_registry",
    "fill_blocking_and_replication",
    "fill_constraints",
    "fill_effect_summaries",
    "fill_experiment_definitions",
    "fill_factors",
    "fill_fixed_configs",
    "fill_fixed_controls",
    "fill_generated_design_matrix",
    "fill_machine_status",
    "fill_objectives",
    "fill_search_spaces",
    "fill_study_design",
    "fill_study_review",
    "fill_study_rigor_checklist",
    "fill_summary_outputs",
    "fill_trial_results",
]
