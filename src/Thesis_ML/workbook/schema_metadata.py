from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from Thesis_ML.config.schema_versions import (
    WORKBOOK_SCHEMA_METADATA_HEADER_KEY,
    WORKBOOK_SCHEMA_METADATA_HEADER_VALUE,
    WORKBOOK_SCHEMA_METADATA_KEY_COLUMN,
    WORKBOOK_SCHEMA_METADATA_ROWS,
    WORKBOOK_SCHEMA_METADATA_START_ROW,
    WORKBOOK_SCHEMA_METADATA_VALUE_COLUMN,
)


def expected_schema_metadata() -> dict[str, str]:
    return {key: value for key, value in WORKBOOK_SCHEMA_METADATA_ROWS}


def write_schema_metadata(ws: Worksheet) -> None:
    ws.cell(
        row=WORKBOOK_SCHEMA_METADATA_START_ROW,
        column=WORKBOOK_SCHEMA_METADATA_KEY_COLUMN,
        value=WORKBOOK_SCHEMA_METADATA_HEADER_KEY,
    )
    ws.cell(
        row=WORKBOOK_SCHEMA_METADATA_START_ROW,
        column=WORKBOOK_SCHEMA_METADATA_VALUE_COLUMN,
        value=WORKBOOK_SCHEMA_METADATA_HEADER_VALUE,
    )
    for offset, (key, value) in enumerate(WORKBOOK_SCHEMA_METADATA_ROWS, start=1):
        ws.cell(
            row=WORKBOOK_SCHEMA_METADATA_START_ROW + offset,
            column=WORKBOOK_SCHEMA_METADATA_KEY_COLUMN,
            value=key,
        )
        ws.cell(
            row=WORKBOOK_SCHEMA_METADATA_START_ROW + offset,
            column=WORKBOOK_SCHEMA_METADATA_VALUE_COLUMN,
            value=value,
        )


def read_schema_metadata(ws: Worksheet) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row in range(
        WORKBOOK_SCHEMA_METADATA_START_ROW + 1,
        WORKBOOK_SCHEMA_METADATA_START_ROW + 50,
    ):
        key_value: Any = ws.cell(row=row, column=WORKBOOK_SCHEMA_METADATA_KEY_COLUMN).value
        if key_value is None:
            continue
        key = str(key_value).strip()
        if not key:
            continue
        value = ws.cell(row=row, column=WORKBOOK_SCHEMA_METADATA_VALUE_COLUMN).value
        metadata[key] = str(value).strip() if value is not None else ""
    return metadata
