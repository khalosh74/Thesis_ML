from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

COL = {
    "header_bg": "1F4E78",
    "header_fg": "FFFFFF",
    "title_bg": "D9E1F2",
    "zebra": "F8FAFC",
    "confirmatory": "E2F0D9",
    "exploratory": "FCE4D6",
    "critical": "FFF2CC",
    "dropped": "E7E6E6",
    "missing": "FBE5E7",
    "ok": "E2F0D9",
    "bad": "FBE5E7",
    "open": "FCE4D6",
}

THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=COL["header_bg"])
        cell.font = Font(color=COL["header_fg"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def style_body(ws, r1: int, r2: int, c1: int, c2: int, zebra: bool = True) -> None:
    fill = PatternFill("solid", fgColor=COL["zebra"])
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if zebra and r % 2 == 0:
                cell.fill = fill


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_list_validation(
    ws, formula: str, col: int, start: int, end: int, allow_blank: bool = True
) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    ws.add_data_validation(dv)
    letter = get_column_letter(col)
    dv.add(f"{letter}{start}:{letter}{end}")


def add_named_list(
    wb: Workbook, list_name: str, sheet_name: str, col: int, start: int, end: int
) -> None:
    letter = get_column_letter(col)
    wb.defined_names.add(
        DefinedName(
            name=list_name,
            attr_text=f"'{sheet_name}'!${letter}${start}:${letter}${end}",
        )
    )


def add_dynamic_named_list(
    wb: Workbook, list_name: str, sheet_name: str, col: int, start: int
) -> None:
    letter = get_column_letter(col)
    formula = (
        f"'{sheet_name}'!${letter}${start}:"
        f"INDEX('{sheet_name}'!${letter}:${letter},MATCH(\"zzz\",'{sheet_name}'!${letter}:${letter}))"
    )
    wb.defined_names.add(DefinedName(name=list_name, attr_text=formula))


def col_idx(columns: list[str], name: str) -> int:
    return columns.index(name) + 1


__all__ = [
    "COL",
    "THIN",
    "add_dynamic_named_list",
    "add_list_validation",
    "add_named_list",
    "add_table",
    "col_idx",
    "set_widths",
    "style_body",
    "style_header",
]
