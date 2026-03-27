from __future__ import annotations

import argparse
import json
import os
import subprocess
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from Thesis_ML.orchestration.workbook_compiler import compile_workbook_workbook
from Thesis_ML.workbook.builder import build_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]


def _timed(label: str, fn: Any) -> tuple[Any, float]:
    start = time.perf_counter()
    value = fn()
    return value, time.perf_counter() - start


def _run_subprocess(cmd: list[str], *, cwd: Path) -> dict[str, Any]:
    start = time.perf_counter()
    completed = subprocess.run(
        cmd,
        cwd=cwd,
        capture_output=True,
        text=True,
        env=os.environ.copy(),
    )
    return {
        "command": cmd,
        "returncode": completed.returncode,
        "duration_seconds": time.perf_counter() - start,
        "stdout": completed.stdout,
        "stderr": completed.stderr,
    }


def _enable_minimal_executable_row(workbook) -> None:
    master_ws = workbook["Master_Experiments"]
    master_headers = [master_ws.cell(1, col).value for col in range(1, master_ws.max_column + 1)]
    master_cols = {str(value): idx + 1 for idx, value in enumerate(master_headers) if value}
    master_ws.cell(2, master_cols["Experiment_ID"], "E16")
    master_ws.cell(2, master_cols["Short_Title"], "Performance smoke baseline row")
    master_ws.cell(2, master_cols["Stage"], "Stage 1 - Target lock")
    master_ws.cell(2, master_cols["Primary_Metric"], "balanced_accuracy")

    experiment_ws = workbook["Experiment_Definitions"]
    experiment_headers = [
        experiment_ws.cell(1, col).value for col in range(1, experiment_ws.max_column + 1)
    ]
    experiment_cols = {str(value): idx + 1 for idx, value in enumerate(experiment_headers) if value}
    experiment_ws.cell(2, experiment_cols["experiment_id"], "E16")
    experiment_ws.cell(2, experiment_cols["enabled"], "Yes")
    experiment_ws.cell(2, experiment_cols["start_section"], "dataset_selection")
    experiment_ws.cell(2, experiment_cols["end_section"], "evaluation")
    experiment_ws.cell(2, experiment_cols["target"], "coarse_affect")
    experiment_ws.cell(2, experiment_cols["cv"], "within_subject_loso_session")
    experiment_ws.cell(2, experiment_cols["model"], "ridge")
    experiment_ws.cell(2, experiment_cols["subject"], "sub-001")
    experiment_ws.cell(2, experiment_cols["reuse_policy"], "auto")


def _resolve_cli_command(command: str, module_fallback: str) -> list[str]:
    try:
        probe = subprocess.run(
            [command, "--help"],
            cwd=REPO_ROOT,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError:
        probe = None
    if probe is not None and probe.returncode == 0:
        return [command]
    return ["python", "-m", module_fallback]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Lightweight performance smoke for workbook + comparison/protocol dry-runs."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("outputs/performance/performance_smoke_summary.json"),
        help="Where to write the performance summary JSON.",
    )
    parser.add_argument(
        "--reports-root",
        type=Path,
        default=Path("outputs/performance/reports"),
        help="Temporary reports root for CLI dry-runs.",
    )
    args = parser.parse_args()

    output_path = (REPO_ROOT / args.output).resolve()
    reports_root = (REPO_ROOT / args.reports_root).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    reports_root.mkdir(parents=True, exist_ok=True)

    results: dict[str, Any] = {}

    workbook, workbook_build_seconds = _timed("build_workbook", build_workbook)
    results["workbook_build_seconds"] = round(workbook_build_seconds, 6)
    _enable_minimal_executable_row(workbook)

    workbook_path = reports_root / "performance_smoke_template.xlsx"
    save_start = time.perf_counter()
    workbook.save(workbook_path)
    results["workbook_save_seconds"] = round(time.perf_counter() - save_start, 6)

    loaded_workbook, workbook_load_seconds = _timed(
        "load_workbook", lambda: load_workbook(workbook_path, data_only=False)
    )
    results["workbook_load_seconds"] = round(workbook_load_seconds, 6)

    _, workbook_compile_seconds = _timed(
        "compile_workbook_workbook", lambda: compile_workbook_workbook(loaded_workbook)
    )
    results["workbook_compile_seconds"] = round(workbook_compile_seconds, 6)

    comparison_cmd_base = _resolve_cli_command(
        "thesisml-run-comparison",
        "Thesis_ML.cli.comparison_runner",
    )
    protocol_cmd_base = _resolve_cli_command(
        "thesisml-run-protocol",
        "Thesis_ML.cli.protocol_runner",
    )

    comparison_cmd = comparison_cmd_base + [
        "--comparison",
        # Canonical modeling-layer comparison workflow.
        "configs/comparisons/model_family_grouped_nested_comparison_v2.json",
        "--all-variants",
        "--reports-root",
        str(reports_root / "comparisons"),
        "--dry-run",
    ]
    protocol_cmd = protocol_cmd_base + [
        "--protocol",
        # Canonical modeling-layer protocol workflow (legacy frozen confirmatory stays in replay/audit paths).
        "configs/protocols/thesis_canonical_nested_v2.json",
        "--all-suites",
        "--reports-root",
        str(reports_root / "confirmatory"),
        "--dry-run",
    ]

    results["comparison_dry_run"] = _run_subprocess(comparison_cmd, cwd=REPO_ROOT)
    results["protocol_dry_run"] = _run_subprocess(protocol_cmd, cwd=REPO_ROOT)

    output_path.write_text(f"{json.dumps(results, indent=2)}\n", encoding="utf-8")

    print(f"Wrote performance smoke summary: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
