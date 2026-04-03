from __future__ import annotations

from openpyxl import Workbook

from Thesis_ML.workbook import template_constants
from Thesis_ML.workbook.governance_sheet_builders import (
    fill_claim_ledger_sheet,
    fill_decision_log_sheet,
    fill_thesis_map_sheet,
)


def _find_row(ws, key: str) -> int:
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "") == key:
            return row
    raise AssertionError(f"Could not find row with key={key}")


def test_template_constants_e09_wording_is_advisory_supporting() -> None:
    experiments = template_constants.build_experiments()
    by_id = {str(item["Experiment_ID"]): item for item in experiments}

    e09 = by_id["E09"]
    assert (
        e09["Decision_Supported"]
        == "Advisory/supporting ROI feature-space sensitivity check"
    )
    assert "simpler atlas-based ROI" in str(e09["Exact_Question"])
    assert "theory-justified ROIs" not in str(e09["Exact_Question"])

    # Ensure nearby unrelated experiment metadata remains stable.
    assert by_id["E10"]["Decision_Supported"] == "Feature reduction policy"


def test_governance_decision_log_d06_uses_non_roi_required_dependency() -> None:
    wb = Workbook()
    ws = wb.active
    fill_decision_log_sheet(ws)

    d06_row = _find_row(ws, "D06")
    assert ws.cell(d06_row, 5).value == "E10"
    rationale = str(ws.cell(d06_row, 6).value or "").lower()
    assert "e09" in rationale
    assert "advisory" in rationale


def test_governance_claim_ledger_c03_removes_e09_from_locked_core_dependencies() -> None:
    wb = Workbook()
    ws = wb.active
    fill_claim_ledger_sheet(ws)

    c03_row = _find_row(ws, "C03")
    assert ws.cell(c03_row, 4).value == "E01,E04,E05,E06,E11"
    boundary = str(ws.cell(c03_row, 7).value or "").lower()
    assert "e09" in boundary
    assert "advisory" in boundary


def test_governance_thesis_map_e09_is_marked_as_advisory() -> None:
    wb = Workbook()
    ws = wb.active
    fill_thesis_map_sheet(ws)

    e09_row = _find_row(ws, "E09")
    assert ws.cell(e09_row, 4).value == "Whole-brain vs simpler atlas-based ROI comparison"
    scope = str(ws.cell(e09_row, 7).value or "").lower()
    assert "advisory" in scope
    assert "not a lock dependency" in scope

