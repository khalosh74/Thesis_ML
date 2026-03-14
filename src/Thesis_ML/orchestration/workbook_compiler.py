from __future__ import annotations

import hashlib
import itertools
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from pydantic import ValidationError

from Thesis_ML.config.schema_versions import (
    SUPPORTED_WORKBOOK_SCHEMA_VERSIONS,
    WORKBOOK_SCHEMA_VERSION,
)
from Thesis_ML.orchestration.compiler import compile_registry_payload
from Thesis_ML.orchestration.contracts import (
    AnalysisPlanSpec,
    BlockingReplicationSpec,
    CompiledStudyManifest,
    ConstraintSpec,
    FactorSpec,
    FixedControlSpec,
    GeneratedDesignCell,
    ReusePolicy,
    SectionName,
    StudyDesignSpec,
    StudyIntent,
    StudyRigorChecklistSpec,
    StudyType,
)
from Thesis_ML.workbook.schema_metadata import read_schema_metadata

_REQUIRED_SHEETS = {"Master_Experiments", "Experiment_Definitions"}

_MASTER_REQUIRED_COLUMNS = {
    "Experiment_ID",
    "Short_Title",
    "Stage",
}

_EXPERIMENT_DEFINITION_REQUIRED_COLUMNS = [
    "experiment_id",
    "enabled",
    "start_section",
    "end_section",
    "base_artifact_id",
    "target",
    "cv",
    "model",
    "subject",
    "train_subject",
    "test_subject",
    "filter_task",
    "filter_modality",
    "reuse_policy",
]
_EXPERIMENT_DEFINITION_OPTIONAL_COLUMNS = {"search_space_id"}

_SEARCH_SPACES_COLUMNS = [
    "search_space_id",
    "enabled",
    "optimization_mode",
    "parameter_name",
    "parameter_values",
    "parameter_scope",
    "objective_metric",
    "max_trials",
    "notes",
]

_STUDY_DESIGN_REQUIRED_COLUMNS = [
    "study_id",
    "study_name",
    "enabled",
    "study_type",
    "intent",
    "question",
    "start_section",
    "end_section",
    "base_artifact_id",
    "primary_metric",
    "secondary_metrics",
    "cv_scheme",
    "replication_mode",
    "num_repeats",
    "random_seed_policy",
    "notes",
]

_STUDY_DESIGN_OPTIONAL_COLUMNS = [
    "generalization_claim",
    "nested_cv",
    "external_validation_planned",
    "blocking_strategy",
    "randomization_strategy",
    "replication_strategy",
    "stopping_rule",
]

_FACTOR_COLUMNS = [
    "study_id",
    "factor_name",
    "section_name",
    "parameter_path",
    "factor_type",
    "levels",
]

_FIXED_CONTROL_COLUMNS = [
    "study_id",
    "parameter_path",
    "value",
]

_CONSTRAINT_COLUMNS = [
    "study_id",
    "if_factor",
    "if_level",
    "disallow_factor",
    "disallow_level",
    "reason",
]

_BLOCKING_AND_REPLICATION_COLUMNS = [
    "study_id",
    "block_type",
    "block_value",
    "repeat_id",
    "seed",
]

_GENERATED_DESIGN_MATRIX_COLUMNS = [
    "study_id",
    "trial_id",
    "cell_id",
    "factor_settings_json",
    "start_section",
    "end_section",
    "base_artifact_id",
    "resolved_params_json",
    "status",
]

_STUDY_RIGOR_CHECKLIST_COLUMNS = [
    "study_id",
    "leakage_risk_reviewed",
    "deployment_boundary_defined",
    "unit_of_analysis_defined",
    "data_hierarchy_defined",
    "missing_data_plan",
    "class_imbalance_plan",
    "subgroup_plan",
    "fairness_or_applicability_notes",
    "reporting_checklist_completed",
    "risk_of_bias_reviewed",
    "confirmatory_lock_applied",
    "analysis_notes",
]

_ANALYSIS_PLAN_COLUMNS = [
    "study_id",
    "primary_contrast",
    "secondary_contrasts",
    "aggregation_level",
    "uncertainty_method",
    "multiplicity_handling",
    "interaction_reporting_policy",
    "interpretation_rules",
    "notes",
]

_CHECKLIST_YES_NO_COLUMNS = (
    "leakage_risk_reviewed",
    "deployment_boundary_defined",
    "unit_of_analysis_defined",
    "data_hierarchy_defined",
    "reporting_checklist_completed",
    "risk_of_bias_reviewed",
    "confirmatory_lock_applied",
)

_SECTION_ORDER: tuple[str, ...] = (
    SectionName.DATASET_SELECTION.value,
    SectionName.FEATURE_CACHE_BUILD.value,
    SectionName.FEATURE_MATRIX_LOAD.value,
    SectionName.SPATIAL_VALIDATION.value,
    SectionName.MODEL_FIT.value,
    SectionName.INTERPRETABILITY.value,
    SectionName.EVALUATION.value,
)
_SECTION_TO_INDEX = {name: idx for idx, name in enumerate(_SECTION_ORDER)}
_FEATURE_MATRIX_LOAD_INDEX = _SECTION_TO_INDEX[SectionName.FEATURE_MATRIX_LOAD.value]

_INT_PATTERN = re.compile(r"^[+-]?\d+$")
_FLOAT_PATTERN = re.compile(r"^[+-]?(?:\d+\.\d+|\d+\.\d*|\.\d+)$")


def _enum_text(value: Any) -> str:
    return str(value.value) if hasattr(value, "value") else str(value)


def _resolve_workbook_schema_version(workbook: Workbook) -> str:
    if "README" not in workbook.sheetnames:
        return WORKBOOK_SCHEMA_VERSION

    metadata = read_schema_metadata(workbook["README"])
    schema_version = metadata.get("workbook_schema_version", "").strip()
    if not schema_version:
        schema_version = WORKBOOK_SCHEMA_VERSION
    if schema_version not in SUPPORTED_WORKBOOK_SCHEMA_VERSIONS:
        supported = ", ".join(sorted(SUPPORTED_WORKBOOK_SCHEMA_VERSIONS))
        raise ValueError(
            f"Unsupported workbook schema version '{schema_version}'. "
            f"Supported versions: {supported}"
        )
    return schema_version


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_index_map(ws, header_row: int = 1) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        header = _normalize_text(cell.value)
        if header:
            mapping[header] = idx
    return mapping


def _read_cell(row: tuple[Any, ...], header_map: dict[str, int], column_name: str) -> Any:
    col_idx = header_map.get(column_name)
    if col_idx is None or col_idx <= 0 or col_idx > len(row):
        return None
    return row[col_idx - 1]


def _sheet_rows(
    workbook: Workbook,
    *,
    sheet_name: str,
    required_columns: list[str],
    header_row: int = 2,
    data_start_row: int = 3,
) -> tuple[dict[str, int], list[tuple[int, tuple[Any, ...]]]]:
    if sheet_name not in workbook.sheetnames:
        return {}, []
    ws = workbook[sheet_name]
    header_map = _header_index_map(ws, header_row=header_row)
    missing = [name for name in required_columns if name not in header_map]
    if missing:
        raise ValueError(f"{sheet_name} is missing required columns: {', '.join(missing)}")
    rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_index, row in enumerate(
        ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if all(_normalize_text(value) == "" for value in row):
            continue
        rows.append((row_index, row))
    return header_map, rows


def _parse_enabled(value: Any, *, row_index: int, sheet_name: str) -> bool:
    text = _normalize_text(value).lower()
    if text in {"", "no", "n", "false", "0"}:
        return False
    if text in {"yes", "y", "true", "1"}:
        return True
    raise ValueError(
        f"Invalid enabled value in {sheet_name} row {row_index}: '{value}'. Use Yes/No."
    )


def _parse_yes_no(
    value: Any,
    *,
    row_index: int,
    sheet_name: str,
    field_name: str,
) -> bool:
    text = _normalize_text(value).lower()
    if text in {"yes", "y", "true", "1"}:
        return True
    if text in {"no", "n", "false", "0"}:
        return False
    raise ValueError(
        f"Invalid {field_name} in {sheet_name} row {row_index}: '{value}'. Use Yes/No."
    )


def _parse_optional_yes_no(
    value: Any,
    *,
    row_index: int,
    sheet_name: str,
    field_name: str,
) -> bool | None:
    text = _normalize_text(value)
    if not text:
        return None
    return _parse_yes_no(value, row_index=row_index, sheet_name=sheet_name, field_name=field_name)


def _validated_section(value: Any, *, field_name: str, row_index: int, sheet_name: str) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    if text not in _SECTION_TO_INDEX:
        allowed = ", ".join(_SECTION_ORDER)
        raise ValueError(
            f"Invalid {field_name} in {sheet_name} row {row_index}: '{text}'. "
            f"Allowed values: {allowed}"
        )
    return text


def _validated_reuse_policy(value: Any, *, row_index: int) -> str:
    text = _normalize_text(value) or ReusePolicy.AUTO.value
    try:
        return ReusePolicy(text).value
    except ValueError as exc:
        allowed = ", ".join(policy.value for policy in ReusePolicy)
        raise ValueError(
            f"Invalid reuse_policy in Experiment_Definitions row {row_index}: '{text}'. "
            f"Allowed values: {allowed}"
        ) from exc


def _assert_base_artifact_usage(
    *,
    row_index: int,
    start_section: str,
    base_artifact_id: str,
    reuse_policy: str,
) -> None:
    start_idx = _SECTION_TO_INDEX[start_section]
    requires_base = start_idx >= _FEATURE_MATRIX_LOAD_INDEX

    if base_artifact_id and not requires_base:
        raise ValueError(
            "Invalid base artifact usage in Experiment_Definitions row "
            f"{row_index}: base_artifact_id is not allowed when start_section="
            f"'{start_section}'."
        )

    if (
        reuse_policy == ReusePolicy.REQUIRE_EXPLICIT_BASE.value
        and requires_base
        and not base_artifact_id
    ):
        raise ValueError(
            "Invalid base artifact usage in Experiment_Definitions row "
            f"{row_index}: reuse_policy='require_explicit_base' requires base_artifact_id "
            "when start_section is feature_matrix_load or later."
        )

    if reuse_policy == ReusePolicy.DISALLOW.value and requires_base:
        raise ValueError(
            "Invalid base artifact usage in Experiment_Definitions row "
            f"{row_index}: reuse_policy='disallow' is incompatible with start_section "
            "feature_matrix_load or later."
        )


def _parse_master_experiment_rows(ws) -> dict[str, dict[str, Any]]:
    header_map = _header_index_map(ws)
    missing_master = [name for name in _MASTER_REQUIRED_COLUMNS if name not in header_map]
    if missing_master:
        raise ValueError(
            "Master_Experiments is missing required columns: " + ", ".join(sorted(missing_master))
        )

    result: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        experiment_id = _normalize_text(_read_cell(row, header_map, "Experiment_ID"))
        if not experiment_id:
            continue
        result[experiment_id] = {
            "experiment_id": experiment_id,
            "title": _normalize_text(_read_cell(row, header_map, "Short_Title")) or experiment_id,
            "stage": _normalize_text(_read_cell(row, header_map, "Stage")) or "Unspecified stage",
            "manipulated_factor": _normalize_text(_read_cell(row, header_map, "Manipulated_Factor"))
            or None,
            "primary_metric": _normalize_text(_read_cell(row, header_map, "Primary_Metric"))
            or "balanced_accuracy",
        }
    return result


def _parse_experiment_definitions_rows(ws) -> list[dict[str, Any]]:
    header_map = _header_index_map(ws)
    missing_columns = [
        name for name in _EXPERIMENT_DEFINITION_REQUIRED_COLUMNS if name not in header_map
    ]
    if missing_columns:
        raise ValueError(
            "Experiment_Definitions is missing required columns: " + ", ".join(missing_columns)
        )

    compiled_trials: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        experiment_id = _normalize_text(_read_cell(row, header_map, "experiment_id"))
        if not experiment_id:
            continue
        enabled = _parse_enabled(
            _read_cell(row, header_map, "enabled"),
            row_index=row_index,
            sheet_name="Experiment_Definitions",
        )
        if not enabled:
            continue

        target = _normalize_text(_read_cell(row, header_map, "target"))
        cv = _normalize_text(_read_cell(row, header_map, "cv"))
        model = _normalize_text(_read_cell(row, header_map, "model"))
        missing_required = [
            name for name, value in (("target", target), ("cv", cv), ("model", model)) if not value
        ]
        if missing_required:
            raise ValueError(
                "Experiment_Definitions row "
                f"{row_index} is missing required values: {', '.join(missing_required)}"
            )

        start_section = _validated_section(
            _read_cell(row, header_map, "start_section"),
            field_name="start_section",
            row_index=row_index,
            sheet_name="Experiment_Definitions",
        ) or SectionName.DATASET_SELECTION.value
        end_section = _validated_section(
            _read_cell(row, header_map, "end_section"),
            field_name="end_section",
            row_index=row_index,
            sheet_name="Experiment_Definitions",
        ) or SectionName.EVALUATION.value
        if _SECTION_TO_INDEX[start_section] > _SECTION_TO_INDEX[end_section]:
            raise ValueError(
                "Invalid section range in Experiment_Definitions row "
                f"{row_index}: start_section='{start_section}' is after end_section='{end_section}'."
            )

        base_artifact_id = _normalize_text(_read_cell(row, header_map, "base_artifact_id")) or None
        reuse_policy = _validated_reuse_policy(
            _read_cell(row, header_map, "reuse_policy"),
            row_index=row_index,
        )
        _assert_base_artifact_usage(
            row_index=row_index,
            start_section=start_section,
            base_artifact_id=base_artifact_id or "",
            reuse_policy=reuse_policy,
        )

        params: dict[str, Any] = {
            "target": target,
            "cv": cv,
            "model": model,
        }
        for key in (
            "subject",
            "train_subject",
            "test_subject",
            "filter_task",
            "filter_modality",
        ):
            value = _normalize_text(_read_cell(row, header_map, key))
            if value:
                params[key] = value

        compiled_trials.append(
            {
                "experiment_id": experiment_id,
                "template_id": f"wb_row_{row_index:03d}",
                "supported": True,
                "params": params,
                "expand": {},
                "sections": list(_SECTION_ORDER),
                "artifacts": [],
                "start_section": start_section,
                "end_section": end_section,
                "base_artifact_id": base_artifact_id,
                "reuse_policy": reuse_policy,
                "search_space_id": (
                    _normalize_text(_read_cell(row, header_map, "search_space_id")) or None
                ),
            }
        )
    return compiled_trials


def _parse_search_values(raw_text: str) -> list[Any]:
    text = raw_text.strip()
    if not text:
        return []
    values: list[Any] = []
    for token in text.split("|"):
        item = token.strip()
        if not item:
            continue
        lowered = item.lower()
        if lowered == "true":
            values.append(True)
            continue
        if lowered == "false":
            values.append(False)
            continue
        try:
            if "." in item:
                values.append(float(item))
            else:
                values.append(int(item))
            continue
        except ValueError:
            values.append(item)
    return values


def _parse_search_spaces_rows(workbook: Workbook) -> list[dict[str, Any]]:
    if "Search_Spaces" not in workbook.sheetnames:
        return []
    ws = workbook["Search_Spaces"]
    header_map = _header_index_map(ws, header_row=2)
    missing = [name for name in _SEARCH_SPACES_COLUMNS if name not in header_map]
    if missing:
        raise ValueError("Search_Spaces is missing required columns: " + ", ".join(missing))

    grouped: dict[str, dict[str, Any]] = {}
    for row_index, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        search_space_id = _normalize_text(_read_cell(row, header_map, "search_space_id"))
        if not search_space_id:
            continue
        enabled = _parse_enabled(
            _read_cell(row, header_map, "enabled"),
            row_index=row_index,
            sheet_name="Search_Spaces",
        )
        if not enabled:
            continue
        parameter_name = _normalize_text(_read_cell(row, header_map, "parameter_name"))
        raw_values = _normalize_text(_read_cell(row, header_map, "parameter_values"))
        if not parameter_name:
            raise ValueError(
                f"Search_Spaces row {row_index} is missing parameter_name for '{search_space_id}'."
            )
        values = _parse_search_values(raw_values)
        if not values:
            raise ValueError(
                f"Search_Spaces row {row_index} has empty parameter_values for '{search_space_id}'."
            )

        max_trials_raw = _read_cell(row, header_map, "max_trials")
        space = grouped.setdefault(
            search_space_id,
            {
                "search_space_id": search_space_id,
                "enabled": True,
                "optimization_mode": _normalize_text(
                    _read_cell(row, header_map, "optimization_mode")
                )
                or "deterministic_grid",
                "objective_metric": _normalize_text(_read_cell(row, header_map, "objective_metric"))
                or "balanced_accuracy",
                "max_trials": int(max_trials_raw) if max_trials_raw not in (None, "") else None,
                "notes": _normalize_text(_read_cell(row, header_map, "notes")) or None,
                "dimensions": [],
            },
        )
        space["dimensions"].append(
            {
                "parameter_name": parameter_name,
                "values": values,
                "parameter_scope": _normalize_text(_read_cell(row, header_map, "parameter_scope"))
                or "parameter",
            }
        )

    return [grouped[key] for key in sorted(grouped)]


def _parse_scalar_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, float, dict, list)):
        return value
    text = _normalize_text(value)
    if not text:
        return ""
    lowered = text.lower()
    if lowered in {"true", "yes"}:
        return True
    if lowered in {"false", "no"}:
        return False
    if _INT_PATTERN.match(text):
        try:
            return int(text)
        except ValueError:
            pass
    if _FLOAT_PATTERN.match(text):
        try:
            return float(text)
        except ValueError:
            pass
    if text.startswith("{") or text.startswith("["):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    return text


def _parse_levels(value: Any, *, row_index: int, factor_name: str, study_id: str) -> list[Any]:
    if isinstance(value, list):
        levels = list(value)
    else:
        text = _normalize_text(value)
        if not text:
            levels = []
        elif text.startswith("["):
            try:
                decoded = json.loads(text)
            except json.JSONDecodeError as exc:
                raise ValueError(
                    f"Factors row {row_index} has invalid JSON levels for '{factor_name}' "
                    f"in study '{study_id}': {exc.msg}"
                ) from exc
            if not isinstance(decoded, list):
                raise ValueError(
                    f"Factors row {row_index} levels JSON must be a list for factor "
                    f"'{factor_name}' in study '{study_id}'."
                )
            levels = list(decoded)
        else:
            levels = [_parse_scalar_value(token) for token in text.split("|") if token.strip()]
    if not levels:
        raise ValueError(
            f"Factors row {row_index} defines no levels for factor '{factor_name}' "
            f"in study '{study_id}'."
        )
    return levels


def _parse_json_object(
    value: Any,
    *,
    row_index: int,
    sheet_name: str,
    column_name: str,
) -> dict[str, Any]:
    text = _normalize_text(value)
    if not text:
        return {}
    try:
        decoded = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"{sheet_name} row {row_index} has invalid JSON in {column_name}: {exc.msg}"
        ) from exc
    if not isinstance(decoded, dict):
        raise ValueError(
            f"{sheet_name} row {row_index} must use JSON object for {column_name}."
        )
    return decoded


def _stable_suffix(payload: dict[str, Any]) -> str:
    canonical = json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha1(canonical.encode("utf-8")).hexdigest()[:10]


def _normalize_level(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    return str(value).strip().lower()


def _is_disallowed_combination(
    factor_settings: dict[str, Any],
    constraints: list[ConstraintSpec],
) -> bool:
    for constraint in constraints:
        if _normalize_level(factor_settings.get(constraint.if_factor)) != _normalize_level(
            constraint.if_level
        ):
            continue
        if _normalize_level(factor_settings.get(constraint.disallow_factor)) == _normalize_level(
            constraint.disallow_level
        ):
            return True
    return False


def _fixed_controls_map(study: StudyDesignSpec) -> dict[str, Any]:
    controls: dict[str, Any] = {}
    for control in study.fixed_controls:
        controls[control.parameter_path] = control.value
    return controls


def _resolved_params_for_cell(
    *,
    study: StudyDesignSpec,
    factor_settings: dict[str, Any],
    base_params: dict[str, Any] | None = None,
) -> dict[str, Any]:
    params = dict(_fixed_controls_map(study))
    if base_params:
        params.update(base_params)
    factor_map = {factor.factor_name: factor for factor in study.factors}
    for factor_name, level in factor_settings.items():
        factor = factor_map.get(factor_name)
        if factor is None:
            params[factor_name] = level
        else:
            params[factor.parameter_path] = level
    if study.cv_scheme and not str(params.get("cv", "")).strip():
        params["cv"] = study.cv_scheme
    return params


def _repeat_plan_for_study(study: StudyDesignSpec) -> list[dict[str, Any]]:
    if study.blocking_replication:
        if study.num_repeats != len(study.blocking_replication):
            raise ValueError(
                f"Study '{study.study_id}' has num_repeats={study.num_repeats} but "
                f"{len(study.blocking_replication)} rows in Blocking_and_Replication."
            )
        return [
            {
                "repeat_id": int(item.repeat_id),
                "seed": item.seed,
                "block_type": item.block_type,
                "block_value": item.block_value,
            }
            for item in study.blocking_replication
        ]
    return [
        {"repeat_id": idx, "seed": None, "block_type": "none", "block_value": None}
        for idx in range(1, int(study.num_repeats) + 1)
    ]


def _seed_for_trial(
    *,
    study: StudyDesignSpec,
    trial_id: str,
    repeat_id: int,
    explicit_seed: int | None,
) -> int | None:
    if explicit_seed is not None:
        return int(explicit_seed)
    policy = str(study.random_seed_policy or "fixed").strip().lower()
    if policy == "fixed":
        return None
    if policy == "per_repeat":
        return int(repeat_id)
    if policy == "per_cell":
        return int(_stable_suffix({"study_id": study.study_id, "trial_id": trial_id}), 16)
    raise ValueError(
        f"Unsupported random_seed_policy='{study.random_seed_policy}' in study '{study.study_id}'."
    )


def _build_factorial_cells(
    study: StudyDesignSpec,
) -> list[GeneratedDesignCell]:
    if study.study_type == StudyType.FRACTIONAL_FACTORIAL:
        raise ValueError(
            f"Study '{study.study_id}' uses study_type='fractional_factorial', which is "
            "not supported in this release. Use full_factorial or custom_matrix."
        )

    factor_names = [factor.factor_name for factor in study.factors]
    if len(set(factor_names)) != len(factor_names):
        raise ValueError(
            f"Study '{study.study_id}' has duplicate factor_name values in Factors sheet."
        )

    if study.study_type == StudyType.FULL_FACTORIAL and not study.factors:
        raise ValueError(f"Study '{study.study_id}' has no factors for full_factorial expansion.")

    level_space = [factor.levels for factor in study.factors]
    combinations = itertools.product(*level_space) if level_space else [tuple()]
    repeats = _repeat_plan_for_study(study)

    cells: list[GeneratedDesignCell] = []
    for combo in combinations:
        factor_settings = {name: combo[idx] for idx, name in enumerate(factor_names)}
        if _is_disallowed_combination(factor_settings, study.constraints):
            continue
        resolved_params = _resolved_params_for_cell(
            study=study,
            factor_settings=factor_settings,
        )
        cell_id = f"{study.study_id}_cell_{_stable_suffix(factor_settings or {'_': 'base'})}"
        for repeat in repeats:
            repeat_id = int(repeat["repeat_id"])
            trial_id = f"{cell_id}_r{repeat_id:03d}"
            notes_parts = []
            if repeat.get("block_type") and str(repeat.get("block_type")) != "none":
                notes_parts.append(
                    f"block_type={repeat.get('block_type')}; block_value={repeat.get('block_value')}"
                )
            cells.append(
                GeneratedDesignCell.model_validate(
                    {
                        "study_id": study.study_id,
                        "trial_id": trial_id,
                        "cell_id": cell_id,
                        "factor_settings": factor_settings,
                        "start_section": _enum_text(study.start_section),
                        "end_section": _enum_text(study.end_section),
                        "base_artifact_id": study.base_artifact_id,
                        "resolved_params": resolved_params,
                        "status": "planned",
                        "repeat_id": repeat_id,
                        "seed": _seed_for_trial(
                            study=study,
                            trial_id=trial_id,
                            repeat_id=repeat_id,
                            explicit_seed=repeat.get("seed"),
                        ),
                        "notes": " | ".join(notes_parts) if notes_parts else None,
                    }
                )
            )
    return cells


def _build_custom_matrix_cells(
    study: StudyDesignSpec,
    *,
    raw_rows: list[dict[str, Any]],
) -> list[GeneratedDesignCell]:
    if not raw_rows:
        raise ValueError(
            f"Study '{study.study_id}' is custom_matrix but no rows were provided in "
            "Generated_Design_Matrix."
        )
    cells: list[GeneratedDesignCell] = []
    for raw in raw_rows:
        row_index = int(raw["row_index"])
        factor_settings = dict(raw.get("factor_settings", {}))
        resolved_params_base = dict(raw.get("resolved_params", {}))
        resolved_params = _resolved_params_for_cell(
            study=study,
            factor_settings=factor_settings,
            base_params=resolved_params_base,
        )
        cell_id = _normalize_text(raw.get("cell_id")) or (
            f"{study.study_id}_cell_{_stable_suffix({'row': row_index, **factor_settings})}"
        )
        trial_id = _normalize_text(raw.get("trial_id")) or f"{cell_id}_r001"
        start_section = _normalize_text(raw.get("start_section")) or _enum_text(study.start_section)
        end_section = _normalize_text(raw.get("end_section")) or _enum_text(study.end_section)
        status = _normalize_text(raw.get("status")) or "planned"
        cells.append(
            GeneratedDesignCell.model_validate(
                {
                    "study_id": study.study_id,
                    "trial_id": trial_id,
                    "cell_id": cell_id,
                    "factor_settings": factor_settings,
                    "start_section": start_section,
                    "end_section": end_section,
                    "base_artifact_id": (
                        _normalize_text(raw.get("base_artifact_id")) or study.base_artifact_id
                    ),
                    "resolved_params": resolved_params,
                    "status": status,
                    "repeat_id": 1,
                    "seed": _seed_for_trial(
                        study=study,
                        trial_id=trial_id,
                        repeat_id=1,
                        explicit_seed=None,
                    ),
                    "notes": f"source_row={row_index}",
                }
            )
        )
    return cells


def _default_stage_for_study(study: StudyDesignSpec) -> str:
    if study.intent == StudyIntent.CONFIRMATORY:
        return "Stage 5 - Confirmatory analysis"
    return "Stage 7 - Exploratory extension"


def _trial_spec_from_cell(study: StudyDesignSpec, cell: GeneratedDesignCell) -> dict[str, Any]:
    return {
        "experiment_id": study.study_id,
        "template_id": cell.cell_id,
        "supported": True,
        "params": dict(cell.resolved_params),
        "expand": {},
        "sections": list(_SECTION_ORDER),
        "artifacts": [],
        "start_section": _enum_text(cell.start_section),
        "end_section": _enum_text(cell.end_section),
        "base_artifact_id": cell.base_artifact_id,
        "reuse_policy": ReusePolicy.AUTO.value,
        "search_space_id": None,
        "study_id": study.study_id,
        "trial_id": cell.trial_id,
        "cell_id": cell.cell_id,
        "repeat_id": int(cell.repeat_id),
        "seed": cell.seed,
        "factor_settings": dict(cell.factor_settings),
        "fixed_controls": _fixed_controls_map(study),
        "design_metadata": {
            "study_type": _enum_text(study.study_type),
            "intent": _enum_text(study.intent),
            "replication_mode": study.replication_mode,
            "random_seed_policy": study.random_seed_policy,
            "cv_scheme": study.cv_scheme,
        },
    }


def _append_warning(warnings: list[str], message: str) -> None:
    text = str(message).strip()
    if text:
        warnings.append(text)


def _validate_study_rigor_policy(
    *,
    study: StudyDesignSpec,
    checklist: StudyRigorChecklistSpec | None,
    analysis_plan: AnalysisPlanSpec | None,
    field_presence: dict[str, bool] | None,
    warnings: list[str],
) -> None:
    strict = study.intent == StudyIntent.CONFIRMATORY

    required_core = {
        "generalization_claim": bool(str(study.generalization_claim or "").strip()),
        "primary_metric": bool(str(study.primary_metric or "").strip()),
        "cv_scheme": bool(str(study.cv_scheme or "").strip()),
    }
    if field_presence:
        for key, is_present in field_presence.items():
            if key in required_core:
                required_core[key] = bool(is_present)

    for field_name, is_present in required_core.items():
        if is_present:
            continue
        message = f"Study '{study.study_id}' is missing {field_name} in Study_Design."
        if strict:
            raise ValueError(message)
        _append_warning(warnings, message)

    if checklist is None:
        message = f"Study '{study.study_id}' has no Study_Rigor_Checklist entry."
        if strict:
            raise ValueError(
                message + " Confirmatory studies require a rigor checklist row."
            )
        _append_warning(warnings, message)
    elif strict and not checklist.confirmatory_lock_applied:
        raise ValueError(
            "Confirmatory study "
            f"'{study.study_id}' requires confirmatory_lock_applied=Yes in Study_Rigor_Checklist."
        )

    if analysis_plan is None:
        message = f"Study '{study.study_id}' has no Analysis_Plan entry."
        if strict:
            raise ValueError(message + " Confirmatory studies require an analysis plan row.")
        _append_warning(warnings, message)
    elif strict:
        missing_fields = [
            field_name
            for field_name, value in (
                ("primary_contrast", analysis_plan.primary_contrast),
                ("multiplicity_handling", analysis_plan.multiplicity_handling),
                ("interpretation_rules", analysis_plan.interpretation_rules),
            )
            if not str(value or "").strip()
        ]
        if missing_fields:
            raise ValueError(
                f"Confirmatory study '{study.study_id}' is missing required Analysis_Plan "
                "fields: "
                + ", ".join(missing_fields)
            )


def _build_study_design_layer(
    workbook: Workbook,
) -> tuple[
    list[StudyDesignSpec],
    list[StudyRigorChecklistSpec],
    list[AnalysisPlanSpec],
    list[GeneratedDesignCell],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[str],
]:
    if "Study_Design" not in workbook.sheetnames:
        return [], [], [], [], [], [], []

    study_header, study_rows = _sheet_rows(
        workbook,
        sheet_name="Study_Design",
        required_columns=_STUDY_DESIGN_REQUIRED_COLUMNS,
    )
    raw_studies: dict[str, dict[str, Any]] = {}
    field_presence_by_study: dict[str, dict[str, bool]] = {}
    for row_index, row in study_rows:
        study_id = _normalize_text(_read_cell(row, study_header, "study_id"))
        if not study_id:
            continue
        if study_id in raw_studies:
            raise ValueError(f"Study_Design has duplicate study_id '{study_id}' (row {row_index}).")
        enabled = _parse_enabled(
            _read_cell(row, study_header, "enabled"),
            row_index=row_index,
            sheet_name="Study_Design",
        )
        study_type = (
            _normalize_text(_read_cell(row, study_header, "study_type"))
            or StudyType.SINGLE_EXPERIMENT.value
        )
        if enabled and study_type == StudyType.FRACTIONAL_FACTORIAL.value:
            raise ValueError(
                "Study_Design row "
                f"{row_index} uses unsupported study_type='fractional_factorial'. "
                "Use full_factorial or custom_matrix."
            )
        raw_primary_metric = _normalize_text(_read_cell(row, study_header, "primary_metric"))
        raw_cv_scheme = _normalize_text(_read_cell(row, study_header, "cv_scheme"))
        raw_generalization_claim = _normalize_text(
            _read_cell(row, study_header, "generalization_claim")
        )
        field_presence_by_study[study_id] = {
            "generalization_claim": bool(raw_generalization_claim),
            "primary_metric": bool(raw_primary_metric),
            "cv_scheme": bool(raw_cv_scheme),
        }
        raw_studies[study_id] = {
            "study_id": study_id,
            "study_name": _normalize_text(_read_cell(row, study_header, "study_name")) or study_id,
            "enabled": enabled,
            "study_type": study_type,
            "intent": _normalize_text(_read_cell(row, study_header, "intent"))
            or StudyIntent.EXPLORATORY.value,
            "question": _normalize_text(_read_cell(row, study_header, "question")) or None,
            "generalization_claim": raw_generalization_claim or None,
            "start_section": _validated_section(
                _read_cell(row, study_header, "start_section"),
                field_name="start_section",
                row_index=row_index,
                sheet_name="Study_Design",
            )
            or SectionName.DATASET_SELECTION.value,
            "end_section": _validated_section(
                _read_cell(row, study_header, "end_section"),
                field_name="end_section",
                row_index=row_index,
                sheet_name="Study_Design",
            )
            or SectionName.EVALUATION.value,
            "base_artifact_id": _normalize_text(_read_cell(row, study_header, "base_artifact_id"))
            or None,
            "primary_metric": raw_primary_metric or "balanced_accuracy",
            "secondary_metrics": _normalize_text(_read_cell(row, study_header, "secondary_metrics"))
            or None,
            "cv_scheme": raw_cv_scheme or None,
            "nested_cv": _parse_optional_yes_no(
                _read_cell(row, study_header, "nested_cv"),
                row_index=row_index,
                sheet_name="Study_Design",
                field_name="nested_cv",
            ),
            "external_validation_planned": _parse_optional_yes_no(
                _read_cell(row, study_header, "external_validation_planned"),
                row_index=row_index,
                sheet_name="Study_Design",
                field_name="external_validation_planned",
            ),
            "blocking_strategy": _normalize_text(
                _read_cell(row, study_header, "blocking_strategy")
            )
            or None,
            "randomization_strategy": _normalize_text(
                _read_cell(row, study_header, "randomization_strategy")
            )
            or None,
            "replication_mode": _normalize_text(_read_cell(row, study_header, "replication_mode"))
            or "none",
            "replication_strategy": _normalize_text(
                _read_cell(row, study_header, "replication_strategy")
            )
            or None,
            "num_repeats": int(_read_cell(row, study_header, "num_repeats") or 1),
            "random_seed_policy": _normalize_text(
                _read_cell(row, study_header, "random_seed_policy")
            )
            or "fixed",
            "stopping_rule": _normalize_text(_read_cell(row, study_header, "stopping_rule")) or None,
            "notes": _normalize_text(_read_cell(row, study_header, "notes")) or None,
            "factors": [],
            "fixed_controls": [],
            "constraints": [],
            "blocking_replication": [],
        }

    factor_header, factor_rows = _sheet_rows(
        workbook,
        sheet_name="Factors",
        required_columns=_FACTOR_COLUMNS,
    )
    for row_index, row in factor_rows:
        study_id = _normalize_text(_read_cell(row, factor_header, "study_id"))
        if not study_id:
            continue
        if study_id not in raw_studies:
            raise ValueError(
                f"Factors row {row_index} references unknown study_id '{study_id}'."
            )
        factor_name = _normalize_text(_read_cell(row, factor_header, "factor_name"))
        raw_studies[study_id]["factors"].append(
            FactorSpec.model_validate(
                {
                    "study_id": study_id,
                    "factor_name": factor_name,
                    "section_name": (
                        _validated_section(
                            _read_cell(row, factor_header, "section_name"),
                            field_name="section_name",
                            row_index=row_index,
                            sheet_name="Factors",
                        )
                        or None
                    ),
                    "parameter_path": _normalize_text(_read_cell(row, factor_header, "parameter_path")),
                    "factor_type": _normalize_text(_read_cell(row, factor_header, "factor_type"))
                    or "categorical",
                    "levels": _parse_levels(
                        _read_cell(row, factor_header, "levels"),
                        row_index=row_index,
                        factor_name=factor_name,
                        study_id=study_id,
                    ),
                }
            )
        )

    control_header, control_rows = _sheet_rows(
        workbook,
        sheet_name="Fixed_Controls",
        required_columns=_FIXED_CONTROL_COLUMNS,
    )
    for row_index, row in control_rows:
        study_id = _normalize_text(_read_cell(row, control_header, "study_id"))
        if not study_id:
            continue
        if study_id not in raw_studies:
            raise ValueError(
                f"Fixed_Controls row {row_index} references unknown study_id '{study_id}'."
            )
        raw_studies[study_id]["fixed_controls"].append(
            FixedControlSpec.model_validate(
                {
                    "study_id": study_id,
                    "parameter_path": _normalize_text(
                        _read_cell(row, control_header, "parameter_path")
                    ),
                    "value": _parse_scalar_value(_read_cell(row, control_header, "value")),
                }
            )
        )

    constraint_header, constraint_rows = _sheet_rows(
        workbook,
        sheet_name="Constraints",
        required_columns=_CONSTRAINT_COLUMNS,
    )
    for row_index, row in constraint_rows:
        study_id = _normalize_text(_read_cell(row, constraint_header, "study_id"))
        if not study_id:
            continue
        if study_id not in raw_studies:
            raise ValueError(
                f"Constraints row {row_index} references unknown study_id '{study_id}'."
            )
        raw_studies[study_id]["constraints"].append(
            ConstraintSpec.model_validate(
                {
                    "study_id": study_id,
                    "if_factor": _normalize_text(_read_cell(row, constraint_header, "if_factor")),
                    "if_level": _parse_scalar_value(_read_cell(row, constraint_header, "if_level")),
                    "disallow_factor": _normalize_text(
                        _read_cell(row, constraint_header, "disallow_factor")
                    ),
                    "disallow_level": _parse_scalar_value(
                        _read_cell(row, constraint_header, "disallow_level")
                    ),
                    "reason": _normalize_text(_read_cell(row, constraint_header, "reason")) or None,
                }
            )
        )

    blocking_header, blocking_rows = _sheet_rows(
        workbook,
        sheet_name="Blocking_and_Replication",
        required_columns=_BLOCKING_AND_REPLICATION_COLUMNS,
    )
    for row_index, row in blocking_rows:
        study_id = _normalize_text(_read_cell(row, blocking_header, "study_id"))
        if not study_id:
            continue
        if study_id not in raw_studies:
            raise ValueError(
                "Blocking_and_Replication row "
                f"{row_index} references unknown study_id '{study_id}'."
            )
        raw_studies[study_id]["blocking_replication"].append(
            BlockingReplicationSpec.model_validate(
                {
                    "study_id": study_id,
                    "block_type": _normalize_text(_read_cell(row, blocking_header, "block_type"))
                    or "none",
                    "block_value": _normalize_text(_read_cell(row, blocking_header, "block_value"))
                    or None,
                    "repeat_id": int(_read_cell(row, blocking_header, "repeat_id") or 1),
                    "seed": (
                        int(_read_cell(row, blocking_header, "seed"))
                        if _read_cell(row, blocking_header, "seed") not in (None, "")
                        else None
                    ),
                }
            )
        )

    generated_header, generated_rows = _sheet_rows(
        workbook,
        sheet_name="Generated_Design_Matrix",
        required_columns=_GENERATED_DESIGN_MATRIX_COLUMNS,
    )
    generated_by_study: dict[str, list[dict[str, Any]]] = {}
    for row_index, row in generated_rows:
        study_id = _normalize_text(_read_cell(row, generated_header, "study_id"))
        if not study_id:
            continue
        generated_by_study.setdefault(study_id, []).append(
            {
                "row_index": row_index,
                "trial_id": _read_cell(row, generated_header, "trial_id"),
                "cell_id": _read_cell(row, generated_header, "cell_id"),
                "factor_settings": _parse_json_object(
                    _read_cell(row, generated_header, "factor_settings_json"),
                    row_index=row_index,
                    sheet_name="Generated_Design_Matrix",
                    column_name="factor_settings_json",
                ),
                "start_section": _read_cell(row, generated_header, "start_section"),
                "end_section": _read_cell(row, generated_header, "end_section"),
                "base_artifact_id": _read_cell(row, generated_header, "base_artifact_id"),
                "resolved_params": _parse_json_object(
                    _read_cell(row, generated_header, "resolved_params_json"),
                    row_index=row_index,
                    sheet_name="Generated_Design_Matrix",
                    column_name="resolved_params_json",
                ),
                "status": _read_cell(row, generated_header, "status"),
            }
        )

    checklist_header, checklist_rows = _sheet_rows(
        workbook,
        sheet_name="Study_Rigor_Checklist",
        required_columns=_STUDY_RIGOR_CHECKLIST_COLUMNS,
    )
    checklist_by_study: dict[str, StudyRigorChecklistSpec] = {}
    for row_index, row in checklist_rows:
        study_id = _normalize_text(_read_cell(row, checklist_header, "study_id"))
        if not study_id:
            continue
        if study_id not in raw_studies:
            raise ValueError(
                "Study_Rigor_Checklist row "
                f"{row_index} references unknown study_id '{study_id}'."
            )
        if study_id in checklist_by_study:
            raise ValueError(
                "Study_Rigor_Checklist has duplicate entries for "
                f"study_id '{study_id}' (row {row_index})."
            )

        required_text_fields = {
            "missing_data_plan": _normalize_text(
                _read_cell(row, checklist_header, "missing_data_plan")
            ),
            "class_imbalance_plan": _normalize_text(
                _read_cell(row, checklist_header, "class_imbalance_plan")
            ),
            "subgroup_plan": _normalize_text(_read_cell(row, checklist_header, "subgroup_plan")),
        }
        missing_required = [name for name, value in required_text_fields.items() if not value]
        if missing_required:
            raise ValueError(
                "Study_Rigor_Checklist row "
                f"{row_index} is missing required values: {', '.join(missing_required)}"
            )

        payload: dict[str, Any] = {
            "study_id": study_id,
            "missing_data_plan": required_text_fields["missing_data_plan"],
            "class_imbalance_plan": required_text_fields["class_imbalance_plan"],
            "subgroup_plan": required_text_fields["subgroup_plan"],
            "fairness_or_applicability_notes": _normalize_text(
                _read_cell(row, checklist_header, "fairness_or_applicability_notes")
            )
            or None,
            "analysis_notes": _normalize_text(_read_cell(row, checklist_header, "analysis_notes"))
            or None,
        }
        for column_name in _CHECKLIST_YES_NO_COLUMNS:
            payload[column_name] = _parse_yes_no(
                _read_cell(row, checklist_header, column_name),
                row_index=row_index,
                sheet_name="Study_Rigor_Checklist",
                field_name=column_name,
            )
        try:
            checklist_by_study[study_id] = StudyRigorChecklistSpec.model_validate(payload)
        except ValidationError as exc:
            raise ValueError(
                f"Invalid Study_Rigor_Checklist entry for study_id '{study_id}' "
                f"(row {row_index}): {exc}"
            ) from exc

    analysis_header, analysis_rows = _sheet_rows(
        workbook,
        sheet_name="Analysis_Plan",
        required_columns=_ANALYSIS_PLAN_COLUMNS,
    )
    analysis_by_study: dict[str, AnalysisPlanSpec] = {}
    for row_index, row in analysis_rows:
        study_id = _normalize_text(_read_cell(row, analysis_header, "study_id"))
        if not study_id:
            continue
        if study_id not in raw_studies:
            raise ValueError(
                f"Analysis_Plan row {row_index} references unknown study_id '{study_id}'."
            )
        if study_id in analysis_by_study:
            raise ValueError(
                f"Analysis_Plan has duplicate entries for study_id '{study_id}' (row {row_index})."
            )

        aggregation_level = _normalize_text(_read_cell(row, analysis_header, "aggregation_level"))
        uncertainty_method = _normalize_text(_read_cell(row, analysis_header, "uncertainty_method"))
        missing_required = [
            field_name
            for field_name, value in (
                ("aggregation_level", aggregation_level),
                ("uncertainty_method", uncertainty_method),
            )
            if not value
        ]
        if missing_required:
            raise ValueError(
                f"Analysis_Plan row {row_index} is missing required values: "
                + ", ".join(missing_required)
            )

        try:
            analysis_by_study[study_id] = AnalysisPlanSpec.model_validate(
                {
                    "study_id": study_id,
                    "primary_contrast": _normalize_text(
                        _read_cell(row, analysis_header, "primary_contrast")
                    )
                    or None,
                    "secondary_contrasts": _normalize_text(
                        _read_cell(row, analysis_header, "secondary_contrasts")
                    )
                    or None,
                    "aggregation_level": aggregation_level,
                    "uncertainty_method": uncertainty_method,
                    "multiplicity_handling": _normalize_text(
                        _read_cell(row, analysis_header, "multiplicity_handling")
                    )
                    or None,
                    "interaction_reporting_policy": _normalize_text(
                        _read_cell(row, analysis_header, "interaction_reporting_policy")
                    )
                    or None,
                    "interpretation_rules": _normalize_text(
                        _read_cell(row, analysis_header, "interpretation_rules")
                    )
                    or None,
                    "notes": _normalize_text(_read_cell(row, analysis_header, "notes")) or None,
                }
            )
        except ValidationError as exc:
            raise ValueError(
                f"Invalid analysis plan '{study_id}' in Analysis_Plan row {row_index}: {exc}"
            ) from exc

    study_specs = [
        StudyDesignSpec.model_validate(raw_studies[study_id]) for study_id in sorted(raw_studies)
    ]
    checklist_specs = [checklist_by_study[study_id] for study_id in sorted(checklist_by_study)]
    analysis_plan_specs = [analysis_by_study[study_id] for study_id in sorted(analysis_by_study)]
    generated_cells: list[GeneratedDesignCell] = []
    study_trials: list[dict[str, Any]] = []
    study_experiments: list[dict[str, Any]] = []
    validation_warnings: list[str] = []

    for study in study_specs:
        if study.enabled:
            _validate_study_rigor_policy(
                study=study,
                checklist=checklist_by_study.get(study.study_id),
                analysis_plan=analysis_by_study.get(study.study_id),
                field_presence=field_presence_by_study.get(study.study_id),
                warnings=validation_warnings,
            )
        if not study.enabled:
            continue
        if study.study_type == StudyType.CUSTOM_MATRIX:
            cells = _build_custom_matrix_cells(
                study,
                raw_rows=generated_by_study.get(study.study_id, []),
            )
        else:
            cells = _build_factorial_cells(study)
        for cell in cells:
            generated_cells.append(cell)
            study_trials.append(_trial_spec_from_cell(study, cell))
        if cells:
            study_experiments.append(
                {
                    "experiment_id": study.study_id,
                    "title": study.study_name,
                    "stage": _default_stage_for_study(study),
                    "manipulated_factor": ", ".join(factor.factor_name for factor in study.factors)
                    or None,
                    "primary_metric": study.primary_metric,
                    "notes": (
                        "Factorial study design compiled from workbook."
                        + (f" Question: {study.question}" if study.question else "")
                        + (f" Notes: {study.notes}" if study.notes else "")
                    ),
                    "variant_templates": [
                        row for row in study_trials if row["experiment_id"] == study.study_id
                    ],
                }
            )

    return (
        study_specs,
        checklist_specs,
        analysis_plan_specs,
        generated_cells,
        study_trials,
        study_experiments,
        validation_warnings,
    )


def compile_workbook_workbook(
    workbook: Workbook,
    *,
    source_workbook_path: Path | None = None,
) -> CompiledStudyManifest:
    workbook_schema_version = _resolve_workbook_schema_version(workbook)
    sheet_names = set(workbook.sheetnames)
    missing_sheets = sorted(_REQUIRED_SHEETS - sheet_names)
    if missing_sheets:
        raise ValueError("Workbook missing required sheet(s): " + ", ".join(missing_sheets))

    master_map = _parse_master_experiment_rows(workbook["Master_Experiments"])
    trial_specs = _parse_experiment_definitions_rows(workbook["Experiment_Definitions"])
    search_spaces = _parse_search_spaces_rows(workbook)
    (
        study_specs,
        study_rigor_checklists,
        analysis_plans,
        generated_cells,
        study_trials,
        study_experiments,
        validation_warnings,
    ) = _build_study_design_layer(workbook)
    all_trials = list(trial_specs) + list(study_trials)

    if not all_trials:
        raise ValueError(
            "No enabled executable rows were found in Experiment_Definitions or Study_Design."
        )

    legacy_experiment_ids = sorted({trial["experiment_id"] for trial in trial_specs})
    unknown_experiments = [
        experiment_id for experiment_id in legacy_experiment_ids if experiment_id not in master_map
    ]
    if unknown_experiments:
        raise ValueError(
            "Experiment_Definitions references unknown experiment_id values: "
            + ", ".join(unknown_experiments)
        )

    experiments_payload: dict[str, dict[str, Any]] = {}
    for experiment_id in legacy_experiment_ids:
        master_row = master_map[experiment_id]
        variants = [trial for trial in trial_specs if trial["experiment_id"] == experiment_id]
        experiments_payload[experiment_id] = {
            "experiment_id": experiment_id,
            "title": master_row["title"],
            "stage": master_row["stage"],
            "manipulated_factor": master_row["manipulated_factor"],
            "primary_metric": master_row["primary_metric"],
            "variant_templates": variants,
        }

    for study_experiment in study_experiments:
        experiment_id = str(study_experiment["experiment_id"])
        study_master_row = master_map.get(experiment_id)
        if experiment_id in experiments_payload:
            experiments_payload[experiment_id]["variant_templates"].extend(
                list(study_experiment.get("variant_templates", []))
            )
            continue
        experiments_payload[experiment_id] = {
            **study_experiment,
            "title": study_master_row["title"] if study_master_row else study_experiment["title"],
            "stage": study_master_row["stage"] if study_master_row else study_experiment["stage"],
            "manipulated_factor": (
                study_master_row["manipulated_factor"]
                if study_master_row and study_master_row.get("manipulated_factor")
                else study_experiment.get("manipulated_factor")
            ),
            "primary_metric": (
                study_master_row["primary_metric"]
                if study_master_row and study_master_row.get("primary_metric")
                else study_experiment["primary_metric"]
            ),
        }

    payload = {
        "schema_version": workbook_schema_version,
        "description": "Compiled from thesis_experiment_program.xlsx",
        "experiments": [experiments_payload[key] for key in sorted(experiments_payload)],
        "search_spaces": search_spaces,
        "study_designs": [study.model_dump(mode="python") for study in study_specs],
        "study_rigor_checklists": [
            checklist.model_dump(mode="python") for checklist in study_rigor_checklists
        ],
        "analysis_plans": [plan.model_dump(mode="python") for plan in analysis_plans],
        "generated_design_matrix": [cell.model_dump(mode="python") for cell in generated_cells],
        "effect_summaries": [],
        "validation_warnings": validation_warnings,
    }
    return compile_registry_payload(payload, source_registry_path=source_workbook_path)


def compile_workbook_file(path: Path) -> CompiledStudyManifest:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=False)
    return compile_workbook_workbook(workbook, source_workbook_path=workbook_path)
