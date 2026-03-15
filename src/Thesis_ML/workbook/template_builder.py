from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from Thesis_ML.config.paths import DEFAULT_WORKBOOK_TEMPLATE
from Thesis_ML.config.schema_versions import (
    SUPPORTED_WORKBOOK_SCHEMA_VERSIONS,
    WORKBOOK_SCHEMA_METADATA_START_ROW,
    WORKBOOK_SCHEMA_VERSION,
)
from Thesis_ML.workbook.governance_sheet_builders import (
    fill_ai_usage_sheet,
    fill_claim_ledger_sheet,
    fill_confirmatory_sheet,
    fill_dashboard_sheet,
    fill_data_profile_sheet,
    fill_data_selection_design_sheet,
    fill_decision_log_sheet,
    fill_dictionary_sheet,
    fill_ethics_sheet,
    fill_grouping_strategy_map_sheet,
    fill_run_log_sheet,
    fill_thesis_map_sheet,
)
from Thesis_ML.workbook.schema_metadata import write_schema_metadata
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_analysis_plan as _fill_analysis_plan_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_artifact_registry as _fill_artifact_registry_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_blocking_and_replication as _fill_blocking_and_replication_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_constraints as _fill_constraints_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_effect_summaries as _fill_effect_summaries_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_experiment_definitions as _fill_experiment_definitions_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_factors as _fill_factors_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_fixed_configs as _fill_fixed_configs_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_fixed_controls as _fill_fixed_controls_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_generated_design_matrix as _fill_generated_design_matrix_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_machine_status as _fill_machine_status_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_objectives as _fill_objectives_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_search_spaces as _fill_search_spaces_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_study_design as _fill_study_design_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_study_review as _fill_study_review_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_study_rigor_checklist as _fill_study_rigor_checklist_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_summary_outputs as _fill_summary_outputs_sheet,
)
from Thesis_ML.workbook.structured_execution_sheets import (
    fill_trial_results as _fill_trial_results_sheet,
)
from Thesis_ML.workbook.template_constants import (
    ANALYSIS_PLAN_COLUMNS,
    ARTIFACT_REGISTRY_COLUMNS,
    BLOCKING_AND_REPLICATION_COLUMNS,
    CONSTRAINTS_COLUMNS,
    EFFECT_SUMMARIES_COLUMNS,
    EXPERIMENT_DEFINITIONS_COLUMNS,
    FACTORS_COLUMNS,
    FIXED_CONFIGS_COLUMNS,
    FIXED_CONTROLS_COLUMNS,
    GENERATED_DESIGN_MATRIX_COLUMNS,
    MACHINE_STATUS_COLUMNS,
    MASTER_COLUMNS,
    OBJECTIVES_COLUMNS,
    RUN_LOG_COLUMNS,
    SEARCH_SPACES_COLUMNS,
    SHEET_ORDER,
    STAGE_V2,
    STUDY_DESIGN_COLUMNS,
    STUDY_REVIEW_COLUMNS,
    STUDY_RIGOR_CHECKLIST_COLUMNS,
    SUMMARY_OUTPUTS_COLUMNS,
    TRIAL_RESULTS_COLUMNS,
    build_experiments,
)
from Thesis_ML.workbook.template_primitives import (
    COL,
    THIN,
    add_dynamic_named_list,
    add_list_validation,
    col_idx,
    set_widths,
    style_body,
    style_header,
)
from Thesis_ML.workbook.template_validation import validate_template_workbook

OUT_XLSX = Path(DEFAULT_WORKBOOK_TEMPLATE)

def fill_readme_sheet(ws) -> None:
    ws.merge_cells("A1:I1")
    ws["A1"] = "Thesis Experiment Program Workbook (v2)"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=COL["title_bg"])
    ws["A1"].alignment = Alignment(horizontal="left")

    blocks = [
        (
            "Purpose",
            "Scientific control system for thesis experiment governance: pre-interpretation design, lock tracking, leakage-aware execution traceability, and chapter-ready evidence mapping.",
        ),
        (
            "Governance layers",
            "Experiment governance documents what each experiment is allowed to conclude. Data-slice governance documents what subset policy and grouping strategy were used in each run and what claims they can support.",
        ),
        (
            "Evidence tiers",
            "Primary confirmatory; Primary-supporting robustness; Secondary decision-support; Exploratory extension. Each tier has explicit interpretation boundaries.",
        ),
        (
            "Stage system",
            "Stage 1 Target lock -> Stage 2 Split lock -> Stage 3 Model lock -> Stage 4 Feature/preprocessing lock -> Stage 5 Confirmatory analysis -> Stage 6 Robustness analysis -> Stage 7 Exploratory extension.",
        ),
        (
            "Freeze policy",
            "Confirmatory eligibility requires D01-D07 locked in Decision_Log. Chapter 4 readiness additionally requires required confirmatory/supporting items completed and marked Ready_for_Chapter_4=YES.",
        ),
        (
            "Data_Selection_Design",
            "Defines allowed data slices (subject/session/task/modality/target scope, inclusion/exclusion, class-balance policy, leakage risk, and thesis-use boundary).",
        ),
        (
            "Grouping_Strategy_Map",
            "Defines split family, train/test grouping units, leakage safeguards, and interpretation boundaries for each grouping strategy identifier.",
        ),
        (
            "Data_Profile",
            "Structured worksheet for manual or imported descriptive counts by subject/session/task/modality/target/class and by Data_Slice_ID, with formula-based sparsity/imbalance flags.",
        ),
        (
            "Claim boundaries",
            "Within-person held-out-session claims and cross-person transfer claims are different scientific claims and must not be interpreted interchangeably.",
        ),
        (
            "Weak-split policy",
            "Weak split results are allowed only as inflation demonstrations and diagnostic contrast. They are not confirmatory evidence.",
        ),
        (
            "Governance additions",
            "Claim_Ledger enforces claim discipline. AI_Usage_Log supports tool transparency and human verification traceability. Ethics_Governance_Notes supports risk/mitigation accountability.",
        ),
        (
            "Workflow",
            "Master_Experiments + Data_Selection_Design + Grouping_Strategy_Map -> Run_Log -> Decision_Log -> Confirmatory_Set -> Claim_Ledger/AI_Usage_Log/Ethics_Governance_Notes -> Thesis_Map -> Dashboard.",
        ),
    ]
    row = 3
    for title, text in blocks:
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="EEF3FB")
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=9)
        ws.cell(row=row, column=2, value=text).alignment = Alignment(wrap_text=True, vertical="top")
        for rr in (row, row + 1):
            for cc in range(1, 10):
                ws.cell(rr, cc).border = THIN
        row += 3

    summary_row = row
    ws[f"A{summary_row}"] = (
        "This workbook is aligned to thesis method-choice/method-application workflow, "
        "data-slice/grouping governance, and leakage-aware reporting requirements."
    )
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row + 1, end_column=9)
    ws[f"A{summary_row}"].fill = PatternFill("solid", fgColor="FFF8E1")
    ws[f"A{summary_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"A{summary_row}"].border = THIN
    for cc in range(1, 10):
        ws.cell(summary_row + 1, cc).border = THIN

    set_widths(
        ws, {"A": 22, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24, "G": 24, "H": 24, "I": 24}
    )
    ws.freeze_panes = "A2"


def fill_master_sheet(ws) -> int:
    for i, h in enumerate(MASTER_COLUMNS, start=1):
        ws.cell(1, i, h)
    style_header(ws, 1, len(MASTER_COLUMNS))

    rows = build_experiments()
    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(MASTER_COLUMNS[:-1], start=1):
            ws.cell(r, c, row.get(name, ""))
        ws.cell(
            r,
            32,
            f'=IF($AA{r}="Dropped","N/A",IF(AND($H{r}<>"",$J{r}<>"",$K{r}<>"",$M{r}<>"",$N{r}<>"",$P{r}<>"",$R{r}<>"",$U{r}<>"",$V{r}<>"",$W{r}<>"",$Z{r}<>"",$Y{r}<>""),"READY","INCOMPLETE"))',
        )

    last = 1 + len(rows)
    style_body(ws, 2, last, 1, len(MASTER_COLUMNS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}{last}"

    set_widths(
        ws,
        {
            "A": 13,
            "B": 30,
            "C": 20,
            "D": 32,
            "E": 28,
            "F": 10,
            "G": 24,
            "H": 46,
            "I": 34,
            "J": 24,
            "K": 42,
            "L": 30,
            "M": 30,
            "N": 34,
            "O": 24,
            "P": 36,
            "Q": 24,
            "R": 18,
            "S": 26,
            "T": 36,
            "U": 36,
            "V": 34,
            "W": 36,
            "X": 34,
            "Y": 30,
            "Z": 36,
            "AA": 14,
            "AB": 18,
            "AC": 32,
            "AD": 28,
            "AE": 32,
            "AF": 16,
        },
    )

    add_list_validation(ws, "=List_Category", 3, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Evidential_Role", 4, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Stage", 5, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Priority", 6, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Reporting_Destination", 25, 2, 500, allow_blank=False)
    add_list_validation(ws, "=List_Status", 27, 2, 500, allow_blank=False)

    rng = f"A2:AF{max(last, 200)}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=['$D2="Primary confirmatory"'],
            fill=PatternFill("solid", fgColor=COL["confirmatory"]),
        ),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=['ISNUMBER(SEARCH("Exploratory",$D2))'],
            fill=PatternFill("solid", fgColor=COL["exploratory"]),
        ),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=['$F2="Critical"'], fill=PatternFill("solid", fgColor=COL["critical"])),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=['$AA2="Dropped"'],
            fill=PatternFill("solid", fgColor=COL["dropped"]),
            font=Font(color="6E6E6E", italic=True),
        ),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[
                'OR($H2="",$J2="",$K2="",$M2="",$N2="",$P2="",$R2="",$U2="",$V2="",$W2="",$Z2="",$Y2="")'
            ],
            fill=PatternFill("solid", fgColor=COL["missing"]),
        ),
    )
    return last


def fill_experiment_definitions_sheet(ws) -> int:
    return _fill_experiment_definitions_sheet(
        ws=ws,
        experiment_definitions_columns=EXPERIMENT_DEFINITIONS_COLUMNS,
    )


def fill_search_spaces_sheet(ws, wb: Workbook) -> int:
    return _fill_search_spaces_sheet(
        ws=ws,
        wb=wb,
        search_spaces_columns=SEARCH_SPACES_COLUMNS,
    )


def fill_study_design_sheet(ws, wb: Workbook) -> int:
    return _fill_study_design_sheet(
        ws=ws,
        wb=wb,
        study_design_columns=STUDY_DESIGN_COLUMNS,
    )


def fill_study_rigor_checklist_sheet(ws) -> int:
    return _fill_study_rigor_checklist_sheet(
        ws=ws,
        study_rigor_checklist_columns=STUDY_RIGOR_CHECKLIST_COLUMNS,
    )


def fill_analysis_plan_sheet(ws) -> int:
    return _fill_analysis_plan_sheet(
        ws=ws,
        analysis_plan_columns=ANALYSIS_PLAN_COLUMNS,
    )


def fill_factors_sheet(ws) -> int:
    return _fill_factors_sheet(
        ws=ws,
        factors_columns=FACTORS_COLUMNS,
    )


def fill_fixed_controls_sheet(ws) -> int:
    return _fill_fixed_controls_sheet(
        ws=ws,
        fixed_controls_columns=FIXED_CONTROLS_COLUMNS,
    )


def fill_constraints_sheet(ws) -> int:
    return _fill_constraints_sheet(
        ws=ws,
        constraints_columns=CONSTRAINTS_COLUMNS,
    )


def fill_blocking_and_replication_sheet(ws) -> int:
    return _fill_blocking_and_replication_sheet(
        ws=ws,
        blocking_and_replication_columns=BLOCKING_AND_REPLICATION_COLUMNS,
    )


def fill_generated_design_matrix_sheet(ws) -> int:
    return _fill_generated_design_matrix_sheet(
        ws=ws,
        generated_design_matrix_columns=GENERATED_DESIGN_MATRIX_COLUMNS,
    )


def fill_effect_summaries_sheet(ws) -> int:
    return _fill_effect_summaries_sheet(
        ws=ws,
        effect_summaries_columns=EFFECT_SUMMARIES_COLUMNS,
    )


def fill_study_review_sheet(ws) -> int:
    return _fill_study_review_sheet(
        ws=ws,
        study_review_columns=STUDY_REVIEW_COLUMNS,
    )


def fill_artifact_registry_sheet(ws) -> int:
    return _fill_artifact_registry_sheet(
        ws=ws,
        artifact_registry_columns=ARTIFACT_REGISTRY_COLUMNS,
    )


def fill_fixed_configs_sheet(ws) -> int:
    return _fill_fixed_configs_sheet(
        ws=ws,
        fixed_configs_columns=FIXED_CONFIGS_COLUMNS,
    )


def fill_objectives_sheet(ws) -> int:
    return _fill_objectives_sheet(
        ws=ws,
        objectives_columns=OBJECTIVES_COLUMNS,
    )


def fill_machine_status_sheet(ws) -> int:
    return _fill_machine_status_sheet(
        ws=ws,
        machine_status_columns=MACHINE_STATUS_COLUMNS,
    )


def fill_trial_results_sheet(ws) -> int:
    return _fill_trial_results_sheet(
        ws=ws,
        trial_results_columns=TRIAL_RESULTS_COLUMNS,
    )


def fill_summary_outputs_sheet(ws) -> int:
    return _fill_summary_outputs_sheet(
        ws=ws,
        summary_outputs_columns=SUMMARY_OUTPUTS_COLUMNS,
    )


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEET_ORDER:
        wb.create_sheet(name)

    fill_readme_sheet(wb["README"])
    fill_master_sheet(wb["Master_Experiments"])
    add_dynamic_named_list(
        wb, "List_Experiment_ID", "Master_Experiments", col_idx(MASTER_COLUMNS, "Experiment_ID"), 2
    )
    fill_experiment_definitions_sheet(wb["Experiment_Definitions"])
    fill_search_spaces_sheet(wb["Search_Spaces"], wb)
    fill_study_design_sheet(wb["Study_Design"], wb)
    fill_study_rigor_checklist_sheet(wb["Study_Rigor_Checklist"])
    fill_analysis_plan_sheet(wb["Analysis_Plan"])
    fill_factors_sheet(wb["Factors"])
    fill_fixed_controls_sheet(wb["Fixed_Controls"])
    fill_constraints_sheet(wb["Constraints"])
    fill_blocking_and_replication_sheet(wb["Blocking_and_Replication"])
    fill_generated_design_matrix_sheet(wb["Generated_Design_Matrix"])
    fill_effect_summaries_sheet(wb["Effect_Summaries"])
    fill_study_review_sheet(wb["Study_Review"])
    fill_artifact_registry_sheet(wb["Artifact_Registry"])
    fill_fixed_configs_sheet(wb["Fixed_Configs"])
    fill_objectives_sheet(wb["Objectives"])
    fill_machine_status_sheet(wb["Machine_Status"])
    fill_trial_results_sheet(wb["Trial_Results"])
    fill_summary_outputs_sheet(wb["Summary_Outputs"])
    fill_data_selection_design_sheet(wb["Data_Selection_Design"], wb)
    fill_grouping_strategy_map_sheet(wb["Grouping_Strategy_Map"], wb)
    fill_data_profile_sheet(wb["Data_Profile"])
    fill_run_log_sheet(wb["Run_Log"])
    fill_decision_log_sheet(wb["Decision_Log"])
    fill_confirmatory_sheet(wb["Confirmatory_Set"])
    fill_thesis_map_sheet(wb["Thesis_Map"])
    fill_dictionary_sheet(wb["Dictionary_Validation"], wb)
    fill_dashboard_sheet(wb["Dashboard"])
    fill_claim_ledger_sheet(wb["Claim_Ledger"])
    fill_ai_usage_sheet(wb["AI_Usage_Log"])
    fill_ethics_sheet(wb["Ethics_Governance_Notes"])

    wb.calculation.fullCalcOnLoad = True
    write_schema_metadata(wb["README"])
    generator_row = WORKBOOK_SCHEMA_METADATA_START_ROW + 7
    wb["README"][f"A{generator_row}"] = "Generated by create_thesis_experiment_workbook.py (v2)"
    wb["README"][f"A{generator_row}"].font = Font(italic=True, color="4B5563")
    return wb


def validate(path: Path) -> dict[str, str]:
    return validate_template_workbook(
        path,
        sheet_order=SHEET_ORDER,
        stage_vocabulary=STAGE_V2,
        run_log_columns=RUN_LOG_COLUMNS,
        experiment_definitions_columns=EXPERIMENT_DEFINITIONS_COLUMNS,
        workbook_schema_version=WORKBOOK_SCHEMA_VERSION,
        supported_workbook_schema_versions=SUPPORTED_WORKBOOK_SCHEMA_VERSIONS,
    )


def main() -> None:
    wb = build_workbook()
    wb.save(OUT_XLSX)
    summary = validate(OUT_XLSX)
    print("Created workbook:", OUT_XLSX.resolve())
    print("Sheet order valid:", summary["sheet_order_ok"])
    print("Missing required sheets:", summary["missing_sheets"])
    print("Legacy required sheets present:", summary["legacy_sheets_present"])
    print("New sheets present:", summary["new_sheets_present"])
    print("Sheet count:", summary["sheet_count"])
    print("Data validations found:", summary["data_validations_found"])
    print("Experiment_Definitions columns valid:", summary["experiment_definitions_columns_ok"])
    print("Run_Log new columns present:", summary["run_log_new_columns_present"])
    print("Required named lists present:", summary["required_named_lists_present"])
    print("Missing named lists:", summary["missing_named_lists"])
    print("Experiment_Ready formula present:", summary["experiment_ready_formula_present"])
    print("Confirmatory formulas present:", summary["confirmatory_formula_present"])
    print("Dashboard formulas present:", summary["dashboard_formula_present"])
    print("Stage vocabulary consistent:", summary["stage_vocab_consistent"])
    print("Stage rows detected:", summary["stage_vocab_rows"])
    print("Schema metadata keys present:", summary["schema_metadata_keys_present"])
    print("Workbook schema version:", summary["workbook_schema_version"])
    print("Workbook schema supported:", summary["workbook_schema_supported"])


if __name__ == "__main__":
    main()

