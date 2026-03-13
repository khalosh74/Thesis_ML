from __future__ import annotations

from collections.abc import Set
from pathlib import Path

from openpyxl import load_workbook

from Thesis_ML.workbook.schema_metadata import (
    expected_schema_metadata,
    read_schema_metadata,
)


def validate_template_workbook(
    path: Path,
    *,
    sheet_order: list[str],
    stage_vocabulary: list[str],
    run_log_columns: list[str],
    experiment_definitions_columns: list[str],
    workbook_schema_version: str,
    supported_workbook_schema_versions: Set[str],
) -> dict[str, str]:
    wb = load_workbook(path)
    sheets = [ws.title for ws in wb.worksheets]
    sheet_ok = sheets == sheet_order
    missing_sheets = [sheet_name for sheet_name in sheet_order if sheet_name not in sheets]

    legacy_required = [
        "README",
        "Master_Experiments",
        "Experiment_Definitions",
        "Search_Spaces",
        "Artifact_Registry",
        "Fixed_Configs",
        "Objectives",
        "Machine_Status",
        "Trial_Results",
        "Summary_Outputs",
        "Run_Log",
        "Decision_Log",
        "Confirmatory_Set",
        "Thesis_Map",
        "Dictionary_Validation",
        "Dashboard",
        "Claim_Ledger",
        "AI_Usage_Log",
        "Ethics_Governance_Notes",
    ]
    legacy_sheets_present = all(sheet_name in sheets for sheet_name in legacy_required)
    new_sheets = [
        "Experiment_Definitions",
        "Search_Spaces",
        "Artifact_Registry",
        "Fixed_Configs",
        "Objectives",
        "Machine_Status",
        "Trial_Results",
        "Summary_Outputs",
        "Data_Selection_Design",
        "Grouping_Strategy_Map",
        "Data_Profile",
    ]
    new_sheets_present = all(sheet_name in sheets for sheet_name in new_sheets)

    master = wb["Master_Experiments"]
    experiment_definitions = wb["Experiment_Definitions"]
    readme = wb["README"]
    confirm = wb["Confirmatory_Set"]
    dash = wb["Dashboard"]
    dictionary = wb["Dictionary_Validation"]
    run_log = wb["Run_Log"]

    dv_count = sum(
        len(wb[name].data_validations.dataValidation)
        for name in [
            "Master_Experiments",
            "Experiment_Definitions",
            "Search_Spaces",
            "Data_Selection_Design",
            "Grouping_Strategy_Map",
            "Run_Log",
            "Decision_Log",
            "Confirmatory_Set",
            "Claim_Ledger",
            "AI_Usage_Log",
            "Ethics_Governance_Notes",
        ]
    )
    stage_values = {master[f"E{row_idx}"].value for row_idx in range(2, master.max_row + 1)}
    stage_vocab = set(stage_vocabulary)
    stage_consistent = stage_values.issubset(stage_vocab)
    run_log_headers = [
        run_log.cell(1, col_idx).value for col_idx in range(1, len(run_log_columns) + 1)
    ]
    run_log_new_cols = [
        "Data_Slice_ID",
        "Grouping_Strategy_ID",
        "Train_Group_Rule",
        "Test_Group_Rule",
        "Transfer_Direction",
        "Sample_Count",
        "Class_Counts",
        "Imbalance_Status",
        "Leakage_Check_Status",
        "Session_Coverage",
        "Task_Coverage",
        "Modality_Coverage",
    ]
    run_log_columns_ok = all(col in run_log_headers for col in run_log_new_cols)
    experiment_definitions_headers = [
        experiment_definitions.cell(1, col_idx).value
        for col_idx in range(1, len(experiment_definitions_columns) + 1)
    ]
    experiment_definitions_columns_ok = (
        experiment_definitions_headers == experiment_definitions_columns
    )

    required_named_lists = [
        "List_Experiment_ID",
        "List_Data_Slice_ID",
        "List_Grouping_Strategy_ID",
        "List_Subject_Scope",
        "List_Session_Scope",
        "List_Task_Scope",
        "List_Modality_Scope",
        "List_Class_Balance_Policy",
        "List_Split_Family",
        "List_Imbalance_Status",
        "List_Leakage_Check_Status",
        "List_Transfer_Direction",
        "List_Execution_Section",
        "List_Reuse_Policy",
        "List_Search_Optimization_Mode",
        "List_Search_Parameter_Scope",
        "List_Search_Space_ID",
    ]
    defined_names = {name for name in wb.defined_names.keys()}
    named_lists_ok = all(name in defined_names for name in required_named_lists)
    missing_named_lists = [name for name in required_named_lists if name not in defined_names]

    confirmatory_formulas_ok = (
        all(
            isinstance(confirm[cell_name].value, str) and confirm[cell_name].value.startswith("=")
            for cell_name in ["F2", "G2", "J2"]
        )
        and isinstance(master["AF2"].value, str)
        and master["AF2"].value.startswith("=")
    )
    dashboard_core_formulas_ok = all(
        isinstance(dash[cell_name].value, str) and dash[cell_name].value.startswith("=")
        for cell_name in ["B13", "B14", "B22", "B23", "B24", "B25", "K4", "N4", "E18"]
    )
    schema_metadata = read_schema_metadata(readme)
    required_schema_metadata = expected_schema_metadata()
    schema_metadata_keys_present = all(key in schema_metadata for key in required_schema_metadata)
    workbook_schema_version_value = schema_metadata.get(
        "workbook_schema_version", workbook_schema_version
    )
    workbook_schema_supported = workbook_schema_version_value in supported_workbook_schema_versions

    return {
        "sheet_order_ok": str(sheet_ok),
        "missing_sheets": ", ".join(missing_sheets) if missing_sheets else "None",
        "legacy_sheets_present": str(legacy_sheets_present),
        "new_sheets_present": str(new_sheets_present),
        "sheet_count": str(len(sheets)),
        "data_validations_found": str(dv_count),
        "experiment_definitions_columns_ok": str(experiment_definitions_columns_ok),
        "run_log_new_columns_present": str(run_log_columns_ok),
        "required_named_lists_present": str(named_lists_ok),
        "missing_named_lists": ", ".join(missing_named_lists) if missing_named_lists else "None",
        "experiment_ready_formula_present": str(
            isinstance(master["AF2"].value, str) and master["AF2"].value.startswith("=")
        ),
        "confirmatory_formula_present": str(confirmatory_formulas_ok),
        "dashboard_formula_present": str(dashboard_core_formulas_ok),
        "stage_vocab_consistent": str(stage_consistent),
        "stage_vocab_rows": str(len([value for value in stage_values if value is not None])),
        "dictionary_stage_head": str(dictionary["C3"].value),
        "schema_metadata_keys_present": str(schema_metadata_keys_present),
        "workbook_schema_version": workbook_schema_version_value,
        "workbook_schema_supported": str(workbook_schema_supported),
    }


__all__ = ["validate_template_workbook"]
