from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from Thesis_ML.config.schema_versions import (
    WORKBOOK_SCHEMA_VERSION,
    WORKBOOK_WRITEBACK_SCHEMA_VERSION,
)
from Thesis_ML.orchestration.decision_support import run_workbook_decision_support_campaign
from Thesis_ML.orchestration.workbook_compiler import compile_workbook_file
from Thesis_ML.workbook.builder import build_workbook
from Thesis_ML.workbook.schema_metadata import read_schema_metadata


def _make_workbook(path: Path) -> None:
    workbook = build_workbook()
    workbook.save(path)


def _set_executable_row(path: Path) -> None:
    workbook = load_workbook(path)
    ws = workbook["Experiment_Definitions"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col = {str(name): idx + 1 for idx, name in enumerate(headers)}

    ws.cell(2, col["experiment_id"], "E16")
    ws.cell(2, col["enabled"], "Yes")
    ws.cell(2, col["start_section"], "dataset_selection")
    ws.cell(2, col["end_section"], "evaluation")
    ws.cell(2, col["base_artifact_id"], "")
    ws.cell(2, col["target"], "coarse_affect")
    ws.cell(2, col["cv"], "within_subject_loso_session")
    ws.cell(2, col["model"], "ridge")
    ws.cell(2, col["subject"], "sub-001")
    ws.cell(2, col["train_subject"], "")
    ws.cell(2, col["test_subject"], "")
    ws.cell(2, col["filter_task"], "")
    ws.cell(2, col["filter_modality"], "")
    ws.cell(2, col["reuse_policy"], "auto")
    workbook.save(path)


def _set_factorial_design(path: Path) -> None:
    workbook = load_workbook(path)

    study_ws = workbook["Study_Design"]
    study_cols = _sheet_header_map(study_ws, 2)
    study_ws.cell(3, study_cols["study_id"], "S01")
    study_ws.cell(3, study_cols["study_name"], "Roundtrip factorial study")
    study_ws.cell(3, study_cols["enabled"], "Yes")
    study_ws.cell(3, study_cols["study_type"], "full_factorial")
    study_ws.cell(3, study_cols["intent"], "exploratory")
    study_ws.cell(3, study_cols["start_section"], "dataset_selection")
    study_ws.cell(3, study_cols["end_section"], "evaluation")
    study_ws.cell(3, study_cols["primary_metric"], "balanced_accuracy")
    study_ws.cell(3, study_cols["cv_scheme"], "within_subject_loso_session")
    study_ws.cell(3, study_cols["num_repeats"], 1)
    study_ws.cell(3, study_cols["replication_mode"], "fixed_repeats")
    study_ws.cell(3, study_cols["random_seed_policy"], "fixed")

    factors_ws = workbook["Factors"]
    factors_cols = _sheet_header_map(factors_ws, 2)
    factors_ws.cell(3, factors_cols["study_id"], "S01")
    factors_ws.cell(3, factors_cols["factor_name"], "model")
    factors_ws.cell(3, factors_cols["parameter_path"], "model")
    factors_ws.cell(3, factors_cols["factor_type"], "categorical")
    factors_ws.cell(3, factors_cols["levels"], "ridge|logreg")
    factors_ws.cell(4, factors_cols["study_id"], "S01")
    factors_ws.cell(4, factors_cols["factor_name"], "filter_task")
    factors_ws.cell(4, factors_cols["parameter_path"], "filter_task")
    factors_ws.cell(4, factors_cols["factor_type"], "categorical")
    factors_ws.cell(4, factors_cols["levels"], "passive|emo")

    fixed_ws = workbook["Fixed_Controls"]
    fixed_cols = _sheet_header_map(fixed_ws, 2)
    fixed_ws.cell(3, fixed_cols["study_id"], "S01")
    fixed_ws.cell(3, fixed_cols["parameter_path"], "target")
    fixed_ws.cell(3, fixed_cols["value"], "coarse_affect")
    fixed_ws.cell(4, fixed_cols["study_id"], "S01")
    fixed_ws.cell(4, fixed_cols["parameter_path"], "subject")
    fixed_ws.cell(4, fixed_cols["value"], "sub-001")

    blocking_ws = workbook["Blocking_and_Replication"]
    blocking_cols = _sheet_header_map(blocking_ws, 2)
    blocking_ws.cell(3, blocking_cols["study_id"], "S01")
    blocking_ws.cell(3, blocking_cols["block_type"], "none")
    blocking_ws.cell(3, blocking_cols["repeat_id"], 1)

    workbook.save(path)


def _write_index_csv(path: Path) -> None:
    df = pd.DataFrame(
        [
            {"sample_id": "s1", "subject": "sub-001", "task": "passive", "modality": "audio"},
            {"sample_id": "s2", "subject": "sub-001", "task": "emo", "modality": "video"},
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def _stub_run_experiment(**kwargs: object) -> dict[str, object]:
    run_id = str(kwargs["run_id"])
    reports_root = Path(kwargs["reports_root"])
    report_dir = reports_root / run_id
    report_dir.mkdir(parents=True, exist_ok=True)

    config_path = report_dir / "config.json"
    metrics_path = report_dir / "metrics.json"
    fold_metrics_path = report_dir / "fold_metrics.csv"
    fold_splits_path = report_dir / "fold_splits.csv"
    predictions_path = report_dir / "predictions.csv"
    spatial_path = report_dir / "spatial_compatibility_report.json"
    interpretability_path = report_dir / "interpretability_summary.json"

    config_path.write_text('{"ok": true}\n', encoding="utf-8")
    metrics_path.write_text(
        json.dumps(
            {
                "accuracy": 0.62,
                "balanced_accuracy": 0.59,
                "macro_f1": 0.58,
                "n_folds": 2,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    fold_metrics_path.write_text("fold,score\n0,0.59\n", encoding="utf-8")
    fold_splits_path.write_text("fold,train,test\n0,a,b\n", encoding="utf-8")
    predictions_path.write_text("y_true,y_pred\nneg,neg\n", encoding="utf-8")
    spatial_path.write_text(
        '{"status":"passed","passed":true,"n_groups_checked":1}\n', encoding="utf-8"
    )
    interpretability_path.write_text('{"status":"not_applicable"}\n', encoding="utf-8")

    return {
        "run_id": run_id,
        "report_dir": str(report_dir),
        "config_path": str(config_path),
        "metrics_path": str(metrics_path),
        "fold_metrics_path": str(fold_metrics_path),
        "fold_splits_path": str(fold_splits_path),
        "predictions_path": str(predictions_path),
        "spatial_compatibility_report_path": str(spatial_path),
        "interpretability_summary_path": str(interpretability_path),
        "artifact_ids": {
            "metrics_bundle": f"metrics_{run_id}",
            "interpretability_bundle": f"interp_{run_id}",
            "experiment_report": f"report_{run_id}",
        },
        "metrics": {
            "accuracy": 0.62,
            "balanced_accuracy": 0.59,
            "macro_f1": 0.58,
            "n_folds": 2,
        },
    }


def _sheet_header_map(ws, header_row: int) -> dict[str, int]:
    return {
        str(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(header_row, col).value
    }


def test_workbook_roundtrip_compile_execute_writeback(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_executable_row(workbook_path)
    index_csv = tmp_path / "dataset_index.csv"
    _write_index_csv(index_csv)

    manifest = compile_workbook_file(workbook_path)
    assert len(manifest.trial_specs) == 1

    from Thesis_ML.orchestration import decision_support as orchestrator

    monkeypatch.setattr(orchestrator, "run_experiment", _stub_run_experiment)

    result = run_workbook_decision_support_campaign(
        workbook_path=workbook_path,
        index_csv=index_csv,
        data_root=tmp_path / "Data",
        cache_dir=tmp_path / "cache",
        output_root=tmp_path / "artifacts" / "decision_support",
        experiment_id=None,
        stage=None,
        run_all=True,
        seed=42,
        n_permutations=0,
        dry_run=False,
        write_back_to_workbook=True,
        append_workbook_run_log=True,
        workbook_output_dir=tmp_path / "workbook_outputs",
    )

    output_workbook_path = Path(str(result["workbook_output_path"]))
    assert output_workbook_path.exists()
    assert output_workbook_path != workbook_path
    assert output_workbook_path.name.startswith("thesis_experiment_program__results_")

    output_wb = load_workbook(output_workbook_path)

    machine_ws = output_wb["Machine_Status"]
    machine_cols = _sheet_header_map(machine_ws, 2)
    machine_rows = [
        row
        for row in range(3, machine_ws.max_row + 1)
        if str(machine_ws.cell(row, machine_cols["machine_id"]).value or "").startswith("campaign_")
    ]
    assert machine_rows
    machine_status_value = machine_ws.cell(machine_rows[-1], machine_cols["status"]).value
    assert machine_status_value in {"Open", "Monitoring", "Closed"}

    trial_ws = output_wb["Trial_Results"]
    trial_cols = _sheet_header_map(trial_ws, 2)
    trial_rows = [
        row
        for row in range(3, trial_ws.max_row + 1)
        if str(trial_ws.cell(row, trial_cols["experiment_id"]).value or "") == "E16"
    ]
    assert trial_rows
    latest_trial_row = trial_rows[-1]
    assert trial_ws.cell(latest_trial_row, trial_cols["status"]).value == "completed"
    assert trial_ws.cell(latest_trial_row, trial_cols["report_path"]).value
    assert trial_ws.cell(latest_trial_row, trial_cols["metrics_path"]).value

    summary_ws = output_wb["Summary_Outputs"]
    summary_cols = _sheet_header_map(summary_ws, 2)
    summary_rows = [
        row
        for row in range(3, summary_ws.max_row + 1)
        if str(summary_ws.cell(row, summary_cols["summary_type"]).value or "") != ""
    ]
    assert summary_rows
    summary_types = {
        str(summary_ws.cell(row, summary_cols["summary_type"]).value) for row in summary_rows
    }
    assert "best_full_pipeline" in summary_types

    run_log_ws = output_wb["Run_Log"]
    run_log_cols = _sheet_header_map(run_log_ws, 1)
    run_log_rows = [
        row
        for row in range(2, run_log_ws.max_row + 1)
        if str(run_log_ws.cell(row, run_log_cols["Experiment_ID"]).value or "") == "E16"
    ]
    assert run_log_rows
    latest_run_log_row = run_log_rows[-1]
    assert run_log_ws.cell(latest_run_log_row, run_log_cols["Result_Summary"]).value == "completed"
    assert run_log_ws.cell(latest_run_log_row, run_log_cols["Artifact_Path"]).value

    source_wb = load_workbook(workbook_path)
    source_schema_metadata = read_schema_metadata(source_wb["README"])
    output_schema_metadata = read_schema_metadata(output_wb["README"])
    assert source_schema_metadata["workbook_schema_version"] == WORKBOOK_SCHEMA_VERSION
    assert output_schema_metadata["workbook_schema_version"] == WORKBOOK_SCHEMA_VERSION
    assert (
        output_schema_metadata["workbook_writeback_schema_version"]
        == WORKBOOK_WRITEBACK_SCHEMA_VERSION
    )

    source_trial_ws = source_wb["Trial_Results"]
    source_trial_cols = _sheet_header_map(source_trial_ws, 2)
    source_rows_with_e16 = [
        row
        for row in range(3, source_trial_ws.max_row + 1)
        if str(source_trial_ws.cell(row, source_trial_cols["experiment_id"]).value or "") == "E16"
    ]
    assert source_rows_with_e16 == []


def test_workbook_roundtrip_factorial_writeback(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook_path = tmp_path / "thesis_experiment_program.xlsx"
    _make_workbook(workbook_path)
    _set_factorial_design(workbook_path)
    index_csv = tmp_path / "dataset_index.csv"
    _write_index_csv(index_csv)

    manifest = compile_workbook_file(workbook_path)
    study_trials = [trial for trial in manifest.trial_specs if trial.study_id == "S01"]
    assert len(study_trials) == 4

    from Thesis_ML.orchestration import decision_support as orchestrator

    monkeypatch.setattr(orchestrator, "run_experiment", _stub_run_experiment)

    result = run_workbook_decision_support_campaign(
        workbook_path=workbook_path,
        index_csv=index_csv,
        data_root=tmp_path / "Data",
        cache_dir=tmp_path / "cache",
        output_root=tmp_path / "artifacts" / "decision_support",
        experiment_id=None,
        stage=None,
        run_all=True,
        seed=42,
        n_permutations=0,
        dry_run=False,
        write_back_to_workbook=True,
        append_workbook_run_log=True,
        workbook_output_dir=tmp_path / "workbook_outputs",
    )

    output_workbook_path = Path(str(result["workbook_output_path"]))
    output_wb = load_workbook(output_workbook_path)

    generated_ws = output_wb["Generated_Design_Matrix"]
    generated_cols = _sheet_header_map(generated_ws, 2)
    generated_rows = [
        row
        for row in range(3, generated_ws.max_row + 1)
        if str(generated_ws.cell(row, generated_cols["study_id"]).value or "") == "S01"
    ]
    assert generated_rows
    statuses = {
        str(generated_ws.cell(row, generated_cols["status"]).value or "") for row in generated_rows
    }
    assert "completed" in statuses

    trial_ws = output_wb["Trial_Results"]
    trial_cols = _sheet_header_map(trial_ws, 2)
    trial_rows = [
        row
        for row in range(3, trial_ws.max_row + 1)
        if str(trial_ws.cell(row, trial_cols["study_id"]).value or "") == "S01"
    ]
    assert trial_rows
    assert trial_ws.cell(trial_rows[-1], trial_cols["cell_id"]).value
    assert trial_ws.cell(trial_rows[-1], trial_cols["factor_settings_json"]).value

    effects_ws = output_wb["Effect_Summaries"]
    effects_cols = _sheet_header_map(effects_ws, 2)
    effect_rows = [
        row
        for row in range(3, effects_ws.max_row + 1)
        if str(effects_ws.cell(row, effects_cols["study_id"]).value or "") == "S01"
    ]
    assert effect_rows
    effect_types = {
        str(effects_ws.cell(row, effects_cols["summary_type"]).value or "") for row in effect_rows
    }
    assert "best_by_study" in effect_types
