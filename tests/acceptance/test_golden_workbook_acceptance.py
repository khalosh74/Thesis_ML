from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from Thesis_ML.orchestration.decision_support import run_workbook_decision_support_campaign
from Thesis_ML.orchestration.workbook_compiler import compile_workbook_file
from Thesis_ML.workbook.validation import validate_workbook


def _header_map(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        key = str(value).strip()
        if key:
            mapping[key] = col
    return mapping


def _rows_with_value(
    ws, *, header_map: dict[str, int], key: str, expected: str, start_row: int
) -> list[int]:
    return [
        row
        for row in range(start_row, ws.max_row + 1)
        if str(ws.cell(row, header_map[key]).value or "") == expected
    ]


def test_golden_workbook_to_execution_writeback_acceptance(
    acceptance_sample_data: dict[str, Path],
    acceptance_sample_workbook: Path,
    acceptance_expected_manifest_shape: dict[str, Any],
    acceptance_expected_output_shape: dict[str, str],
) -> None:
    """
    Golden acceptance path:
    workbook -> compiled manifest -> tiny real execution -> workbook write-back.
    """

    manifest = compile_workbook_file(acceptance_sample_workbook)
    assert manifest.schema_version == acceptance_expected_manifest_shape["schema_version"]
    assert (
        manifest.compiled_manifest_schema_version
        == acceptance_expected_manifest_shape["compiled_manifest_schema_version"]
    )
    assert len(manifest.experiments) == 1
    assert (
        manifest.experiments[0].experiment_id == acceptance_expected_manifest_shape["experiment_id"]
    )
    assert len(manifest.trial_specs) == acceptance_expected_manifest_shape["trial_count"]
    trial = manifest.trial_specs[0]
    for key, value in acceptance_expected_manifest_shape["required_trial_params"].items():
        assert trial.params[key] == value

    run_result = run_workbook_decision_support_campaign(
        workbook_path=acceptance_sample_workbook,
        index_csv=acceptance_sample_data["index_csv"],
        data_root=acceptance_sample_data["data_root"],
        cache_dir=acceptance_sample_data["cache_dir"],
        output_root=acceptance_sample_data["output_root"],
        experiment_id=None,
        stage=None,
        run_all=True,
        seed=42,
        n_permutations=0,
        dry_run=False,
        write_back_to_workbook=True,
        workbook_output_dir=acceptance_sample_data["workbook_output_dir"],
        append_workbook_run_log=True,
    )

    workbook_output_path = Path(str(run_result["workbook_output_path"]))
    assert workbook_output_path.exists()
    assert workbook_output_path != acceptance_sample_workbook

    output_workbook = load_workbook(workbook_output_path, data_only=False)

    trial_ws = output_workbook["Trial_Results"]
    trial_cols = _header_map(trial_ws, header_row=2)
    trial_rows = _rows_with_value(
        trial_ws,
        header_map=trial_cols,
        key="experiment_id",
        expected="E16",
        start_row=3,
    )
    assert trial_rows
    latest_trial_row = trial_rows[-1]
    assert (
        trial_ws.cell(latest_trial_row, trial_cols["status"]).value
        == acceptance_expected_output_shape["trial_status"]
    )
    assert str(trial_ws.cell(latest_trial_row, trial_cols["metrics_path"]).value or "").strip()
    assert str(trial_ws.cell(latest_trial_row, trial_cols["report_path"]).value or "").strip()

    summary_ws = output_workbook["Summary_Outputs"]
    summary_cols = _header_map(summary_ws, header_row=2)
    summary_rows = _rows_with_value(
        summary_ws,
        header_map=summary_cols,
        key="summary_type",
        expected=acceptance_expected_output_shape["summary_type"],
        start_row=3,
    )
    assert summary_rows

    run_log_ws = output_workbook["Run_Log"]
    run_log_cols = _header_map(run_log_ws, header_row=1)
    run_log_rows = _rows_with_value(
        run_log_ws,
        header_map=run_log_cols,
        key="Experiment_ID",
        expected="E16",
        start_row=2,
    )
    assert run_log_rows
    latest_run_log_row = run_log_rows[-1]
    assert (
        run_log_ws.cell(latest_run_log_row, run_log_cols["Result_Summary"]).value
        == acceptance_expected_output_shape["run_log_result"]
    )
    assert str(
        run_log_ws.cell(latest_run_log_row, run_log_cols["Artifact_Path"]).value or ""
    ).strip()


def test_canonical_designed_study_end_to_end_acceptance(
    acceptance_sample_data: dict[str, Path],
    acceptance_canonical_designed_study_workbook: Path,
) -> None:
    study_id = "S_CANONICAL_2X2"
    workbook_summary = validate_workbook(acceptance_canonical_designed_study_workbook)
    assert workbook_summary["sheet_order_ok"] == "True"
    assert workbook_summary["workbook_schema_supported"] == "True"

    manifest = compile_workbook_file(acceptance_canonical_designed_study_workbook)
    study_trials = [trial for trial in manifest.trial_specs if trial.study_id == study_id]
    assert len(study_trials) == 4
    study_reviews = [review for review in manifest.study_reviews if review.study_id == study_id]
    assert len(study_reviews) == 1
    assert study_reviews[0].execution_disposition == "allowed"
    assert study_reviews[0].expected_design_cells == 4
    generated_cells = [cell for cell in manifest.generated_design_matrix if cell.study_id == study_id]
    assert len(generated_cells) == 4

    run_result = run_workbook_decision_support_campaign(
        workbook_path=acceptance_canonical_designed_study_workbook,
        index_csv=acceptance_sample_data["index_csv"],
        data_root=acceptance_sample_data["data_root"],
        cache_dir=acceptance_sample_data["cache_dir"],
        output_root=acceptance_sample_data["output_root"],
        experiment_id=None,
        stage=None,
        run_all=True,
        seed=42,
        n_permutations=0,
        dry_run=False,
        write_back_to_workbook=True,
        workbook_output_dir=acceptance_sample_data["workbook_output_dir"],
        append_workbook_run_log=True,
    )

    study_review_summary_path = Path(str(run_result["study_review_summary_path"]))
    assert study_review_summary_path.exists()
    study_review_summary = json.loads(study_review_summary_path.read_text(encoding="utf-8"))
    summary_studies = list(study_review_summary.get("studies", []))
    assert any(str(item.get("study_id")) == study_id for item in summary_studies)

    workbook_output_path = Path(str(run_result["workbook_output_path"]))
    output_workbook = load_workbook(workbook_output_path, data_only=False)

    generated_ws = output_workbook["Generated_Design_Matrix"]
    generated_cols = _header_map(generated_ws, header_row=2)
    generated_rows = _rows_with_value(
        generated_ws,
        header_map=generated_cols,
        key="study_id",
        expected=study_id,
        start_row=3,
    )
    assert len(generated_rows) == 4

    trial_ws = output_workbook["Trial_Results"]
    trial_cols = _header_map(trial_ws, header_row=2)
    trial_rows = _rows_with_value(
        trial_ws,
        header_map=trial_cols,
        key="study_id",
        expected=study_id,
        start_row=3,
    )
    assert len(trial_rows) == 4
    trial_statuses = {
        str(trial_ws.cell(row, trial_cols["status"]).value or "").strip() for row in trial_rows
    }
    assert trial_statuses <= {"completed", "failed", "dry_run"}
    assert "blocked" not in trial_statuses

    effect_ws = output_workbook["Effect_Summaries"]
    effect_cols = _header_map(effect_ws, header_row=2)
    effect_rows = _rows_with_value(
        effect_ws,
        header_map=effect_cols,
        key="study_id",
        expected=study_id,
        start_row=3,
    )
    assert effect_rows

    review_ws = output_workbook["Study_Review"]
    review_cols = _header_map(review_ws, header_row=2)
    review_rows = _rows_with_value(
        review_ws,
        header_map=review_cols,
        key="study_id",
        expected=study_id,
        start_row=3,
    )
    assert review_rows
    latest_review_row = review_rows[-1]
    assert review_ws.cell(latest_review_row, review_cols["execution_disposition"]).value == "allowed"
    assert int(review_ws.cell(latest_review_row, review_cols["error_count"]).value or 0) == 0
